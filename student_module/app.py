import os
import re
import json
import socket
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Response
from flask import abort
from flask_mail import Mail
from flask_bcrypt import Bcrypt
from dotenv import load_dotenv
from models import db, Student, Faculty, Timetable, LoginCredential, Parent, Guardian, Academic, OtherInfo, WorkExperience, Note, Course, Semester, Subject, SubjectAllocation, InternalMark, UniversityMark, UniversityResult, Attendance, MentoringSession, MentorLeave, LeaveRequest, Activity, DailyAttendance, StudentAttendance, Alert, WeeklyStudyPlan, StudyPlanSubject, StudySessionLog, MentorIntervention, InterventionOutcome, Batch, AlumniStudent, AlumniMentorHistory, Notification, MentorPrivateNote, StudentMark, SubjectHandlerMark, SubjectHandlerAttendance, SubjectAcademicEntry, SubjectDataAuditLog, SubjectAnalysisLog, Certificate, MentorMessage, SubjectHandlerMessage, PlaygroundNote, RemedialClass
from analytics.engine import run_full_analysis
from utils import send_otp_email, generate_otp, get_department_from_admission, normalize_dept_name
from datetime import datetime, date, timedelta
from functools import wraps
from sqlalchemy import inspect, text, func
from sqlalchemy.engine import make_url
from werkzeug.utils import secure_filename

from flask_cors import CORS
from sqlalchemy import event

def get_current_semester(start_year, duration_years):
    """
    Calculate the current semester based on start year and course duration.
    Assumes 2 semesters per year (Spring/Fall or odd/even).
    """
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # Calculate years since start
    years_since_start = current_year - start_year
    
    # Determine semester within the current academic year
    # Typically semester 1 runs Jan-June, semester 2 runs July-Dec
    if current_month <= 6:
        # First half of year (Jan-Jun) is semester 1 of the current academic year
        semester_in_year = 1
    else:
        # Second half of year (Jul-Dec) is semester 2 of the current academic year
        semester_in_year = 2
    
    # Calculate total semester number
    if years_since_start < 0:
        # Future start date, return 1 as default
        return 1
    elif years_since_start == 0:
        # First year, return the semester in the first year
        return semester_in_year
    else:
        # Calculate based on complete years plus current semester
        completed_semesters = years_since_start * 2
        current_semester = completed_semesters + semester_in_year
        
        # Ensure it doesn't exceed the total semesters for the course
        total_semesters = duration_years * 2
        return min(current_semester, total_semesters)


def is_imca_course_name(course_name):
    name = (course_name or '').strip().upper()
    return 'IMCA' in name or 'INTEGRATED MCA' in name


def is_computer_applications_course_name(course_name):
    name = (course_name or '').strip().upper()
    return (
        'COMPUTER APPLICATIONS' in name
        or name == 'MCA'
        or is_imca_course_name(name)
    )


def is_mba_course_name(course_name):
    name = (course_name or '').strip().upper()
    return (
        'BUSINESS ADMINISTRATION' in name
        or name == 'MBA'
        or 'BUSINESS' in name
    )


def get_normalized_course_duration(course_name, duration_years):
    raw = int(duration_years or 4)
    if is_mba_course_name(course_name):
        return 2
    if is_computer_applications_course_name(course_name):
        return 2
    return raw


def get_batch_end_year(start_year, duration_years, course_name=None):
    start_year = int(start_year)
    duration_years = get_normalized_course_duration(course_name, duration_years)
    return start_year + duration_years


def get_display_course_name(batch, course_name):
    if is_mba_course_name(course_name):
        return 'MBA'
    if is_computer_applications_course_name(course_name):
        has_imca_student = Student.query.filter(
            Student.batch_id == batch.id,
            Student.admission_number.ilike('%IMCA%')
        ).first() is not None
        if has_imca_student or (int(batch.end_year) - int(batch.start_year) >= 5):
            return 'IMCA'
        return 'MCA'
    return course_name


def _grade_point_from_value(mark_or_grade):
    if mark_or_grade is None or mark_or_grade == "":
        return None

    try:
        numeric = float(mark_or_grade)
    except (TypeError, ValueError):
        numeric = None

    if numeric is not None:
        if numeric >= 90:
            return 10.0
        if numeric >= 80:
            return 9.0
        if numeric >= 70:
            return 8.0
        if numeric >= 60:
            return 7.0
        if numeric >= 50:
            return 6.0
        if numeric >= 45:
            return 5.0
        if numeric >= 40:
            return 4.0
        return 0.0

    grade = str(mark_or_grade).strip().upper()
    mapping = {
        'S': 10.0,
        'O': 10.0,
        'A+': 9.0,
        'A': 8.5,
        'B+': 8.0,
        'B': 7.5,
        'C+': 7.0,
        'C': 6.5,
        'D': 6.0,
        'P': 5.5,
        'F': 0.0,
        'FE': 0.0,
    }
    return mapping.get(grade)


def _calculate_student_cgpa_metrics(student_id):
    marks = StudentMark.query.filter_by(student_id=student_id.upper()).all()
    semester_points = {}

    for row in marks:
        semester = row.semester
        if semester is None:
            continue

        score = _subject_combined_score(row)
        if score is None:
            fallback = row.university_mark if row.university_mark is not None else row.university_grade
            point = _grade_point_from_value(fallback)
        else:
            point = round(score / 10.0, 2)
        if point is None:
            continue

        semester_points.setdefault(int(semester), []).append(point)

    semester_sgpa = {
        semester: round(sum(points) / len(points), 2)
        for semester, points in semester_points.items()
        if points
    }

    cgpa = round(sum(semester_sgpa.values()) / len(semester_sgpa), 2) if semester_sgpa else None
    latest_semester = max(semester_sgpa) if semester_sgpa else None
    latest_sgpa = semester_sgpa.get(latest_semester) if latest_semester is not None else None

    return {
        "cgpa": cgpa,
        "sgpa": latest_sgpa,
        "semester_sgpa": semester_sgpa,
        "latest_semester": latest_semester,
    }


def _valid_metric(value):
    try:
        if value is None:
            return None
        numeric = float(value)
        if numeric < 0:
            return None
        return numeric
    except (TypeError, ValueError):
        return None


def _average_non_null(values):
    items = [float(value) for value in values if value is not None]
    return round(sum(items) / len(items), 2) if items else None


def _normalized_internal_score(row):
    avg_internal = _average_non_null([row.internal1, row.internal2, row.internal3])
    if avg_internal is None:
        return None
    # Seeded/demo internals are typically stored on a 0-50 scale, so normalize to 100.
    return round(avg_internal * 2, 1) if avg_internal <= 50 else round(avg_internal, 1)


def _subject_combined_score(row):
    university_score = _valid_metric(getattr(row, "university_mark", None))
    internal_score = _normalized_internal_score(row)

    if university_score is not None and internal_score is not None:
        return round((university_score * 0.7) + (internal_score * 0.3), 1)
    if university_score is not None:
        return round(university_score, 1)
    if internal_score is not None:
        return round(internal_score, 1)
    return None


def _classification_for_score(score):
    if score is None:
        return "unknown"
    if score < 40:
        return "critical"
    if score <= 60:
        return "weak"
    return "stable"


def _trend_from_values(values):
    cleaned = [float(value) for value in values if value is not None]
    if len(cleaned) < 2:
        return "stagnant"
    delta = cleaned[-1] - cleaned[0]
    if delta >= 5:
        return "improving"
    if delta <= -5:
        return "declining"
    return "stagnant"


def _build_today_plan(subject_items):
    default_slots = ["06:30-07:30", "18:00-19:00", "19:30-20:15"]
    plan = []
    for index, item in enumerate(subject_items[:3]):
        subject = item.get("subject") or "Priority Subject"
        classification = item.get("classification") or "weak"
        plan.append({
            "time": default_slots[index],
            "subject": subject,
            "type": "concept/practice",
            "tasks": [
                f"Concept learning: rebuild one weak {subject} concept from class notes.",
                f"Active recall: close the notes and write a 5-point summary for {subject}.",
                f"Practice/testing: solve 5 questions for {subject} and review every mistake.",
            ],
            "classification": classification,
        })
    return plan


def sync_passed_out_students_to_alumni():
    current_year = datetime.now().year
    touched = 0

    expired_batches = Batch.query.filter(
        Batch.end_year <= current_year - 1
    ).all()
    for batch in expired_batches:
        if batch.status != 'completed':
            batch.status = 'completed'
            touched += 1
        students = Student.query.filter_by(batch_id=batch.id).all()
        for student in students:
            if student.status != 'Passed Out':
                student.status = 'Passed Out'
                touched += 1
            if not student.passout_year:
                student.passout_year = batch.end_year
                touched += 1

    cleaned = 0
    active_students = Student.query.filter(Student.status != 'Passed Out').all()
    for student in active_students:
        existing = AlumniStudent.query.filter_by(admission_number=student.admission_number).first()
        if not existing:
            continue

        batch = Batch.query.get(student.batch_id) if student.batch_id else None
        if batch and batch.status == 'active':
            db.session.delete(existing)
            cleaned += 1

    synced = 0
    passed_out_students = Student.query.filter_by(status='Passed Out').all()
    for student in passed_out_students:
        existing = AlumniStudent.query.filter_by(admission_number=student.admission_number).first()
        batch = Batch.query.get(student.batch_id) if student.batch_id else None
        passout_year = student.passout_year or (batch.end_year if batch else None)
        if existing:
            existing.name = student.full_name
            existing.email = student.email
            existing.department = student.branch
            existing.course_id = batch.course_id if batch else existing.course_id
            existing.batch_id = batch.id if batch else existing.batch_id
            existing.mentor_id = student.mentor_id
            existing.passout_year = passout_year
            touched += 1
        else:
            alumni = AlumniStudent(
                admission_number=student.admission_number,
                name=student.full_name,
                email=student.email,
                department=student.branch,
                course_id=batch.course_id if batch else None,
                batch_id=batch.id if batch else None,
                mentor_id=student.mentor_id,
                passout_year=passout_year
            )
            db.session.add(alumni)
            synced += 1
        if passout_year and not student.passout_year:
            student.passout_year = passout_year

    if touched or cleaned or synced:
        db.session.commit()
    return {'cleaned': cleaned, 'synced': synced}

load_dotenv()

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
app.secret_key = os.getenv('SECRET_KEY', 'default_secret_key')

# Database Configuration
database_url = os.getenv("DATABASE_URL", "sqlite:///mentorai.db")
try:
    url = make_url(database_url)
    backend = url.get_backend_name()
    is_local_mysql = backend in {"mysql", "mariadb"} and (url.host or "").lower() in {"localhost", "127.0.0.1", "::1"}
    if is_local_mysql and not os.getenv("FORCE_MYSQL"):
        mysql_port = url.port or 3306
        try:
            sock = socket.create_connection((url.host, mysql_port), timeout=0.5)
            sock.close()
        except OSError:
            Path(app.instance_path).mkdir(parents=True, exist_ok=True)
            sqlite_path = (Path(app.instance_path) / "mentorai.db").resolve()
            database_url = f"sqlite:///{sqlite_path.as_posix()}"
except Exception:
    pass
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Mail Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')

from datetime import timedelta

mail = Mail(app)
bcrypt = Bcrypt(app)
db.init_app(app)

# Session config
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)
app.config['SESSION_COOKIE_SAMESITE']    = 'Lax'
app.config['MAX_CONTENT_LENGTH']         = 10 * 1024 * 1024  # 10MB upload limit

with app.app_context():
    if 'sqlite' in str(db.engine.url):
        with db.engine.connect() as conn:
            conn.execute(text("PRAGMA foreign_keys=ON"))
    db.create_all()
    # Migrations: Add missing schema manually to avoid operational errors
    try:
        inspector = inspect(db.engine)
        
        # Students table migration
        if 'students' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('students')]
            
            if 'password_hash' not in columns:
                print("Migrating: Adding password_hash to students table...")
                with db.engine.connect() as conn:
                    conn.execute(text("ALTER TABLE students ADD COLUMN password_hash VARCHAR(255)"))
                    conn.commit()
            
            if 'mentor_id' not in columns:
                print("Migrating: Adding mentor_id to students table...")
                with db.engine.connect() as conn:
                    if 'sqlite' in str(db.engine.url):
                        conn.execute(text("ALTER TABLE students ADD COLUMN mentor_id INTEGER REFERENCES faculty(id)"))
                    else:
                        conn.execute(text("ALTER TABLE students ADD COLUMN mentor_id INT"))
                        conn.execute(text("ALTER TABLE students ADD CONSTRAINT fk_student_mentor FOREIGN KEY (mentor_id) REFERENCES faculty(id)"))
                    conn.commit()
            
            if 'status' not in columns:
                print("Migrating: Adding status to students table...")
                with db.engine.connect() as conn:
                    conn.execute(text("ALTER TABLE students ADD COLUMN status VARCHAR(20) DEFAULT 'Live'"))
                    conn.commit()
                    
        # Faculty table migration
        if 'faculty' in inspector.get_table_names():
            f_cols = [c['name'] for c in inspector.get_columns('faculty')]
            with db.engine.connect() as conn:
                f_added = False
                if 'is_mentor_eligible' not in f_cols:
                    print("Migrating: Adding is_mentor_eligible to faculty table...")
                    if 'sqlite' in str(db.engine.url):
                        conn.execute(text("ALTER TABLE faculty ADD COLUMN is_mentor_eligible BOOLEAN DEFAULT 1"))
                    else:
                        conn.execute(text("ALTER TABLE faculty ADD COLUMN is_mentor_eligible BOOLEAN DEFAULT TRUE"))
                    f_added = True
                
                if 'email' not in f_cols:
                    print("Migrating: Adding email to faculty table...")
                    conn.execute(text("ALTER TABLE faculty ADD COLUMN email VARCHAR(100)"))
                    f_added = True
                    
                if 'is_hod' not in f_cols:
                    print("Migrating: Adding is_hod to faculty table...")
                    if 'sqlite' in str(db.engine.url):
                        conn.execute(text("ALTER TABLE faculty ADD COLUMN is_hod BOOLEAN DEFAULT 0"))
                    else:
                        conn.execute(text("ALTER TABLE faculty ADD COLUMN is_hod BOOLEAN DEFAULT FALSE"))
                    f_added = True
                
                if 'is_subject_handler' not in f_cols:
                    print("Migrating: Adding is_subject_handler to faculty table...")
                    if 'sqlite' in str(db.engine.url):
                        conn.execute(text("ALTER TABLE faculty ADD COLUMN is_subject_handler BOOLEAN DEFAULT 0"))
                    else:
                        conn.execute(text("ALTER TABLE faculty ADD COLUMN is_subject_handler BOOLEAN DEFAULT FALSE"))
                    f_added = True
                
                if f_added:
                    conn.commit()
                    
        # Mentor interventions migration
        if 'mentor_interventions' in inspector.get_table_names():
            mi_cols = [c['name'] for c in inspector.get_columns('mentor_interventions')]
            with db.engine.connect() as conn:
                mi_added = False
                if 'escalated' not in mi_cols:
                    if 'sqlite' in str(db.engine.url):
                        conn.execute(text("ALTER TABLE mentor_interventions ADD COLUMN escalated BOOLEAN DEFAULT 0"))
                    else:
                        conn.execute(text("ALTER TABLE mentor_interventions ADD COLUMN escalated BOOLEAN DEFAULT FALSE"))
                    conn.execute(text("ALTER TABLE mentor_interventions ADD COLUMN escalated_at DATETIME"))
                    mi_added = True
                    
                if mi_added:
                    conn.commit()
                    
        # Timetables table migration
        if 'timetables' in inspector.get_table_names():
            t_cols = [c['name'] for c in inspector.get_columns('timetables')]
            with db.engine.connect() as conn:
                added = False
                if 'file_path' not in t_cols:
                    print("Migrating: Adding file_path to timetables table...")
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN file_path VARCHAR(255)"))
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN course_id INTEGER"))
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN batch_id INTEGER"))
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN semester INTEGER"))
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN academic_year VARCHAR(20)"))
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN uploaded_at DATETIME"))
                    conn.execute(text("ALTER TABLE timetables ADD COLUMN uploaded_by INTEGER"))
                    added = True
                if added:
                    conn.commit()

        # Attendance table migration â€“ add subject_code if missing
        if 'attendance' in inspector.get_table_names():
            att_cols = [c['name'] for c in inspector.get_columns('attendance')]
            if 'subject_code' not in att_cols:
                with db.engine.connect() as conn:
                    print("Migrating: Adding subject_code to attendance table...")
                    conn.execute(text("ALTER TABLE attendance ADD COLUMN subject_code VARCHAR(50)"))
                    conn.commit()

        if 'certificates' in inspector.get_table_names():
            cert_cols = [c['name'] for c in inspector.get_columns('certificates')]
            with db.engine.connect() as conn:
                added = False
                if 'title' not in cert_cols:
                    conn.execute(text("ALTER TABLE certificates ADD COLUMN title VARCHAR(150)"))
                    added = True
                if 'issuing_org' not in cert_cols:
                    conn.execute(text("ALTER TABLE certificates ADD COLUMN issuing_org VARCHAR(150)"))
                    added = True
                if 'issue_date' not in cert_cols:
                    conn.execute(text("ALTER TABLE certificates ADD COLUMN issue_date DATE"))
                    added = True
                if 'expiry_date' not in cert_cols:
                    conn.execute(text("ALTER TABLE certificates ADD COLUMN expiry_date DATE"))
                    added = True
                if 'share_with_mentor' not in cert_cols:
                    print("Migrating: Adding share_with_mentor to certificates table...")
                    if 'sqlite' in str(db.engine.url):
                        conn.execute(text("ALTER TABLE certificates ADD COLUMN share_with_mentor BOOLEAN DEFAULT 0"))
                    else:
                        conn.execute(text("ALTER TABLE certificates ADD COLUMN share_with_mentor BOOLEAN DEFAULT FALSE"))
                    added = True
                if added:
                    conn.commit()

        if 'mentoring_sessions' in inspector.get_table_names():
            ms_cols = [c['name'] for c in inspector.get_columns('mentoring_sessions')]
            with db.engine.connect() as conn:
                added = False
                if 'attendance_marked_at' not in ms_cols:
                    conn.execute(text("ALTER TABLE mentoring_sessions ADD COLUMN attendance_marked_at DATETIME"))
                    added = True
                if 'absence_reason' not in ms_cols:
                    conn.execute(text("ALTER TABLE mentoring_sessions ADD COLUMN absence_reason TEXT"))
                    added = True
                if added:
                    conn.commit()

    except Exception as e:
        print(f"Migration check failed or skipped: {e}")

    # Auto-promote any expired batches on startup
    try:
        from services.batch_service import (
            promote_expired_batches_to_alumni,
            auto_allocate_unassigned_students,
            auto_heal_invalid_mentor_assignments,
        )
        result = promote_expired_batches_to_alumni()
        if result.get("expired_batches_found", 0) > 0:
            print(f"[STARTUP] Auto-promoted expired batches: {result}")

        healed_result = auto_heal_invalid_mentor_assignments()
        if healed_result.get("invalid_students_found", 0) > 0:
            print(f"[STARTUP] Auto-healed invalid mentor assignments: {healed_result}")

        allocation_result = auto_allocate_unassigned_students()
        if allocation_result.get("unassigned_students_found", 0) > 0:
            print(f"[STARTUP] Auto-allocated unassigned students: {allocation_result}")
    except Exception as e:
        print(f"[STARTUP ERROR] Batch promotion failed: {e}")

    # Seed Admin User
    if not Faculty.query.filter_by(username='admin').first():
        pw_hash = bcrypt.generate_password_hash('admin123').decode('utf-8')
        admin = Faculty(
            username='admin',
            password_hash=pw_hash,
            name='Administrator',
            designation='Admin',
            department='All',
            status='Live'
        )
        db.session.add(admin)
        db.session.commit()
        print("Admin user created: admin/admin123")

# ============= VALIDATION FUNCTIONS =============

def validate_name(name):
    """Validate name: only alphabets and spaces"""
    pattern = r'^[A-Za-z\s]+$'
    return bool(re.match(pattern, name))

def validate_email(email):
    """Validate email: basic pattern check"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))

def validate_mobile(mobile):
    """Validate mobile: exactly 10 digits"""
    pattern = r'^\d{10}$'
    return bool(re.match(pattern, mobile))

def validate_password(password):
    """Validate password: minimum 6 characters"""
    return len(password) >= 6

# ============= SESSION DECORATOR =============

def login_required(f):
    """Decorator to protect routes that require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ============= ROUTES =============

@app.route('/')
def index():
    return render_template('home.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            # Basic Info Only
            admission_number = request.form['admission_number'].strip().upper()
            full_name = request.form['full_name'].strip()
            email = request.form['email'].strip().lower()
            password = request.form['password']
            confirm_password = request.form.get('confirm_password', '')
            branch = request.form.get('branch', '').strip()
            
            # Validation
            if not admission_number or not full_name or not email or not password:
                flash('All required fields must be filled!', 'error')
                return redirect(url_for('register'))
            
            # Validate name
            if not validate_name(full_name):
                flash('Full Name must contain only alphabets and spaces!', 'error')
                return redirect(url_for('register'))
            
            # Validate password
            if not validate_password(password):
                flash('Password must be at least 6 characters long!', 'error')
                return redirect(url_for('register'))
            
            # Check password match
            if password != confirm_password:
                flash('Password and Confirm Password do not match!', 'error')
                return redirect(url_for('register'))
            
            # Validate email
            if not validate_email(email):
                flash('Only Gmail addresses (@gmail.com) are allowed!', 'error')
                return redirect(url_for('register'))
            
            # Check if student exists in bulk records
            student = Student.query.get(admission_number)
            if not student:
                flash('Registration Failed. Your admission number was not found in the system. Please contact the administrator.', 'error')
                return redirect(url_for('register'))
                
            # Check if already registered
            if LoginCredential.query.filter_by(admission_number=admission_number).first():
                flash('This student is already registered. Please log in instead if you have a password.', 'error')
                return redirect(url_for('register'))
                
            # Validate details (case-insensitive and trimmed)
            db_name = student.full_name.strip().lower()
            db_email = student.email.strip().lower() if student.email else ""
            
            if db_name != full_name.lower() or db_email != email.lower():
                flash('Registration Failed. The provided name or email does not match our records. Please enter the exact details submitted during admission.', 'error')
                return redirect(url_for('register'))

            # All checks passed, create Credentials with bcrypt
            hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
            new_cred = LoginCredential(admission_number=admission_number, password_hash=hashed_pw)
            
            # Save to database
            db.session.add(new_cred)
            db.session.commit()
            
            flash('Registration Successful. Your details have been verified with the institution records. Your account has been created successfully. You can now log in.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
            
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Prevent caching of login page
    if request.method == 'POST':
        admission_number = request.form['admission_number'].strip().upper()
        password = request.form['password']
        
        cred = LoginCredential.query.get(admission_number)
        
        if cred and bcrypt.check_password_hash(cred.password_hash, password):
            session.clear()  # Clear any existing session
            session['user_id'] = admission_number
            session.permanent = False  # Session expires when browser closes
            
            # Fetch student profile
            student = Student.query.get(admission_number)
            flash(f'Welcome back, {student.full_name}!', 'success')
            
            # Check if profile is complete
            if not student.profile_completed:
                flash('Please complete your profile to continue.', 'info')
                return redirect(url_for('complete_profile'))
            
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid Admission Number or Password!', 'error')
            
    response = render_template('login.html')
    # Prevent browser caching
    response = app.make_response(response)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/logout')
def logout():
    session.clear()  # Destroy all session data
    flash('You have been logged out successfully.', 'info')
    response = redirect(url_for('login'))
    # Prevent browser caching
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        student = Student.query.filter_by(email=email).first()
        
        if student:
            otp = generate_otp()
            session['reset_otp'] = otp
            session['reset_email'] = email
            if send_otp_email(mail, email, otp):
                flash('OTP sent to your email.', 'info')
                return redirect(url_for('verify_otp'))
            else:
                flash('Failed to send email. Check configuration.', 'error')
        else:
            flash('Email not found!', 'error')
            
    return render_template('forgot_password.html')

@app.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    if 'reset_email' not in session:
        return redirect(url_for('forgot_password'))
        
    if request.method == 'POST':
        entered_otp = request.form['otp']
        if entered_otp == session.get('reset_otp'):
            return redirect(url_for('reset_password'))
        else:
            flash('Invalid OTP!', 'error')
            
    return render_template('verify_otp.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if 'reset_email' not in session:
        return redirect(url_for('forgot_password'))
        
    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form.get('confirm_password', '')
        
        if not validate_password(new_password):
            flash('Password must be at least 6 characters long!', 'error')
            return redirect(url_for('reset_password'))
        
        if new_password != confirm_password:
            flash('Passwords do not match!', 'error')
            return redirect(url_for('reset_password'))
        
        email = session['reset_email']
        student = Student.query.filter_by(email=email).first()
        
        if student:
            cred = LoginCredential.query.get(student.admission_number)
            cred.password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
            db.session.commit()
            
            session.pop('reset_otp', None)
            session.pop('reset_email', None)
            
            flash('Password reset successful! Login now.', 'success')
            return redirect(url_for('login'))
            
    return render_template('reset_password.html')

@app.route('/dashboard')
@login_required
def dashboard():
    student = Student.query.get(session['user_id'])
    
    # Calculate profile completion percentage
    mandatory_fields = [
        student.full_name,
        student.admission_number,
        student.email,
        student.mobile_number,
        student.permanent_address,
        getattr(student.parents, 'father_name', None) if student.parents else None,
        getattr(student.parents, 'mother_name', None) if student.parents else None,
        getattr(student.academics, 'school_10th', None) if student.academics else None,
        getattr(student.academics, 'school_12th', None) if student.academics else None
    ]
    
    total_mandatory = len(mandatory_fields)
    filled_mandatory = sum(1 for field in mandatory_fields if field)
    completion_pct = int((filled_mandatory / total_mandatory) * 100) if total_mandatory > 0 else 0
    
    from academic_utils import calculate_academic_status
    acad = calculate_academic_status(student.batch, student.branch)
    current_sem = acad.get('current_semester', 1)
    
    response = render_template('dashboard.html', student=student, completion_pct=completion_pct, current_sem=current_sem)
    # Prevent browser caching
    response = app.make_response(response)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/complete_profile', methods=['GET', 'POST'])
@login_required
def complete_profile():
    student = Student.query.get(session['user_id'])
    
    if request.method == 'POST':
        try:
            # Personal Details
            student.roll_number = request.form.get('roll_number', '').strip()
            student.branch = request.form.get('branch', '').strip()
            
            # Normalize batch for MCA/IMCA
            raw_batch = request.form.get('batch', '').strip()
            if student.branch == 'Department of Computer Applications' or student.branch == 'MCA' or student.branch == 'IMCA':
                 # Normalize branch name just in case
                 student.branch = 'Department of Computer Applications'
                 
                 course_code = None
                 import re
                 match = re.search(r'(MCA|IMCA)', student.admission_number)
                 if match:
                    course_code = match.group(1)
                 
                 if course_code and raw_batch and not raw_batch.upper().startswith(course_code):
                      raw_batch = f"{course_code} {raw_batch}"
            
            student.batch = raw_batch
            
            dob_str = request.form.get('dob', '').strip()
            if dob_str:
                student.dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                
            student.age = int(request.form.get('age')) if request.form.get('age') else None
            student.blood_group = request.form.get('blood_group', '').strip()
            student.religion = request.form.get('religion', '').strip()
            student.caste_category = request.form.get('caste_category', '').strip()
            
            # Handle photo update
            if 'photo' in request.files:
                photo = request.files['photo']
                if photo and photo.filename:
                    import uuid
                    filename = f"{student.admission_number}_{uuid.uuid4().hex[:8]}{os.path.splitext(photo.filename)[1]}"
                    photo_dir = os.path.join(app.root_path, 'static', 'photos')
                    os.makedirs(photo_dir, exist_ok=True)
                    photo.save(os.path.join(photo_dir, filename))
                    student.photo_path = f"photos/{filename}"
            
            # Update Parents
            parent = Parent.query.filter_by(student_admission_number=student.admission_number).first()
            if not parent:
                parent = Parent(student_admission_number=student.admission_number)
            
            parent.father_name = request.form.get('father_name', '').strip()
            parent.father_profession = request.form.get('father_profession', '').strip()
            parent.father_age = int(request.form.get('father_age')) if request.form.get('father_age') else None
            parent.father_mobile = request.form.get('father_mobile', '').strip()
            parent.mother_name = request.form.get('mother_name', '').strip()
            parent.mother_profession = request.form.get('mother_profession', '').strip()
            parent.mother_age = int(request.form.get('mother_age')) if request.form.get('mother_age') else None
            parent.mother_mobile = request.form.get('mother_mobile', '').strip()
            
            # Update Academics
            acad = Academic.query.filter_by(student_admission_number=student.admission_number).first()
            if not acad:
                acad = Academic(student_admission_number=student.admission_number)
                
            acad.school_10th = request.form.get('school_10th', '').strip()
            acad.board_10th = request.form.get('board_10th', '').strip()
            acad.percentage_10th = float(request.form.get('percentage_10th')) if request.form.get('percentage_10th') else None
            
            acad.school_12th = request.form.get('school_12th', '').strip()
            acad.board_12th = request.form.get('board_12th', '').strip()
            acad.percentage_12th = float(request.form.get('percentage_12th')) if request.form.get('percentage_12th') else None
            
            acad.college_ug = request.form.get('college_ug', '').strip()
            acad.university_ug = request.form.get('university_ug', '').strip()
            acad.percentage_ug = float(request.form.get('percentage_ug')) if request.form.get('percentage_ug') else None
            
            acad.sgpa = float(request.form.get('sgpa')) if request.form.get('sgpa') else None
            acad.cgpa = float(request.form.get('cgpa')) if request.form.get('cgpa') else None
            
            # Update Other Info
            other = OtherInfo.query.filter_by(student_admission_number=student.admission_number).first()
            if not other:
                other = OtherInfo(student_admission_number=student.admission_number)
                
            other.siblings_details = request.form.get('siblings_details', '').strip()
            other.accommodation_type = request.form.get('accommodation_type', '').strip()
            other.transport_mode = request.form.get('transport_mode', '').strip()
            
            # Mark profile as completed
            student.profile_completed = True
            
            db.session.add(parent)
            db.session.add(acad)
            db.session.add(other)
            db.session.commit()
            
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')
            
    return render_template('complete_profile.html', student=student)


# ============= API ROUTES (for Node.js frontend) =============

@app.route('/api/register', methods=['POST'])
def api_register():
    """API endpoint for registration"""
    data = request.get_json()
    
    try:
        admission_number = data.get('admission_number', '').strip().upper()
        full_name = data.get('full_name', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        # Check if student exists in bulk records
        student = Student.query.get(admission_number)
        if not student:
            return jsonify({
                'success': False, 
                'message': 'Registration Failed. Your admission number was not found in the system. Please contact the administrator.'
            }), 404
            
        # Check if already registered
        if LoginCredential.query.filter_by(admission_number=admission_number).first():
            return jsonify({
                'success': False, 
                'message': 'This student is already registered. Please log in instead if you have a password.'
            }), 400
            
        # Validate details (case-insensitive and trimmed)
        db_name = student.full_name.strip().lower()
        db_email = student.email.strip().lower() if student.email else ""
        
        if db_name != full_name.lower() or db_email != email.lower():
            return jsonify({
                'success': False, 
                'message': 'Registration Failed. The provided name or email does not match our records. Please enter the exact details submitted during admission.'
            }), 400
            
        # All checks passed, create credentials
        hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
        new_cred = LoginCredential(admission_number=admission_number, password_hash=hashed_pw)
        
        # Optionally, mark as registered if needed implicitly
        # student.status = 'Registered' or profile_completed logic etc.
        
        db.session.add(new_cred)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Registration Successful. Your details have been verified with the institution records. Your account has been created successfully. You can now log in.',
            'data': {
                'admission_number': admission_number,
                'name': student.full_name,
                'department': student.branch
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/login', methods=['POST'])
def api_login():
    """Unified API endpoint for Student and Faculty login"""
    data = request.get_json() or {}

    # Accept 'admission_number' (from old student login) OR generic 'username/identifier'
    raw_username = data.get('username', '').strip()
    raw_adm = data.get('admission_number', '').strip().upper()
    identifier_upper = raw_adm or raw_username.upper()
    password = data.get('password', '')

    if not (raw_adm or raw_username) or not password:
        return jsonify({'success': False, 'message': 'Username and password required'}), 400

    # 1. Try Student Login (admission number = upper-cased)
    cred = LoginCredential.query.get(identifier_upper)
    if cred and cred.password_hash and bcrypt.check_password_hash(cred.password_hash, password):
        student = Student.query.get(identifier_upper)
        session['user_id'] = student.admission_number
        session['user_role'] = 'student'
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'data': {
                'role': 'student',
                'admission_number': student.admission_number,
                'name': student.full_name,
                'email': student.email,
                'department': student.branch,
                'batch': student.batch,
                'photo_path': student.photo_path,
                'profile_completed': student.profile_completed
            }
        }), 200

    # 2. Try Faculty/Admin Login
    # Try original-case username first (e.g. 'admin'), then uppercased
    user = Faculty.query.filter_by(username=raw_username).first()
    if not user and raw_adm:
        user = Faculty.query.filter_by(username=raw_adm).first()

    if user and user.password_hash and bcrypt.check_password_hash(user.password_hash, password):
        if user.status != 'Live':
            return jsonify({'success': False, 'message': 'Account is inactive'}), 403

        def _faculty_allowed_roles_api(faculty: Faculty) -> list[str]:
            if faculty.designation and faculty.designation.strip().lower() == 'admin':
                return ['admin']

            roles: list[str] = []
            if bool(faculty.is_hod):
                roles.append('hod')
            if bool(faculty.is_mentor_eligible) or bool(getattr(faculty, 'mentees', None)):
                roles.append('mentor')
            if bool(faculty.is_subject_handler) or _faculty_has_active_subjects_api(faculty.id):
                roles.append('subject-handler')

            if not roles:
                roles = ['mentor']

            seen: set[str] = set()
            ordered: list[str] = []
            for r in roles:
                if r not in seen:
                    ordered.append(r)
                    seen.add(r)
            return ordered

        def _faculty_default_role_api(faculty: Faculty) -> tuple[str, list[str]]:
            roles = _faculty_allowed_roles_api(faculty)
            if 'admin' in roles:
                return 'admin', roles
            if roles == ['hod']:
                return 'hod', roles
            if 'mentor' in roles and 'subject-handler' in roles:
                return ('subject-handler' if _faculty_has_active_subjects_api(faculty.id) else 'mentor'), roles
            return roles[0], roles

        active_role, allowed_roles = _faculty_default_role_api(user)
        session['user_id'] = user.id
        session['user_role'] = active_role
        session['user_roles'] = allowed_roles
        session['user_designation'] = user.designation
        session['user_dept'] = user.department
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'data': {
                'role': active_role,  # mentor / subject-handler / hod / admin
                'allowed_roles': allowed_roles,
                'id': user.id,
                'username': user.username,
                'name': user.name,
                'department': user.department,
                'designation': user.designation,
                'designation_role': (user.designation or '').strip().lower(),
                'is_subject_handler': bool(user.is_subject_handler),
                'is_hod': bool(user.is_hod),
                'is_mentor_eligible': bool(user.is_mentor_eligible)
            }
        }), 200

    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401


@app.route('/api/faculty/switch-role', methods=['POST'])
def api_faculty_switch_role():
    """Switch active role for a logged-in faculty between mentor and subject-handler (and hod if permitted)."""
    try:
        if not session.get('user_id'):
            return jsonify({'success': False, 'message': 'Unauthorized. Please login.'}), 401

        current_role = _current_role()
        if current_role == 'student':
            return jsonify({'success': False, 'message': 'Faculty access required'}), 403

        payload = request.get_json(silent=True) or {}
        requested = str(payload.get('role') or '').strip().lower().replace('_', '-')
        if requested not in ('mentor', 'subject-handler', 'hod', 'admin'):
            return jsonify({'success': False, 'message': 'Invalid role'}), 400

        faculty = Faculty.query.get(int(session.get('user_id')))
        if not faculty:
            session.clear()
            return jsonify({'success': False, 'message': 'Session user not found'}), 401

        def _faculty_allowed_roles_api(f: Faculty) -> list[str]:
            if f.designation and f.designation.strip().lower() == 'admin':
                return ['admin']
            roles: list[str] = []
            if bool(f.is_hod):
                roles.append('hod')
            if bool(f.is_mentor_eligible) or bool(getattr(f, 'mentees', None)):
                roles.append('mentor')
            if bool(f.is_subject_handler) or _faculty_has_active_subjects_api(f.id):
                roles.append('subject-handler')
            if not roles:
                roles = ['mentor']
            seen: set[str] = set()
            ordered: list[str] = []
            for r in roles:
                if r not in seen:
                    ordered.append(r)
                    seen.add(r)
            return ordered

        allowed = _faculty_allowed_roles_api(faculty)
        if requested not in allowed:
            return jsonify({'success': False, 'message': 'Forbidden: role not permitted'}), 403

        session['user_role'] = requested
        session['user_roles'] = allowed

        return jsonify({'success': True, 'data': {'role': requested, 'allowed_roles': allowed}}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/profile/<admission_number>', methods=['GET'])
def api_get_profile(admission_number):
    try:
        student = Student.query.get(admission_number)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        # Serialize Basic Info
        data = {
            'admission_number': student.admission_number,
            'full_name': student.full_name,
            'email': student.email,
            'department': student.branch,
            'photo_path': student.photo_path,
            'roll_number': student.roll_number,
            'dob': student.dob.isoformat() if student.dob else None,
            'mobile_number': student.mobile_number,
            'blood_group': student.blood_group,
            'religion': student.religion,
            'caste': student.caste_category, # Model has caste_category
            #'category': student.category, # Model has caste_category used above? Or separate? 
            # Model: caste_category. API used 'caste' and 'category'. I'll map caste_category to caste.
            #'community': student.community, # Not in model
            #'is_catholic': student.is_catholic, # Not in model, derived?
            'parish_name': student.parish, # Model: parish
            'diocese_name': student.diocese, # Model: diocese
            'permanent_address': student.permanent_address,
            'temporary_address': student.contact_address, # Model: contact_address
            'profile_completed': student.profile_completed,
            
            'parents': {},
            'guardian': {},
            'accommodation': {},
            'academics': {},
            'work_experience': [],
            'other_info': {}
        }
        
        # Serialize Nested Relationships
        if student.parents:
            p = student.parents
            data['parents'] = {
                'father_name': p.father_name,
                'father_occupation': p.father_profession, # Map profession -> occupation
                'father_mobile': p.father_mobile,
                'mother_name': p.mother_name,
                'mother_occupation': p.mother_profession, # Map profession -> occupation
                'mother_mobile': p.mother_mobile,
            }
            
        if student.guardian:
            g = student.guardian
            data['guardian'] = {
                'name': g.name,
                'mobile': g.mobile_number,
                'address': g.address
            }
            
        # Accommodation is in OtherInfo in schema
        if student.other_info:
            o = student.other_info
            
            # Accommodation Data
            data['accommodation'] = {
                'type': o.accommodation_type,
                #'room_number': o.room_number, # Not in model
                'stay_from': o.stay_from.isoformat() if o.stay_from else None,
                'stay_to': o.stay_to.isoformat() if o.stay_to else None,
                'hostel_name': o.hostel_name,
                'transport_mode': o.transport_mode,
                'vehicle_number': o.vehicle_number,
                'staying_with': o.staying_with
            }
            
            # Other Info Data
            data['other_info'] = {
                'siblings_details': o.siblings_details,
                #'achievements': o.achievements, # Not in OtherInfo model
                #'hobbies': o.hobbies # Not in OtherInfo model
            }
            
        if student.academics:
            a = student.academics
            data['academics'] = {
                'school_10th': a.school_10th,
                'board_10th': a.board_10th,
                'percentage_10th': a.percentage_10th,
                #'medium_10th': a.medium_of_instruction, # Use shared field?
                'school_12th': a.school_12th,
                'board_12th': a.board_12th,
                'percentage_12th': a.percentage_12th,
                #'medium_12th': a.medium_of_instruction, # Use shared field?
                'college_ug': a.college_ug,
                'university_ug': a.university_ug,
                'percentage_ug': a.percentage_ug,
                'entrance_rank': a.entrance_rank,
                'nature_of_admission': a.nature_of_admission
            }
        
        # Work Experience
        if student.work_experience:
            data['work_experience'] = [{
                'organization': w.organization,
                'job_title': w.job_title,
                'duration': w.duration
            } for w in student.work_experience]
        
        return jsonify({'success': True, 'data': data}), 200
        
    except Exception as e:
        print(f"Error fetching profile: {e}")
        return jsonify({'success': False, 'message': 'Internal Server Error'}), 500

@app.route('/api/complete_profile', methods=['POST'])
def api_complete_profile():
    try:
        data = {}
        # Handle multipart/form-data (for file upload) or JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            import json
            if 'data' in request.form:
                data = json.loads(request.form['data'])
        else:
            data = request.get_json() or {}

        admission_number = data.get('admission_number')
        
        if not admission_number:
            return jsonify({'success': False, 'message': 'Admission number required'}), 400
            
        student = Student.query.get(admission_number)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Handle Photo Upload
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and photo.filename:
                import uuid
                filename = f"{student.admission_number}_{uuid.uuid4().hex[:8]}{os.path.splitext(photo.filename)[1]}"
                photo_dir = os.path.join(app.root_path, 'static', 'photos')
                os.makedirs(photo_dir, exist_ok=True)
                photo.save(os.path.join(photo_dir, filename))
                student.photo_path = f"photos/{filename}"

        # Personal Details
        if data.get('full_name'):
            student.full_name = data.get('full_name')
        if data.get('email'):
            student.email = data.get('email')
        
        student.roll_number = data.get('roll_number')
        if data.get('dob'):
             student.dob = datetime.strptime(data.get('dob'), '%Y-%m-%d').date()
        student.age = data.get('age')
        student.blood_group = data.get('blood_group')
        student.religion = data.get('religion')
        student.diocese = data.get('diocese')
        student.parish = data.get('parish')
        student.caste_category = data.get('caste_category')
        student.permanent_address = data.get('permanent_address')
        student.contact_address = data.get('contact_address')
        student.mobile_number = data.get('mobile_number')
        
        # Parent Details
        parent = Parent.query.filter_by(student_admission_number=admission_number).first()
        if not parent:
            parent = Parent(student_admission_number=admission_number)
            
        parent_data = data.get('parents', {})
        parent.father_name = parent_data.get('father_name')
        parent.father_profession = parent_data.get('father_profession')
        parent.father_place_of_work = parent_data.get('father_place_of_work')
        parent.father_mobile = parent_data.get('father_mobile')
        parent.mother_name = parent_data.get('mother_name')
        parent.mother_profession = parent_data.get('mother_profession')
        parent.mother_place_of_work = parent_data.get('mother_place_of_work')
        parent.mother_mobile = parent_data.get('mother_mobile')
        
        db.session.add(parent)
        
        # Guardian Details
        guardian_data = data.get('guardian')
        if guardian_data:
            guardian = Guardian.query.filter_by(student_admission_number=admission_number).first()
            if not guardian:
                guardian = Guardian(student_admission_number=admission_number)
            guardian.name = guardian_data.get('name')
            guardian.address = guardian_data.get('address')
            guardian.mobile_number = guardian_data.get('mobile_number')
            db.session.add(guardian)

        # Academic Details
        acad = Academic.query.filter_by(student_admission_number=admission_number).first()
        if not acad:
            acad = Academic(student_admission_number=admission_number)
            
        acad_data = data.get('academics', {})
        acad.school_10th = acad_data.get('school_10th')
        acad.board_10th = acad_data.get('board_10th')
        acad.percentage_10th = acad_data.get('percentage_10th')
        acad.school_12th = acad_data.get('school_12th')
        acad.board_12th = acad_data.get('board_12th')
        acad.percentage_12th = acad_data.get('percentage_12th')
        acad.entrance_rank = acad_data.get('entrance_rank')
        acad.nature_of_admission = acad_data.get('nature_of_admission')
        acad.medium_of_instruction = acad_data.get('medium_of_instruction')
        
        if acad_data.get('college_ug'):
             acad.college_ug = acad_data.get('college_ug')
             acad.university_ug = acad_data.get('university_ug')
             acad.percentage_ug = acad_data.get('percentage_ug')
        
        db.session.add(acad)
        
        # Work Experience
        WorkExperience.query.filter_by(student_admission_number=admission_number).delete()
        for work in data.get('work_experience', []):
            new_work = WorkExperience(
                student_admission_number=admission_number,
                organization=work.get('organization'),
                job_title=work.get('job_title'),
                duration=work.get('duration')
            )
            db.session.add(new_work)

        # Other Info
        other = OtherInfo.query.filter_by(student_admission_number=admission_number).first()
        if not other:
            other = OtherInfo(student_admission_number=admission_number)
            
        other_data = data.get('other_info', {})
        other.accommodation_type = other_data.get('accommodation_type')
        other.staying_with = other_data.get('staying_with')
        other.hostel_name = other_data.get('hostel_name')
        if other_data.get('stay_from'):
             other.stay_from = datetime.strptime(other_data.get('stay_from'), '%Y-%m-%d').date()
        if other_data.get('stay_to'):
             other.stay_to = datetime.strptime(other_data.get('stay_to'), '%Y-%m-%d').date()
             
        other.transport_mode = other_data.get('transport_mode')
        other.vehicle_number = other_data.get('vehicle_number')
        
        db.session.add(other)

        student.profile_completed = True
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Profile updated successfully'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= ADMIN & FACULTY ROUTES =============

@app.route('/api/admin/login', methods=['POST'])
def api_admin_login():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({'success': False, 'message': 'Username and password required'}), 400

    user = Faculty.query.filter_by(username=username).first()
    
    if user and bcrypt.check_password_hash(user.password_hash, password):
        if user.status != 'Live':
             return jsonify({'success': False, 'message': 'Account is inactive'}), 403
             
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'data': {
                'id': user.id,
                'username': user.username,
                'name': user.name,
                'role': user.designation, # Admin, HOD, Mentor, Subject Handler
                'department': user.department,
                'designation': user.designation,
                'is_subject_handler': bool(user.is_subject_handler),
                'is_hod': bool(user.is_hod),
                'is_mentor_eligible': bool(user.is_mentor_eligible)
            }
        }), 200
    
    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

@app.route('/api/admin/teachers', methods=['GET'])
def api_get_teachers():
    try:
        # Fetch all faculty, ordered by Department then Designation
        teachers = Faculty.query.filter(Faculty.username != 'admin').all()
        
        data = []
        for t in teachers:
            data.append({
                'id': t.id,
                'name': t.name,
                'username': t.username,
                'designation': t.designation,
                'department': t.department,
                'status': t.status
            })
            
        return jsonify({'success': True, 'data': data}), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/faculty/<int:faculty_id>/status', methods=['PUT'])
def api_update_faculty_status(faculty_id):
    try:
        data = request.get_json()
        new_status = data.get('status')
        if not new_status:
             return jsonify({'success': False, 'message': 'Status is required'}), 400
             
        faculty = Faculty.query.get(faculty_id)
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404

        from services.batch_service import is_mentor_eligible, rebalance_department_batches

        old_status = faculty.status
        affected_batch_ids = sorted({
            s.batch_id for s in Student.query.filter_by(mentor_id=faculty.id, status='Live').all()
            if s.batch_id is not None
        })

        faculty.status = new_status
        faculty.is_mentor_eligible = is_mentor_eligible(faculty)
        db.session.flush()

        rebalance_result = None
        if old_status != new_status:
            rebalance_result = rebalance_department_batches(faculty.department, affected_batch_ids)
        db.session.commit()

        message = f'Status updated to {new_status}'
        if rebalance_result and rebalance_result.get('processed_batches'):
            message += f". Rebalanced {rebalance_result['total_students']} students across {len(rebalance_result['processed_batches'])} batch(es)"
        elif rebalance_result and rebalance_result.get('errors'):
            message += f". Rebalance warning: {'; '.join(rebalance_result['errors'])}"

        return jsonify({
            'success': True,
            'message': message,
            'rebalanced_batches': rebalance_result.get('processed_batches', []) if rebalance_result else []
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/faculty/<int:faculty_id>', methods=['DELETE'])
def api_delete_faculty(faculty_id):
    try:
        faculty = Faculty.query.get(faculty_id)
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404

        from services.batch_service import rebalance_department_batches

        affected_batch_ids = sorted({
            s.batch_id for s in Student.query.filter_by(mentor_id=faculty.id, status='Live').all()
            if s.batch_id is not None
        })

        faculty.status = 'Inactive'
        faculty.is_mentor_eligible = False
        db.session.flush()

        rebalance_result = rebalance_department_batches(faculty.department, affected_batch_ids)
        db.session.delete(faculty)
        db.session.commit()

        message = 'Faculty deleted successfully'
        if rebalance_result.get('processed_batches'):
            message += f". Rebalanced {rebalance_result['total_students']} students across {len(rebalance_result['processed_batches'])} batch(es)"
        elif rebalance_result.get('errors'):
            message += f". Rebalance warning: {'; '.join(rebalance_result['errors'])}"

        return jsonify({
            'success': True,
            'message': message,
            'rebalanced_batches': rebalance_result.get('processed_batches', [])
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= STUDENT ACCESS ROUTES =============

@app.route('/api/admin/students', methods=['GET'])
def api_get_students_list():
    try:
        from academic_utils import calculate_academic_status
        students = Student.query.all()
        data = []
        changed = False
        for s in students:
            acad = calculate_academic_status(s.batch, s.branch)
            new_status = 'Passed Out' if acad['student_status'] == 'alumni' else 'Live'
            
            if s.status != new_status and s.status not in ['Hold', 'Dropout']:
                s.status = new_status
                changed = True

            data.append({
                'admission_number': s.admission_number,
                'name': s.full_name,
                'department': s.branch,
                'batch': s.batch,
                'status': s.status,
                'mentor_id': s.mentor_id,
                'academic_info': acad
            })
            
        if changed:
            db.session.commit()
            
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/student/<string:admission_number>/status', methods=['PUT'])
def api_update_student_status(admission_number):
    try:
        data = request.get_json()
        new_status = data.get('status')
        if not new_status:
            return jsonify({'success': False, 'message': 'Status is required'}), 400
            
        student = Student.query.get(admission_number)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
            
        student.status = new_status
        db.session.commit()
        return jsonify({'success': True, 'message': f'Status updated to {new_status}'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= TIMETABLE ROUTES =============

@app.route('/api/admin/timetable/upload', methods=['POST'])
def api_upload_timetable():
    try:
        department = request.form.get('department')
        batch = request.form.get('batch')
        
        if 'file' not in request.files:
             return jsonify({'success': False, 'message': 'File is required'}), 400
             
        file = request.files['file']
        
        if not department or not batch or file.filename == '':
            return jsonify({'success': False, 'message': 'Department, Batch and File are required'}), 400
            
        # Delete existing timetable for this dept/batch to replace it
        if department and batch:
            Timetable.query.filter_by(department=department, batch=batch).delete()
            db.session.commit() # Commit deletion first
        
        import pandas as pd
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
            
        # Expect columns: Day, Period, Subject
        # Optional: Handler/Faculty, Time
        df.columns = [c.lower().strip() for c in df.columns]

        pending_faculty_slots = set()
        count = 0
        for _, row in df.iterrows():
            day = row.get('day')
            period = row.get('period')
            subject = row.get('subject')
            
            if pd.isna(day) or pd.isna(period) or pd.isna(subject): continue
            
            handler_name = row.get('handler') or row.get('handler_name') or row.get('faculty')
            if pd.isna(handler_name): handler_name = None
            
            time_slot = row.get('time') or row.get('time_slot')
            if pd.isna(time_slot): time_slot = None

            # Try to link handler to Faculty (simple name match)
            handler_id = None
            if handler_name:
                faculty = Faculty.query.filter(Faculty.name.ilike(f"%{handler_name}%")).first()
                if faculty:
                    handler_id = faculty.id

            period_value = int(period) if str(period).isdigit() else 0
            day_value = str(day).strip()

            if handler_id:
                slot_key = (handler_id, day_value.lower(), period_value)
                if slot_key in pending_faculty_slots:
                    raise ValueError(
                        f"Faculty '{handler_name}' is assigned multiple classes in the uploaded file for {day_value} period {period_value}"
                    )

                conflict = Timetable.query.filter(
                    Timetable.handler_id == handler_id,
                    Timetable.day.ilike(day_value),
                    Timetable.period == period_value,
                ).first()
                if conflict:
                    raise ValueError(
                        f"Faculty '{handler_name}' already has a class for {day_value} period {period_value} "
                        f"({conflict.department} {conflict.batch})"
                    )
                pending_faculty_slots.add(slot_key)
            
            entry = Timetable(
                department=department,
                batch=batch,
                day=day_value,
                period=period_value,
                time_slot=str(time_slot) if time_slot else None,
                subject=str(subject),
                handler_name=str(handler_name) if handler_name else None,
                handler_id=handler_id
            )
            db.session.add(entry)
            
            # --- START PHASE 1 SUBJECT ALLOCATION LOGIC ---
            course = Course.query.filter_by(name=department).first()
            if not course:
                course = Course(name=department)
                db.session.add(course)
                db.session.flush()

            semester = Semester.query.filter_by(name=batch).first()
            if not semester:
                semester = Semester(name=batch)
                db.session.add(semester)
                db.session.flush()

            subject_obj = Subject.query.filter_by(name=str(subject), semester_id=semester.id).first()
            if not subject_obj:
                subject_obj = Subject(name=str(subject), course_id=course.id, semester_id=semester.id)
                db.session.add(subject_obj)
                db.session.flush()

            if handler_id:
                allocation = SubjectAllocation.query.filter_by(subject_id=subject_obj.id, faculty_id=handler_id).first()
                if not allocation:
                    allocation = SubjectAllocation(subject_id=subject_obj.id, faculty_id=handler_id)
                    db.session.add(allocation)
            # --- END PHASE 1 LOGIC ---
            
            count += 1
            
        db.session.commit()
        return jsonify({'success': True, 'message': f'Uploaded {count} timetable entries for {department} - {batch}'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= MARKS & EXAM ROUTES (PHASE 1) =============

def faculty_teaches_subject(faculty_id, subject_id):
    return SubjectAllocation.query.filter_by(faculty_id=faculty_id, subject_id=subject_id).first() is not None

def student_in_subject(student_id, subject_id):
    subject = Subject.query.get(subject_id)
    student = Student.query.get(student_id)
    if not subject or not student: return False
    course = Course.query.get(subject.course_id)
    semester = Semester.query.get(subject.semester_id)
    # Using branch -> course.name and batch -> semester.name matching logic 
    return student.branch == course.name and student.batch == semester.name

@app.route("/api/marks/upload-internal", methods=["POST"])
def api_upload_internal():
    try:
        data = request.json
        student_id = data.get("student_id")
        subject_id = data.get("subject_id")
        exam_type = data.get("exam_type")
        marks = data.get("marks")
        faculty_id = data.get("faculty_id")

        if not all([student_id, subject_id, exam_type, marks is not None, faculty_id]):
            return jsonify({'success': False, 'message': 'Missing data'}), 400

        if not faculty_teaches_subject(faculty_id, subject_id):
            return jsonify({'success': False, 'message': 'Forbidden: You do not teach this subject'}), 403

        if not student_in_subject(student_id, subject_id):
            return jsonify({'success': False, 'message': 'Forbidden: Student does not belong to this subject course/semester'}), 403

        if not (0 <= float(marks) <= 100):
            return jsonify({'success': False, 'message': 'Marks must be between 0 and 100'}), 400

        mark_record = InternalMark.query.filter_by(
            student_id=student_id, subject_id=subject_id, exam_type=exam_type
        ).first()

        if mark_record:
            mark_record.marks = float(marks)
            mark_record.uploaded_by = faculty_id
        else:
            mark_record = InternalMark(
                student_id=student_id, subject_id=subject_id,
                exam_type=exam_type, marks=float(marks), uploaded_by=faculty_id
            )
            db.session.add(mark_record)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Internal marks uploaded successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route("/api/marks/upload-university", methods=["POST"])
def api_upload_university():
    try:
        student_id = request.form.get("student_id")
        semester_id = request.form.get("semester_id")
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'PDF file is required'}), 400
        
        file = request.files['file']
        if not file.filename.endswith('.pdf'):
            return jsonify({'success': False, 'message': 'Only PDF files allowed'}), 400
            
        if not student_id or not semester_id:
            return jsonify({'success': False, 'message': 'student_id and semester_id required'}), 400
        
        import os
        os.makedirs("uploads", exist_ok=True)
        pdf_path = f"uploads/univ_marks_{student_id}_{semester_id}.pdf"
        file.save(pdf_path)

        record = UniversityMark.query.filter_by(student_id=student_id, semester_id=semester_id).first()
        if record:
            record.pdf_path = pdf_path
            record.verified = False
        else:
            record = UniversityMark(student_id=student_id, semester_id=semester_id, pdf_path=pdf_path)
            db.session.add(record)

        db.session.commit()
        return jsonify({'success': True, 'message': 'University marks uploaded successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

def _extract_batch_years(batch_label):
    import re
    match = re.search(r'(\d{4})\s*-\s*(\d{4})', str(batch_label or ''))
    return (int(match.group(1)), int(match.group(2))) if match else None

def _normalize_batch_label(batch_label):
    import re
    label = str(batch_label or '').strip().lower()
    label = re.sub(r'\s+', ' ', label)
    label = label.replace('_', ' ').replace('-', '-')
    return label

def _filter_timetable_by_batch(rows, requested_batch):
    """Prefer exact batch labels; fall back to same year range only when needed."""
    requested_norm = _normalize_batch_label(requested_batch)
    requested_years = _extract_batch_years(requested_batch)

    exact = [
        row for row in rows
        if _normalize_batch_label(row.batch) == requested_norm
        or _normalize_batch_label(row.batch).endswith(f" {requested_norm}")
        or requested_norm.endswith(f" {_normalize_batch_label(row.batch)}")
    ]
    if exact:
        return exact

    if requested_years:
        return [row for row in rows if _extract_batch_years(row.batch) == requested_years]

    return []

@app.route('/api/timetable/view', methods=['GET'])
def api_view_timetable():
    try:
        department = request.args.get('department')
        batch = request.args.get('batch')
        faculty_id = request.args.get('faculty_id')
        
        if department and batch:
            dept_rows = Timetable.query.filter(Timetable.department == department).all()
            entries = _filter_timetable_by_batch(dept_rows, batch)
        elif faculty_id:
            entries = Timetable.query.filter_by(handler_id=faculty_id).order_by(Timetable.day, Timetable.period).all()
        else:
             return jsonify({'success': False, 'message': 'Missing filters'}), 400

        entries = sorted(entries, key=lambda row: (row.day or '', row.period or 0))
        
        data = []
        for t in entries:
            data.append({
                'day': t.day,
                'period': t.period,
                'time_slot': t.time_slot,
                'subject': t.subject,
                'handler': t.handler_name,
                'department': t.department,
                'batch': t.batch
            })
            
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

ALLOWED_DEPARTMENTS = [
    'Department of Computer Applications', # Canonical for MCA/IMCA
    'Computer Science and Engineering (CSE)',
    'Mechanical Engineering (ME)',
    'Civil Engineering (CE)',
    'Electrical and Electronics Engineering (EEE)',
    'Electronics and Communication Engineering (ECE)',
    'Electronics and Computer Engineering (ECM)',
    'Department of Business Administration',
    'Basic Sciences & Humanities',
    # Legacy / Alternate codes kept for matching potential raw inputs before normalization (though util handles most)
    'CSE', 'MCA', 'IMCA', 'MBA', 'Civil', 'Mechanical', 'Electrical', 'Electronics',
    'CS', 'AI', 'AD', 'CY', 'CSE-AI', 'CSE-DS', 'CSE-CY', 'CSE-AD'
]

@app.route('/api/admin/upload_teachers', methods=['POST'])
def api_upload_teachers():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
            
        if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
             return jsonify({'success': False, 'message': 'Invalid file format. Upload CSV or Excel.'}), 400

        import pandas as pd
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
            
        # Standardize columns: username, password, name, designation, department
        required_cols = ['username', 'password', 'name', 'designation', 'department']
        # Check if columns exist (case insensitive)
        df.columns = [c.lower().strip() for c in df.columns]
        
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
             return jsonify({'success': False, 'message': f'Missing columns: {", ".join(missing)}'}), 400
             
        success_count = 0
        errors = []
        
        for index, row in df.iterrows():
            dept = row['department']
            # Validation: Department
            # Map department codes to full names (canonical)
            # Use unified normalization from utils for best results
            normalized_dept = normalize_dept_name(dept)
            
            if not normalized_dept:
                errors.append(f"Row {index+1}: Invalid Department '{dept}'")
                continue
            
            # Additional validation against allowed list if needed
            found = False
            for d in ALLOWED_DEPARTMENTS:
                 if d.lower() == normalized_dept.lower():
                     found = True
                     break
            if not found:
                 # If normalize_dept_name returned something not in ALLOWED, warn or allow?
                 # Assuming normalize fn helps map to allowed.
                 pass

                
            username = str(row['username']).strip()
            if Faculty.query.filter_by(username=username).first():
                errors.append(f"Row {index+1}: Username '{username}' already exists")
                continue
                
            pw_hash = bcrypt.generate_password_hash(str(row['password'])).decode('utf-8')
            
            new_faculty = Faculty(
                username=username,
                password_hash=pw_hash,
                name=row['name'],
                designation=row['designation'],
                department=normalized_dept,
                status='Live'
            )
            db.session.add(new_faculty)
            success_count += 1
            
        db.session.commit()
        
        msg = f"Successfully created {success_count} users."
        if errors:
            msg += f" {len(errors)} errors: " + "; ".join(errors[:5])
            if len(errors) > 5: msg += "..."
            
        return jsonify({'success': True, 'message': msg, 'errors': errors}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/upload_students', methods=['POST'])
def api_upload_students():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400

        import pandas as pd
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
            
        # Expected: admission_number, name, roll_number, department, batch, email
        df.columns = [c.lower().strip() for c in df.columns]
        
        if 'batch' not in df.columns:
            return jsonify({'success': False, 'message': 'Batch column is required'}), 400
        
        processed_students = []
        affected_batches = set()
        success_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                with db.session.begin_nested():
                    adm_no = str(row['admission_number']).strip().upper()
                    
                    # Validate required fields
                    if not all([row.get('name'), row.get('email')]):
                        errors.append(f"Row {idx+1}: Missing required fields")
                        continue
                    
                    # Check email uniqueness proactively
                    student_email = str(row.get('email')).strip()
                    existing_email = Student.query.filter_by(email=student_email).first()
                    if existing_email and existing_email.admission_number != adm_no:
                        # Auto-fix email for sample data collisions to allow graceful upload
                        parts = student_email.split('@')
                        if len(parts) == 2:
                            student_email = f"{parts[0]}_{adm_no.lower()}@{parts[1]}"
                        else:
                            student_email = f"{student_email}_{adm_no.lower()}@mentai.edu"
                        
                        # Verify the auto-fixed email isn't somehow also taken
                        if Student.query.filter_by(email=student_email).first():
                             student_email = f"student_{adm_no.lower()}@mentai.edu"
                    
                    # Find or create department (course)
                    dept_input = str(row.get('department', '')).strip()
                    
                    # Map department codes to full names using utils
                    found_dept = normalize_dept_name(dept_input) if dept_input else None
                    
                    if not found_dept:
                        errors.append(f"Row {idx+1}: Department not found or invalid '{dept_input}' for {row.get('name')}")
                        continue
                    
                    # Find course by name
                    course = Course.query.filter_by(name=found_dept).first()
                    if not course:
                        # Create course with default duration
                        course = Course(name=found_dept, duration_years=4)
                        db.session.add(course)
                        db.session.flush()  # Get ID without committing
                    
                    # Parse batch to get start_year
                    batch_str = str(row['batch']).strip()
                    
                    # Extract years from batch string (e.g., "2024-2026")
                    import re
                    year_match = re.search(r'(\d{4})-(\d{4})', batch_str)
                    if not year_match:
                        errors.append(f"Row {idx+1}: Invalid batch format for {row.get('name')}. Expected format: YYYY-YYYY")
                        continue
                    
                    start_year = int(year_match.group(1))
                    expected_end_year = int(year_match.group(2))
                    
                    # Find or create batch
                    batch = Batch.query.filter_by(
                        course_id=course.id,
                        start_year=start_year
                    ).first()
                    
                    if not batch:
                        # Create batch with the provided end year from the uploaded data
                        batch = Batch(
                            course_id=course.id,
                            start_year=start_year,
                            end_year=expected_end_year,
                            status='active'
                        )
                        db.session.add(batch)
                        db.session.flush()  # Get ID without committing
                    
                    # Find or create student
                    student = Student.query.get(adm_no)
                    # Auto-generate roll number from last 2 digits of admission number if not provided
                    roll_num = row.get('roll_number')
                    if not roll_num:
                        roll_num = adm_no[-2:]  # Last 2 digits of admission number
                    
                    if not student:
                        student = Student(
                            admission_number=adm_no,
                            full_name=row.get('name'),
                            roll_number=roll_num,
                            email=student_email,
                            branch=found_dept,
                            batch_id=batch.id,
                            batch=f"{start_year}-{expected_end_year}",  # Set batch string for mentor allocation
                            status='Live'  # Ensure status is set to Live for new students
                        )
                    else:
                        student.full_name = row.get('name', student.full_name)
                        if row.get('roll_number'):
                            student.roll_number = row.get('roll_number')
                        student.email = student_email
                        student.branch = found_dept or student.branch
                        student.batch_id = batch.id
                        student.batch = f"{start_year}-{expected_end_year}"  # Update batch string
                        if student.status == 'Passed Out':
                            student.status = 'Live'
                    
                    db.session.add(student)
                    db.session.flush() # Will trigger IntegrityError here instead of at commit if there's a problem, rolling back just this row
                    
                    processed_students.append(student)
                    affected_batches.add((found_dept, batch.id, f"{start_year}-{expected_end_year}"))
                    success_count += 1
                    
            except Exception as e:
                # The nested transaction is automatically rolled back if an exception occurs
                errors.append(f"Row {idx+1}: {str(e)}")
        
        db.session.commit()
        
        from services.batch_service import auto_allocate_batches
        allocation_result = auto_allocate_batches(list(affected_batches), mode="incremental")
        
        msg = f"Successfully processed {success_count} students."
        if errors:
            msg += f" {len(errors)} errors: " + "; ".join(errors[:5])
            if len(errors) > 5: msg += "..."
        if allocation_result.get('processed'):
            msg += f" Auto-allocated mentoring for {allocation_result.get('total_assigned', 0)} student(s)."
        if allocation_result.get('errors'):
            msg += " Allocation warnings: " + "; ".join(allocation_result['errors'][:3])
            if len(allocation_result['errors']) > 3:
                msg += "..."
            
        return jsonify({
            'success': True,
            'message': msg,
            'allocation_report': allocation_result,
            'errors': errors
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/batch/archive', methods=['POST'])
def api_archive_batch():
    """
    Manually archive a batch â†’ moves all its Live students to alumni_students,
    marks their status as 'Passed Out', and sets batch.status = 'completed'.
    Body: { "department": "...", "batch": "MCA 2023-2025" }
    """
    try:
        data = request.get_json() or {}
        department = data.get('department', '').strip()
        batch_label = data.get('batch', '').strip()

        if not batch_label:
            return jsonify({'success': False, 'message': 'batch label required'}), 400

        from services.batch_service import extract_year_range
        years = extract_year_range(batch_label)
        if not years:
            return jsonify({'success': False, 'message': f'Cannot parse years from: {batch_label}'}), 400

        start_year, end_year = years
        matched_batches = Batch.query.filter_by(start_year=start_year, end_year=end_year).all()

        # Narrow by department if given
        if department and matched_batches:
            dept_filtered = [
                b for b in matched_batches
                if Student.query.filter(
                    Student.batch_id == b.id,
                    Student.branch.ilike(department)
                ).first() is not None
            ]
            if dept_filtered:
                matched_batches = dept_filtered

        if not matched_batches:
            return jsonify({'success': False, 'message': f'No batch found for {batch_label}'}), 404

        total_promoted = 0
        archived_batches = []

        for batch in matched_batches:
            if batch.status == 'completed':
                archived_batches.append(f'{batch.start_year}-{batch.end_year} (already completed)')
                continue

            students = Student.query.filter(
                Student.batch_id == batch.id,
                Student.status.in_(['Live', 'Dropout', 'Pending'])
            ).all()

            for s in students:
                existing = AlumniStudent.query.filter_by(admission_number=s.admission_number).first()
                if not existing:
                    alumni = AlumniStudent(
                        admission_number=s.admission_number,
                        name=s.full_name,
                        email=s.email,
                        department=s.branch,
                        course_id=batch.course_id,
                        batch_id=batch.id,
                        mentor_id=s.mentor_id,
                        passout_year=batch.end_year
                    )
                    db.session.add(alumni)

                if s.mentor_id:
                    history = AlumniMentorHistory(
                        admission_number=s.admission_number,
                        mentor_id=s.mentor_id,
                        start_date=s.created_at or datetime.utcnow(),
                        end_date=datetime.utcnow()
                    )
                    db.session.add(history)

                s.status = 'Passed Out'
                s.mentor_id = None
                s.passout_year = batch.end_year
                total_promoted += 1

            batch.status = 'completed'
            archived_batches.append(f'{batch.start_year}-{batch.end_year}')

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Batch {batch_label} archived. {total_promoted} students moved to alumni.',
            'archived_batches': archived_batches,
            'students_promoted': total_promoted
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/batch/unarchive', methods=['POST'])
def api_unarchive_batch():
    """
    Reverses archiving: reactivates batch and restores students to Live.
    Body: { "batch": "MCA 2023-2025" }
    """
    try:
        data = request.get_json() or {}
        batch_label = data.get('batch', '').strip()

        if not batch_label:
            return jsonify({'success': False, 'message': 'batch label required'}), 400

        from services.batch_service import extract_year_range
        years = extract_year_range(batch_label)
        if not years:
            return jsonify({'success': False, 'message': f'Cannot parse years from: {batch_label}'}), 400

        start_year, end_year = years
        matched_batches = Batch.query.filter_by(start_year=start_year, end_year=end_year).all()

        if not matched_batches:
            return jsonify({'success': False, 'message': f'No batch found for {batch_label}'}), 404

        total_restored = 0
        for batch in matched_batches:
            batch.status = 'active'
            alumni_records = AlumniStudent.query.filter_by(batch_id=batch.id).all()
            for a in alumni_records:
                s = Student.query.get(a.admission_number)
                if s and s.status == 'Passed Out':
                    s.status = 'Live'
                    s.passout_year = None
                    total_restored += 1
                db.session.delete(a)

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Batch {batch_label} restored to active. {total_restored} students re-activated.',
            'students_restored': total_restored
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


def _transfer_mentor_notes_to_admin(student_admission_number: str):
    notes = MentorPrivateNote.query.filter_by(
        student_admission_number=student_admission_number
    ).all()
    for note in notes:
        note.transferred_to_admin = True
        note.transferred_at = datetime.utcnow()
        note.visibility = 'alumni_admin'


def _promote_batch_students_to_alumni(batch, students):
    promoted_count = 0

    for s in students:
        existing = AlumniStudent.query.filter_by(admission_number=s.admission_number).first()
        if not existing:
            alumni = AlumniStudent(
                admission_number=s.admission_number,
                name=s.full_name,
                email=s.email,
                department=s.branch,
                course_id=batch.course_id,
                batch_id=batch.id,
                mentor_id=s.mentor_id,
                passout_year=batch.end_year
            )
            db.session.add(alumni)

        if s.mentor_id:
            history = AlumniMentorHistory(
                admission_number=s.admission_number,
                mentor_id=s.mentor_id,
                start_date=s.created_at or datetime.utcnow(),
                end_date=datetime.utcnow()
            )
            db.session.add(history)

        _transfer_mentor_notes_to_admin(s.admission_number)

        s.status = 'Passed Out'
        s.mentor_id = None
        s.passout_year = batch.end_year
        promoted_count += 1

    batch.status = 'completed'
    return promoted_count


@app.route('/api/admin/semester/promote-all', methods=['POST'])
def api_promote_all_to_next_semester():
    """
    Promotes all expired batches (end_year <= current_year) to alumni.
    MCA 2023-2025 (end=2025 < 2026): promoted.
    BTech 2022-2026 (end=2026 <= 2026): promoted.
    Auto-creates new 2026 batches for any course that just graduated.
    Optional body: { "department": "..." } to restrict to one department.
    """
    try:
        data = request.get_json() or {}
        filter_dept = data.get('department', '').strip()
        current_year = datetime.now().year

        expired_batches = Batch.query.filter(
            Batch.status == 'active',
            Batch.end_year <= current_year
        ).all()

        promoted_summary = []
        new_batches_created = []
        total_promoted = 0

        for batch in expired_batches:
            course = Course.query.get(batch.course_id)
            if not course:
                continue

            if filter_dept:
                dept_match = Student.query.filter(
                    Student.batch_id == batch.id,
                    Student.branch.ilike(filter_dept)
                ).first()
                if not dept_match:
                    continue

            students = Student.query.filter(
                Student.batch_id == batch.id,
                Student.status.in_(['Live', 'Dropout', 'Pending'])
            ).all()

            batch_promoted = _promote_batch_students_to_alumni(batch, students)
            total_promoted += batch_promoted
            promoted_summary.append({
                'course': course.name,
                'batch': f'{batch.start_year}-{batch.end_year}',
                'students_promoted': batch_promoted
            })

            # Auto-create new intake batch for this course if not already present
            new_start = current_year
            if not Batch.query.filter_by(course_id=batch.course_id, start_year=new_start).first():
                duration = get_normalized_course_duration(course.name, course.duration_years or 4)
                new_batch = Batch(
                    course_id=batch.course_id,
                    start_year=new_start,
                    end_year=get_batch_end_year(new_start, duration, course.name),
                    status='active'
                )
                db.session.add(new_batch)
                new_batches_created.append(f'{course.name} {new_start}-{get_batch_end_year(new_start, duration, course.name)}')

        db.session.commit()

        if not promoted_summary:
            return jsonify({
                'success': True,
                'message': 'No expired batches found that need promotion.',
                'promoted_summary': [],
                'total_promoted': 0,
                'new_batches': []
            }), 200

        return jsonify({
            'success': True,
            'message': f'Promoted {total_promoted} students from {len(promoted_summary)} batch(es) to alumni.',
            'promoted_summary': promoted_summary,
            'total_promoted': total_promoted,
            'new_batches': new_batches_created
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/semester/promote-batch', methods=['POST'])
def api_promote_batch():
    """
    Promotes a specific batch to the next semester.
    If it is the FINAL semester (MCA/MBA sem 4, BTech sem 8, IMCA sem 10),
    moves all students to alumni and marks batch completed.
    Otherwise returns next-semester info (no DB change needed — semester is calendar-computed).

    Body: { "batch_label": "MCA 2024-2026", "department": "...", "duration_years": 2 }
    """
    try:
        data = request.get_json() or {}
        batch_label   = data.get('batch_label', '').strip()
        department    = data.get('department', '').strip()
        duration_years = int(data.get('duration_years', 4))

        if not batch_label:
            return jsonify({'success': False, 'message': 'batch_label required'}), 400

        from services.batch_service import extract_year_range
        years = extract_year_range(batch_label)
        if not years:
            return jsonify({'success': False, 'message': f'Cannot parse years from: {batch_label}'}), 400

        start_year, end_year = years
        now = datetime.now()
        current_year  = now.year
        current_month = now.month

        years_elapsed = current_year - start_year
        sem_in_year   = 1 if current_month <= 6 else 2
        max_sem       = duration_years * 2
        current_sem   = min(years_elapsed * 2 + sem_in_year, max_sem)

        if current_sem >= max_sem:
            # ── Final semester → promote to alumni ──────────────────────────
            matched_batches = Batch.query.filter_by(
                start_year=start_year, end_year=end_year
            ).all()

            # Narrow by department students if dept provided
            if department and matched_batches:
                dept_filtered = [
                    b for b in matched_batches
                    if Student.query.filter(
                        Student.batch_id == b.id,
                        Student.branch.ilike(department)
                    ).first() is not None
                ]
                if dept_filtered:
                    matched_batches = dept_filtered

            if not matched_batches:
                return jsonify({'success': False, 'message': f'No active batch found for {batch_label}'}), 404

            total_promoted = 0
            for batch in matched_batches:
                if batch.status == 'completed':
                    continue

                students = Student.query.filter(
                    Student.batch_id == batch.id,
                    Student.status.in_(['Live', 'Dropout', 'Pending'])
                ).all()

                total_promoted += _promote_batch_students_to_alumni(batch, students)

                # Auto-create new intake batch for this course if not present
                if not Batch.query.filter_by(
                    course_id=batch.course_id, start_year=current_year
                ).first():
                    course = Course.query.get(batch.course_id)
                    dur = course.duration_years if course else duration_years
                    new_batch = Batch(
                        course_id=batch.course_id,
                        start_year=current_year,
                        end_year=current_year + dur,
                        status='active'
                    )
                    db.session.add(new_batch)

            db.session.commit()

            return jsonify({
                'success': True,
                'to_alumni': True,
                'current_sem': current_sem,
                'max_sem': max_sem,
                'students_promoted': total_promoted,
                'message': (
                    f'Batch {batch_label} (Sem {current_sem}/{max_sem}) promoted to Alumni. '
                    f'{total_promoted} students graduated.'
                )
            }), 200

        else:
            # ── Not final semester — no DB change, semester is calendar-computed ───
            next_sem = current_sem + 1
            return jsonify({
                'success': True,
                'to_alumni': False,
                'current_sem': current_sem,
                'next_sem': next_sem,
                'max_sem': max_sem,
                'message': (
                    f'Batch {batch_label} advanced from Semester {current_sem} → Semester {next_sem}. '
                    f'Students remain active.'
                )
            }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============= ATTENDANCE & MENTORSHIP ROUTES =============


@app.route('/api/admin/attendance/daily_upload', methods=['POST'])
def api_upload_daily_attendance():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
            
        file = request.files['file']
        date_str = request.form.get('date')
        
        if not date_str:
             return jsonify({'success': False, 'message': 'Date is required'}), 400
             
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Read file
        import pandas as pd
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file)
        else:
            return jsonify({'success': False, 'message': 'Invalid file format'}), 400
            
        # Expected cols: admission_number, h1, h2, h3, h4, h5, h6, h7 (or 1, 2...7)
        df.columns = [str(c).lower().strip() for c in df.columns]
        
        count = 0
        errors = []
        
        # Mapping P/A/1/0 to 1/0
        def parse_status(val):
            s = str(val).upper().strip()
            if s in ['P', 'PRESENT', '1', 'TRUE']: return 1
            return 0
            
        for index, row in df.iterrows():
            adm_no = str(row.get('admission_number') or row.get('admission no') or '').strip().upper()
            if not adm_no or adm_no == 'NAN': continue
            
            # Check if student exists
            student = Student.query.get(adm_no)
            if not student:
                errors.append(f"Row {index+1}: Student {adm_no} not found")
                continue
                
            # Create or Update Attendance Record
            record = DailyAttendance.query.filter_by(student_admission_number=adm_no, date=attendance_date).first()
            if not record:
                record = DailyAttendance(student_admission_number=adm_no, date=attendance_date)
            
            # Parse hours (look for h1, hour1, 1, etc)
            record.hour_1 = parse_status(row.get('h1') or row.get('1') or 0)
            record.hour_2 = parse_status(row.get('h2') or row.get('2') or 0)
            record.hour_3 = parse_status(row.get('h3') or row.get('3') or 0)
            record.hour_4 = parse_status(row.get('h4') or row.get('4') or 0)
            record.hour_5 = parse_status(row.get('h5') or row.get('5') or 0)
            record.hour_6 = parse_status(row.get('h6') or row.get('6') or 0)
            record.hour_7 = parse_status(row.get('h7') or row.get('7') or 0)
            
            db.session.add(record)
            
            # ALSO populating StudentAttendance for our new analytics table
            status_val = 'P' if record.hour_1 == 1 else 'A'
            sa = StudentAttendance.query.filter_by(student_admission_number=adm_no, date=attendance_date).first()
            if not sa:
                sa = StudentAttendance(student_admission_number=adm_no, date=attendance_date)
            sa.status = status_val
            db.session.add(sa)
            
            count += 1
            
        db.session.commit()
        
        # Trigger analytics (Safe: After upload, compute trends)
        run_full_analysis()
        
        return jsonify({
            'success': True, 
            'message': f'Processed {count} records and triggered analytics engine.',
            'errors': errors[:10] # Return first 10 errors
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/upload_attendance', methods=['POST'])
def api_upload_attendance():
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
            
        count = 0
        for row in data:
            student_id = row.get('student_id')
            date_str = row.get('date')
            status = row.get('status')
            
            if not student_id or not date_str or not status: continue
            
            # Upsert
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            record = StudentAttendance.query.filter_by(student_admission_number=student_id, date=attendance_date).first()
            if not record:
                record = StudentAttendance(student_admission_number=student_id, date=attendance_date)
                
            record.status = status
            db.session.add(record)
            count += 1
            
        db.session.commit()
        
        # Trigger analytics (Safe: After upload, compute trends)
        run_full_analysis()
        
        return jsonify({'success': True, 'message': 'Upload successful & analytics triggered'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/alerts/<role>/<id>', methods=['GET'])
def api_get_alerts(role, id):
    try:
        from sqlalchemy import desc
        if role == 'student':
            alerts = Alert.query.filter_by(student_admission_number=id).order_by(desc(Alert.created_at)).limit(10).all()
        elif role == 'mentor':
            alerts = Alert.query.filter_by(mentor_id=id).order_by(desc(Alert.created_at)).limit(20).all()
        else:
            return jsonify({'success': False, 'message': 'Invalid role'}), 400
            
        return jsonify({
            'success': True,
            'data': [{
                'id': a.id,
                'student_id': a.student_admission_number,
                'type': a.type,
                'message': a.message,
                'is_read': a.is_read,
                'created_at': a.created_at.strftime('%Y-%m-%d %H:%M')
            } for a in alerts]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= PHASE 2 TREND ANALYTICS ROUTES =============
from analytics.trends import compute_student_analytics
from models import StudentAnalytics

def calculate_analytics(student_id):
    # Fetch chronological daily attendance
    from sqlalchemy import asc
    from models import StudentMark
    attendances = DailyAttendance.query.filter_by(student_admission_number=student_id).order_by(asc(DailyAttendance.date)).all()
    
    # Flatten all class hours to a chronological list of 1s and 0s.
    att_records = []
    for a in attendances:
        for hour in range(1, 8):
            val = getattr(a, f'hour_{hour}', None)
            if val in (0, 1, True, False):
                att_records.append(1 if val == 1 or val is True else 0)

    if not att_records:
        subject_attendance = Attendance.query.filter_by(student_admission_number=student_id).all()
        for row in subject_attendance:
            total = int(row.total_classes or 0)
            attended = int(row.attended_classes or 0)
            if total > 0:
                att_records.extend([1] * min(attended, total))
                att_records.extend([0] * max(total - attended, 0))
        
    # Fetch internal marks
    internal_marks_objs = InternalMark.query.filter_by(student_id=student_id).all()
    marks_dict = {}
    for m in internal_marks_objs:
        if m.subject_id not in marks_dict:
            marks_dict[m.subject_id] = {}
        marks_dict[m.subject_id][m.exam_type] = m.marks

    if not marks_dict:
        student_marks = StudentMark.query.filter_by(student_id=student_id).all()
        for mark in student_marks:
            key = mark.subject_code or f"semester_{mark.semester or 'unknown'}"
            marks_dict.setdefault(key, {})
            for idx, value in enumerate([mark.internal1, mark.internal2, mark.internal3], start=1):
                if value is not None:
                    marks_dict[key][f"Internal{idx}"] = float(value)
            if mark.university_mark is not None and not marks_dict[key]:
                marks_dict[key]["University"] = float(mark.university_mark)
        
    metrics = compute_student_analytics(student_id, att_records, marks_dict)
    
    # Store or update in DB (Option B implementation)
    sa = StudentAnalytics.query.filter_by(student_id=student_id).first()
    if not sa:
        sa = StudentAnalytics(student_id=student_id)
        db.session.add(sa)
    
    sa.attendance_percentage = metrics['attendance_percentage']
    sa.attendance_slope = metrics['attendance_slope']
    sa.avg_internal_marks = metrics['avg_internal_marks']
    sa.marks_slope = metrics['marks_slope']
    sa.failure_count = metrics['failure_count']
    sa.marks_variance = metrics['marks_variance']
    sa.risk_score = metrics['risk_score']
    sa.status = metrics['status']
    sa.last_updated = datetime.utcnow()
    
    try:
        from analytics.train_model import predict_student_risk
        sa.ml_risk_probability = predict_student_risk(sa)
    except Exception as e:
        sa.ml_risk_probability = 0.0

    # ----- PHASE 5: COMPLIANCE-AWARE RISK RECALIBRATION -----
    try:
        from models import WeeklyStudyPlan
        recent_plan = WeeklyStudyPlan.query.filter_by(student_id=student_id).order_by(WeeklyStudyPlan.week_start.desc()).first()
        compliance_fraction = None
        
        if recent_plan:
            allocated = sum(sub.allocated_hours for sub in recent_plan.subjects)
            if allocated > 0:
                completed = sum(sum(log.hours_completed for log in sub.sessions) for sub in recent_plan.subjects)
                compliance_fraction = min(1.0, completed / allocated)
        
        if compliance_fraction is not None:
            if compliance_fraction >= 0.8:
                risk_modifier = -0.05
            elif compliance_fraction >= 0.4:
                risk_modifier = 0.0
            else:
                risk_modifier = 0.10
        else:
            risk_modifier = 0.0
            
        sa.compliance_modifier = risk_modifier

        r_det = (sa.risk_score or 0.0) / 100.0
        r_ml = (sa.ml_risk_probability or 0.0) / 100.0

        r_det_adj = max(0.0, min(1.0, r_det + risk_modifier))
        r_ml_adj = max(0.0, min(1.0, r_ml + risk_modifier))

        # When ML risk is unavailable or returns zero, keep the deterministic risk
        # as the primary score instead of averaging it down by half.
        if (sa.ml_risk_probability or 0.0) > 0:
            adjusted_risk = (0.5 * r_det_adj + 0.5 * r_ml_adj) * 100.0
        else:
            adjusted_risk = r_det_adj * 100.0

        sa.adjusted_risk = round(max(0.0, min(100.0, adjusted_risk)), 2)
        if sa.adjusted_risk >= 60:
            sa.status = "Declining"
        elif sa.adjusted_risk <= 30 and sa.attendance_slope and sa.attendance_slope > 0 and sa.marks_slope and sa.marks_slope > 0:
            sa.status = "Improving"
        elif sa.adjusted_risk < 60 and sa.status == "Declining":
            sa.status = "Stable"

        # ----- PHASE 7: CAUSAL OUTCOME TRACKING -----
        try:
            from datetime import date, timedelta
            today = date.today()
            current_week_start = today - timedelta(days=today.weekday())
            prev_week_start = current_week_start - timedelta(days=7)

            # Check for intervention from exactly one week ago
            past_interv = MentorIntervention.query.filter_by(
                student_id=student_id,
                week_start=prev_week_start
            ).first()

            if past_interv:
                # Idempotency check: unique intervention_id
                existing_outcome = InterventionOutcome.query.filter_by(intervention_id=past_interv.id).first()
                if not existing_outcome:
                    delta = sa.adjusted_risk - past_interv.risk_snapshot
                    
                    if delta <= -10:
                        label = "Improved"
                    elif delta < 10:
                        label = "Static"
                    else:
                        label = "Regressed"
                        
                    sub_comp = (compliance_fraction * 100.0) if compliance_fraction is not None else 0.0
                    shift = sub_comp - past_interv.compliance_snapshot
                    
                    outcome = InterventionOutcome(
                        intervention_id=past_interv.id,
                        evaluated_week_start=current_week_start,
                        initial_risk=past_interv.risk_snapshot,
                        subsequent_risk=sa.adjusted_risk,
                        delta=delta,
                        initial_compliance=past_interv.compliance_snapshot,
                        subsequent_compliance=sub_comp,
                        compliance_shift=shift,
                        outcome_label=label
                    )
                    db.session.add(outcome)
                    # We don't commit yet; the main function commits at the end
        except Exception as eval_e:
            print(f"Phase 7 Outcome Evaluation Error: {eval_e}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        sa.compliance_modifier = 0.0
        if (sa.ml_risk_probability or 0.0) > 0:
            sa.adjusted_risk = (0.5 * ((sa.risk_score or 0.0) / 100.0) + 0.5 * ((sa.ml_risk_probability or 0.0) / 100.0)) * 100.0
        else:
            sa.adjusted_risk = float(sa.risk_score or 0.0)

    db.session.commit()
    
    metrics['ml_risk_probability'] = sa.ml_risk_probability
    metrics['compliance_modifier'] = sa.compliance_modifier
    metrics['adjusted_risk'] = sa.adjusted_risk
    return metrics

@app.route('/api/analytics/student/<string:student_id>', methods=['GET'])
def api_get_student_analytics(student_id):
    try:
        # Check if student exists
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
            
        metrics = calculate_analytics(student_id) # On-request (Dynamic generation/update) for test purposes
        return jsonify({'success': True, 'data': metrics}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/analytics/mentor/<int:mentor_id>', methods=['GET'])
def api_get_mentor_analytics(mentor_id):
    try:
        mentees = Student.query.filter_by(mentor_id=mentor_id, status='Live').all()
        data = []
        for m in mentees:
            # Check DB or compute
            sa = StudentAnalytics.query.filter_by(student_id=m.admission_number).first()
            if not sa:
                metrics = calculate_analytics(m.admission_number)
            else:
                metrics = {
                    "attendance_percentage": sa.attendance_percentage,
                    "attendance_slope": sa.attendance_slope,
                    "avg_internal_marks": sa.avg_internal_marks,
                    "marks_slope": sa.marks_slope,
                    "failure_count": sa.failure_count,
                    "marks_variance": sa.marks_variance,
                    "risk_score": sa.risk_score,
                    "status": sa.status,
                    "ml_risk_probability": sa.ml_risk_probability,
                    "compliance_modifier": sa.compliance_modifier,
                    "adjusted_risk": sa.adjusted_risk
                }
            data.append({
                "student_id": m.admission_number,
                "name": m.full_name,
                "batch": m.batch or "",
                "risk_score": metrics.get("risk_score", 0.0) or 0.0,
                "attendance_trend": metrics.get("attendance_slope", 0.0) or 0.0,
                "marks_trend": metrics.get("marks_slope", 0.0) or 0.0,
                "status": metrics.get("status", "Stable") or "Stable",
                "ml_risk_probability": metrics.get("ml_risk_probability", 0.0) or 0.0,
                "adjusted_risk": metrics.get("adjusted_risk", 0.0) or 0.0,
                "attendance_percentage": metrics.get("attendance_percentage", 0.0) or 0.0,
                "avg_internal_marks": metrics.get("avg_internal_marks", 0.0) or 0.0
            })
            
        return jsonify({'success': True, 'data': data}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/mentorship/view', methods=['GET'])
def api_get_mentors_view():
    try:
        dept = request.args.get('department')
        batch = request.args.get('batch')  # Could be "MCA 2024-2026" or "2024-2026"
        
        if not dept:
            return jsonify({'success': False, 'message': 'Department required'}), 400
            
        dept = normalize_dept_name(dept)
        
        # Build a flexible faculty/student department filter.
        # Faculty and Student records may be stored as short codes ("MCA", "IMCA") OR full names.
        from services.batch_service import is_mentor_eligible
        
        ca_dept = 'Department of Computer Applications'
        if dept == ca_dept:
            # Match MCA, IMCA, Computer Applications, or explicit full name
            faculty_filter = db.or_(
                Faculty.department.ilike('MCA'),
                Faculty.department.ilike('IMCA'),
                Faculty.department.ilike('%Computer Applications%'),
            )
            student_dept_filter = db.or_(
                Student.branch.ilike('MCA'),
                Student.branch.ilike('IMCA'),
                Student.branch.ilike('%Computer Applications%'),
            )
        else:
            faculty_filter = Faculty.department.ilike(dept)
            student_dept_filter = Student.branch.ilike(f'%{dept}%')
        
        all_faculty = Faculty.query.filter(
            faculty_filter,
            Faculty.status == 'Live'
        ).all()
        
        mentors = [f for f in all_faculty if is_mentor_eligible(f)]
        
        data = []
        for m in mentors:
            # Total Load (all batches)
            total_load = Student.query.filter_by(mentor_id=m.id, status='Live').count()
            
            # Batch Specific Query - Use year range matching
            batch_query = Student.query.filter(
                Student.mentor_id == m.id,
                Student.status == 'Live'
            )
            
            if batch:
                # Extract years from batch label (handles both "MCA 2024-2026" and "2024-2026")
                from services.batch_service import extract_year_range
                target_years = extract_year_range(batch)
                
                if target_years:
                    # Match by year range instead of exact batch string
                    batch_query = batch_query.filter(
                        db.or_(
                            Student.batch.ilike(f"%{target_years[0]}-{target_years[1]}%"),
                            db.and_(
                                Student.batch.ilike(f"%{target_years[0]}%"),
                                Student.batch.ilike(f"%{target_years[1]}%")
                            )
                        )
                    )
            
            mentees = batch_query.all()
            
            data.append({
                'id': m.id,
                'name': m.name,
                'designation': m.designation,
                'total_load': total_load,
                'batch_mentee_count': len(mentees),
                'mentees': [{'admission_number': s.admission_number, 'name': s.full_name, 'batch': s.batch} for s in mentees]
            })
            
        # Find unassigned students using the same flexible filter
        unassigned_query = Student.query.filter(
            student_dept_filter,
            Student.status == 'Live',
            Student.mentor_id == None
        )
        
        if batch:
            from services.batch_service import extract_year_range
            target_years = extract_year_range(batch)
            if target_years:
                unassigned_query = unassigned_query.filter(
                    db.or_(
                        Student.batch.ilike(f"%{target_years[0]}-{target_years[1]}%"),
                        db.and_(
                            Student.batch.ilike(f"%{target_years[0]}%"),
                            Student.batch.ilike(f"%{target_years[1]}%")
                        )
                    )
                )
             
        unassigned_count = unassigned_query.count()
        
        return jsonify({
            'success': True, 
            'data': data,
            'unassigned_count': unassigned_count
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/mentor/mentees', methods=['GET'])
def api_get_my_mentees():
    try:
        username = request.args.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Mentor username required'}), 400
            
        # Find mentor
        mentor = Faculty.query.filter_by(username=username).first()
        if not mentor:
            return jsonify({'success': False, 'message': 'Mentor not found'}), 404
            
        # Get mentees
        mentees = Student.query.filter_by(mentor_id=mentor.id).all()
        
        from academic_utils import calculate_academic_status
        mentee_data = []
        for s in mentees:
            acad = calculate_academic_status(s.batch, s.branch)
            mentee_data.append({
                'id': s.admission_number,
                'name': s.full_name,
                'batch': f"{s.branch} {s.batch if s.batch else ''}".strip(),
                'status': s.status,
                'risk': 'low',
                'lastMeeting': 'Recent',
                'academic_info': acad
            })
            
        return jsonify({
            'success': True,
            'data': mentee_data
        }), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/stats', methods=['GET'])
def api_get_admin_stats():
    try:
        # All active/enrolled students (not passed out)
        total_students = Student.query.filter(Student.status != 'Passed Out').count()
        # All active mentors assigned to actual departments (exclude Basic Sciences & Humanities and include only Professors and designated mentors - NOT HODs)
        active_mentors = Faculty.query.filter(
            Faculty.status == 'Live',
            ~Faculty.department.ilike('%Basic Sciences%'),
            ~Faculty.department.ilike('%Humanities%'),
            ~Faculty.designation.ilike('%HOD%'),  # Exclude HODs from mentorship
            (Faculty.designation.ilike('%Professor%')) | 
            (Faculty.designation.ilike('%Mentor%'))
        ).count()
        
        # Count unique departments from both students and faculty, handling Computer Applications unification
        try:
            student_depts_raw = db.session.query(Student.branch).distinct().all()
            faculty_depts_raw = db.session.query(Faculty.department).distinct().all()
                    
            # Combine and normalize departments, handling Computer Applications unification
            all_depts = set()
            
            # Process student departments
            for dept in student_depts_raw:
                if dept[0]:  # Check if not None
                    dept_name = dept[0]
                    # Map MCA/IMCA/Computer Applications to a unified Computer Applications department
                    if dept_name in ['MCA', 'IMCA', 'Computer Applications'] or 'Computer Applications' in dept_name:
                        all_depts.add('Computer Applications')
                    else:
                        all_depts.add(dept_name)
            
            # Process faculty departments
            for dept in faculty_depts_raw:
                if dept[0]:  # Check if not None
                    dept_name = dept[0]
                    # Map MCA/IMCA/Computer Applications to a unified Computer Applications department
                    if dept_name in ['MCA', 'IMCA', 'Computer Applications'] or 'Computer Applications' in dept_name:
                        all_depts.add('Computer Applications')
                    else:
                        all_depts.add(dept_name)
                    
            departments = len(all_depts)
        except Exception as e:
            # Fallback in case of database error
            departments = 0
        
        # Count total courses
        total_courses = Course.query.count()
        
        return jsonify({
            'success': True,
            'data': {
                'total_students': total_students,
                'active_mentors': active_mentors,
                'departments': departments,
                'total_courses': total_courses
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/departments', methods=['GET'])
def api_get_departments():
    try:
        # Get unique departments from students
        student_departments = db.session.query(Student.branch).distinct().all()
        student_depts = [dept[0] for dept in student_departments if dept[0]]
        
        # Get unique departments from faculty
        faculty_departments = db.session.query(Faculty.department).distinct().all()
        faculty_depts = [dept[0] for dept in faculty_departments if dept[0]]
        
        # Get all courses (departments) from the Course table
        all_courses = Course.query.all()
        course_depts = [(course.name, course.code) for course in all_courses if course.name]
        
        # Combine and deduplicate - keep original department names to preserve consistency
        all_departments = list(set(student_depts + faculty_depts + [name for name, _ in course_depts]))
        
        # Sort alphabetically
        all_departments.sort()
        
        # Create detailed response with faculty counts and codes
        departments_data = []
        for dept in all_departments:
            # Count faculty members for this department
            faculty_count = Faculty.query.filter_by(department=dept).count()
            
            # Count students for this department (with 'Live' status)
            students_count = Student.query.filter(
                Student.branch == dept,
                Student.status != 'Passed Out'
            ).count()
            
            # Get the department code from Course table
            try:
                course = Course.query.filter_by(name=dept).first()
                dept_code = course.code if course and course.code else dept[:3].upper()
            except:
                # Handle case where 'code' column doesn't exist in the database yet
                dept_code = dept[:3].upper()
            
            departments_data.append({
                'name': dept,
                'code': dept_code,
                'faculty_count': faculty_count,
                'students_count': students_count
            })
        
        return jsonify({
            'success': True,
            'data': departments_data
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/departments', methods=['POST'])
def api_add_department():
    try:
        data = request.get_json()
        dept_name = data.get('name', '').strip()
        dept_code = data.get('code', '').strip().upper()
        
        if not dept_name:
            return jsonify({'success': False, 'message': 'Department name is required'}), 400
        
        # If no code provided, derive from department name
        if not dept_code:
            # Map department names to codes
            dept_to_code = {
                'Computer Applications': 'CA',
                'Computer Science & Engineering': 'CSE',
                'Civil Engineering': 'CE',
                'Mechanical Engineering': 'ME',
                'Electrical & Electronics Engineering': 'EEE',
                'Electronics & Communication Engineering': 'ECE',
                'Management Studies': 'MBA',
                'Basic Sciences & Humanities': 'BSH',
                'MCA': 'MCA',
                'IMCA': 'IMCA'
            }
            dept_code = dept_to_code.get(dept_name, dept_name[:3].upper())
        
        # Check if department already exists
        existing_course = Course.query.filter_by(name=dept_name).first()
        if existing_course:
            return jsonify({'success': False, 'message': 'Department/Course already exists'}), 400
        
        # Create a new course record for the department
        new_course = Course(name=dept_name, code=dept_code)
        db.session.add(new_course)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Department/Course "{dept_name}" added successfully with code "{dept_code}"',
            'data': {'name': dept_name, 'code': dept_code}
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/departments/<string:dept_name>', methods=['DELETE'])
def api_remove_department(dept_name):
    try:
        # Find and delete the course record
        course = Course.query.filter_by(name=dept_name).first()
        if not course:
            return jsonify({'success': False, 'message': 'Department/Course not found'}), 404
        
        # Check if there are students or faculty in this department
        students_in_dept = Student.query.filter_by(branch=dept_name).count()
        faculty_in_dept = Faculty.query.filter_by(department=dept_name).count()
        
        if students_in_dept > 0 or faculty_in_dept > 0:
            return jsonify({
                'success': False, 
                'message': f'Cannot delete department with {students_in_dept} students and {faculty_in_dept} faculty members'
            }), 400
        
        db.session.delete(course)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Department/Course "{dept_name}" removed successfully'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= PHASE 4C PLANNER ROUTES =============
from analytics.planner import generate_and_lock_weekly_plan
from models import WeeklyStudyPlan, StudyPlanSubject, StudySessionLog, MentorIntervention
from datetime import date, timedelta

@app.route('/api/planner/<string:student_id>', methods=['GET'])
def api_get_student_planner(student_id):
    try:
        plan = generate_and_lock_weekly_plan(student_id)
        if not plan:
            return jsonify({'success': False, 'message': 'Could not generate plan'}), 400

        # Serialize
        subjects = []
        for sub in plan.subjects:
            subjects.append({
                "id": sub.id,
                "subject_id": sub.subject_id,
                "subject_name": sub.subject.name if sub.subject else "Unknown",
                "allocated_hours": sub.allocated_hours,
                "priority": sub.priority,
                "completed_hours": sum(log.hours_completed for log in sub.sessions)
            })

        return jsonify({
            'success': True,
            'data': {
                'id': plan.id,
                'week_start': plan.week_start.isoformat(),
                'week_end': plan.week_end.isoformat(),
                'total_hours': plan.total_hours,
                'booster_applied': plan.booster_applied,
                'deterministic_risk': plan.deterministic_risk,
                'ml_probability': plan.ml_probability,
                'subjects': subjects
            }
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/planner/log-session', methods=['POST'])
def api_log_session():
    try:
        data = request.json
        plan_subject_id = data.get('plan_subject_id')
        hours_completed = data.get('hours_completed', 0)

        if not plan_subject_id:
            return jsonify({'success': False, 'message': 'plan_subject_id required'}), 400

        try:
            hours = float(hours_completed)
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid hours format'}), 400

        # Server-side validation
        if hours <= 0:
            return jsonify({'success': False, 'message': 'Hours must be greater than 0'}), 400
        if hours > 24:
            return jsonify({'success': False, 'message': 'Cannot log more than 24 hours in a single session'}), 400

        log = StudySessionLog(
            plan_subject_id=plan_subject_id,
            date=date.today(),
            hours_completed=hours
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Session logged successfully'}), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/planner/mentor/<int:mentor_id>', methods=['GET'])
def api_get_mentor_planner(mentor_id):
    try:
        mentees = Student.query.filter_by(mentor_id=mentor_id, status='Live').all()
        data = []
        for m in mentees:
            plan = generate_and_lock_weekly_plan(m.admission_number)
            if not plan:
                continue

            completed = 0.0
            allocated = sum(sub.allocated_hours for sub in plan.subjects)
            for sub in plan.subjects:
                completed += sum(log.hours_completed for log in sub.sessions)

            compliance_raw = (completed / allocated) * 100 if allocated > 0 else 0
            compliance = min(100.0, compliance_raw)

            data.append({
                "student_id": m.admission_number,
                "name": m.full_name,
                "risk_score": plan.deterministic_risk,
                "ml_probability": plan.ml_probability,
                "compliance": round(compliance, 1),
                "plan_generated": True,
                "total_allocated": allocated,
                "total_completed": completed
            })

        return jsonify({'success': True, 'data': data}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= PHASE 6 INTERVENTION ROUTES =============

@app.route('/api/intervention/create', methods=['POST'])
def api_create_intervention():
    try:
        data = request.json
        student_id = data.get('student_id')
        mentor_id = data.get('mentor_id')
        intervention_type = data.get('intervention_type')
        notes = data.get('notes', '')

        if not all([student_id, mentor_id, intervention_type]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400

        # Validate Mentor
        student = Student.query.filter_by(admission_number=student_id).first()
        if not student or student.mentor_id != int(mentor_id):
            return jsonify({'success': False, 'message': 'Unauthorized mentor for this student'}), 403

        # Validate Duplicate in Week
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
        
        existing = MentorIntervention.query.filter_by(
            student_id=student_id,
            week_start=week_start
        ).first()

        if existing:
            return jsonify({'success': False, 'message': 'Intervention already logged for this week'}), 400

        # Fetch Snapshots
        sa = StudentAnalytics.query.filter_by(student_id=student_id).first()
        risk_snap = sa.adjusted_risk if sa and sa.adjusted_risk else 0.0

        comp_snap = 0.0
        recent_plan = WeeklyStudyPlan.query.filter_by(student_id=student_id, week_start=week_start).first()
        if recent_plan:
            allocated = sum(sub.allocated_hours for sub in recent_plan.subjects)
            if allocated > 0:
                completed = sum(sum(log.hours_completed for log in sub.sessions) for sub in recent_plan.subjects)
                comp_snap = min(100.0, (completed / allocated) * 100)

        new_intervention = MentorIntervention(
            student_id=student_id,
            mentor_id=mentor_id,
            week_start=week_start,
            risk_snapshot=risk_snap,
            compliance_snapshot=comp_snap,
            intervention_type=intervention_type,
            notes=notes,
            locked=True
        )
        db.session.add(new_intervention)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Intervention logged successfully'}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/intervention/mentor/<int:mentor_id>', methods=['GET'])
def api_get_mentor_interventions(mentor_id):
    try:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
        
        interventions = MentorIntervention.query.filter_by(
            mentor_id=mentor_id,
            week_start=week_start
        ).all()
        
        logged_map = {}
        for interv in interventions:
            logged_map[interv.student_id] = True
            
        return jsonify({'success': True, 'data': logged_map}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============= ACADEMIC LIFECYCLE MANAGEMENT API ROUTES =============

@app.route('/api/admin/courses', methods=['GET'])
def api_get_courses():
    """Get all courses"""
    try:
        courses = Course.query.all()
        result = []
        for course in courses:
            result.append({
                'id': course.id,
                'name': course.name,
                'duration_years': get_normalized_course_duration(course.name, getattr(course, 'duration_years', 4))
            })
        
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/course/create', methods=['POST'])
def api_create_course():
    """Create a new course"""
    try:
        data = request.get_json()
        name = data.get('name')
        duration_years = data.get('duration_years', 4)
        
        if not name:
            return jsonify({'success': False, 'message': 'Course name is required'}), 400
        
        # Check if course already exists
        existing_course = Course.query.filter_by(name=name).first()
        if existing_course:
            return jsonify({'success': False, 'message': 'Course already exists'}), 400
        
        new_course = Course(name=name, duration_years=duration_years)
        db.session.add(new_course)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Course {name} created successfully',
            'data': {
                'id': new_course.id,
                'name': new_course.name,
                'duration_years': new_course.duration_years
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/batches', methods=['GET'])
def api_get_batches():
    """Get all batches with course information"""
    try:
        batches = db.session.query(Batch, Course).join(Course).all()
        current_year = datetime.now().year

        grouped = {}
        for batch, course in batches:
            display_course_name = get_display_course_name(batch, course.name)

            key = (display_course_name, batch.start_year, batch.end_year)
            effective_end_year = batch.end_year
            if key not in grouped:
                grouped[key] = {
                    'id': batch.id,
                    'course_name': display_course_name,
                    'course_id': course.id,
                    'start_year': batch.start_year,
                    'end_year': effective_end_year,
                    'raw_statuses': set(),
                    'duplicate_count': 0,
                }

            grouped[key]['id'] = min(grouped[key]['id'], batch.id)
            grouped[key]['raw_statuses'].add((batch.status or '').strip().lower())
            grouped[key]['duplicate_count'] += 1

        course_active_starts = {}
        for item in grouped.values():
            statuses = item.get('raw_statuses', set())
            has_active = 'active' in statuses
            has_completed = 'completed' in statuses

            if has_active or not has_completed:
                course_active_starts.setdefault(item['course_id'], []).append(int(item['start_year']))

        result = []
        for item in grouped.values():
            statuses = item.pop('raw_statuses', set())
            has_active = 'active' in statuses
            has_completed = 'completed' in statuses

            if has_completed and not has_active:
                status_tier = 'alumni'
                status_label = 'Alumni'
                status = 'completed'
            else:
                course_name = item.get('course_name', '')
                years_left = int(item['end_year']) - int(current_year)
                active_starts = course_active_starts.get(item['course_id'], [])
                has_newer_active_batch = any(start_year > int(item['start_year']) for start_year in active_starts)

                if years_left == 0:
                    status_tier = 'final_year'
                    status_label = 'Final Year'
                else:
                    status_tier = 'ongoing'
                    status_label = 'Ongoing'
                status = 'active'

            item['status'] = status
            item['status_tier'] = status_tier
            item['status_label'] = status_label
            item['is_completed'] = status_tier == 'alumni'
            result.append(item)

        result.sort(key=lambda row: (row['end_year'], row['start_year'], row['course_name']), reverse=True)
        
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/batch/create', methods=['POST'])
def api_create_batch():
    """Create a new batch and automatically roll qualifying senior batches into alumni."""
    try:
        data = request.get_json()
        course_id = data.get('course_id')
        start_year = data.get('start_year')
        
        if not course_id or not start_year:
            return jsonify({'success': False, 'message': 'Course ID and Start Year are required'}), 400
        
        # Get course duration
        course = Course.query.get(course_id)
        if not course:
            return jsonify({'success': False, 'message': 'Course not found'}), 404
        
        # Calculate end_year
        end_year = get_batch_end_year(start_year, course.duration_years, course.name)
        
        existing_batch = Batch.query.filter_by(course_id=course_id, start_year=start_year).first()
        if existing_batch:
            has_students = Student.query.filter_by(batch_id=existing_batch.id).first() is not None
            has_alumni = AlumniStudent.query.filter_by(batch_id=existing_batch.id).first() is not None
            if not has_students and not has_alumni:
                db.session.delete(existing_batch)
                db.session.flush()
            else:
                return jsonify({
                    'success': False,
                    'message': f'Batch {course.name} {start_year}-{existing_batch.end_year} already exists'
                }), 409

        # When a new intake is added, any active batch ending in that year or earlier becomes alumni.
        completion_cutoff_year = max(datetime.now().year, int(start_year))
        completed_batches = db.session.query(Batch, Course).join(Course).filter(
            Batch.end_year <= completion_cutoff_year,
            Batch.status == 'active'
        ).all()

        completed_batch_list = []
        moved_student_count = 0
        for old_batch, old_course in completed_batches:
            students = Student.query.filter(
                Student.batch_id == old_batch.id,
                Student.status.in_(['Live', 'Dropout', 'Pending'])
            ).all()

            moved_student_count += _promote_batch_students_to_alumni(old_batch, students)
            completed_batch_list.append(f"{old_course.name} {old_batch.start_year}-{old_batch.end_year}")

        # Create the new batch
        new_batch = Batch(
            course_id=course_id,
            start_year=start_year,
            end_year=end_year,
            status='active'
        )
        
        db.session.add(new_batch)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': (
                f'Batch {course.name} {start_year}-{end_year} created successfully. '
                f'Automatically moved {moved_student_count} student(s) from {len(completed_batch_list)} completed batch(es) to alumni.'
                if completed_batch_list else
                f'Batch {course.name} {start_year}-{end_year} created successfully'
            ),
            'data': {
                'id': new_batch.id,
                'course_name': course.name,
                'start_year': start_year,
                'end_year': end_year
            },
            'completed_batches': completed_batch_list,
            'moved_student_count': moved_student_count
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/batch/confirm_completion', methods=['POST'])
def api_confirm_batch_completion():
    """Confirm completion of batches and move students to alumni"""
    try:
        data = request.get_json()
        confirmed = data.get('confirmed', False)
        
        if not confirmed:
            return jsonify({'success': False, 'message': 'Confirmation required'}), 400
        
        # Find all completed batches (end_year <= current_year and status is active)
        current_year = datetime.now().year
        completed_batches = Batch.query.filter(
            (Batch.end_year <= current_year) & (Batch.status == 'active')
        ).all()
        
        if not completed_batches:
            return jsonify({'success': False, 'message': 'No completed batches found'}), 404
        
        # Process each completed batch
        moved_student_count = 0
        for batch in completed_batches:
            # Get students in this batch
            students = Student.query.filter_by(batch_id=batch.id).all()

            moved_student_count += _promote_batch_students_to_alumni(batch, students)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Moved {moved_student_count} students to alumni from {len(completed_batches)} batches'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# Update the mentor allocation logic to work with the new batch system
@app.route('/api/admin/mentorship/allocate', methods=['POST'])
def api_allocate_mentors():
    try:
        data = request.get_json()
        department = data.get('department')
        batch_label = data.get('batch')  # This could be "MCA 2024-2026", "2024-2026", or None
        
        if not department:
            return jsonify({'success': False, 'message': 'Department required'}), 400
        
        from models import Batch
        from services.batch_service import extract_year_range, redistribute_mentors

        # ── Case 1: No batch specified → redistribute ALL students in dept ──
        if not batch_label:
            # Get all active batches in any department (we'll filter by student branch)
            all_batches = Batch.query.filter_by(status='active').all()
            total_students_moved = 0
            total_mentors_used = 0
            errors = []
            for b in all_batches:
                batch_lbl = f"{b.start_year}-{b.end_year}"
                res = redistribute_mentors(
                    department=department,
                    batch_id=b.id,
                    batch_label=batch_lbl,
                    mode="full"
                )
                if res.get('error'):
                    errors.append(res['error'])
                else:
                    total_students_moved += res.get('total_students', 0)
                    total_mentors_used = max(total_mentors_used, res.get('total_mentors', 0))
            
            if total_students_moved == 0 and errors:
                return jsonify({'success': False, 'message': '; '.join(errors)}), 500
            
            return jsonify({
                'success': True,
                'message': f"Successfully allocated {total_students_moved} students across {total_mentors_used} mentors in {department}"
            }), 200

        # ── Case 2: Specific batch provided ─────────────────────────────────
        target_years = extract_year_range(batch_label)
        
        if not target_years:
            return jsonify({
                'success': False,
                'message': f'Invalid batch format: {batch_label}'
            }), 400
        
        # Find batch_id that matches the years and has students in this department
        batch_id = None
        all_batches = Batch.query.all()
        for b in all_batches:
            if (b.start_year == target_years[0] and 
                b.end_year == target_years[1]):
                # Prefer the batch that actually has students in this department
                sample_student = Student.query.filter(
                    Student.branch.ilike(department.strip()),
                    Student.batch_id == b.id
                ).first()
                if sample_student:
                    batch_id = b.id
                    break
        
        if not batch_id:
            # Fallback: take any batch with matching years
            for b in all_batches:
                if (b.start_year == target_years[0] and 
                    b.end_year == target_years[1]):
                    batch_id = b.id
                    break
        
        if not batch_id:
            return jsonify({
                'success': False,
                'message': f'No matching batch found for {batch_label}'
            }), 404
        
        clean_batch_label = f"{target_years[0]}-{target_years[1]}"
        
        result = redistribute_mentors(
            department=department,
            batch_id=batch_id,
            batch_label=clean_batch_label,
            mode="full"
        )
        
        if result.get("error"):
            return jsonify({
                'success': False,
                'message': result.get("error")
            }), 500
        
        return jsonify({
            'success': True,
            'message': f"Successfully allocated {result.get('total_students', 0)} students to {result.get('total_mentors', 0)} mentors",
            'distribution': result.get('distribution', {})
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500



# ============= ALUMNI TRACKING API ROUTES =============

@app.route('/api/admin/alumni/departments', methods=['GET'])
def api_get_alumni_departments():
    """Get all departments with alumni counts"""
    try:
        sync_passed_out_students_to_alumni()
        # Get distinct departments with alumni
        departments = db.session.query(
            AlumniStudent.department,
            db.func.count(AlumniStudent.id).label('alumni_count')
        ).group_by(AlumniStudent.department).all()
        
        result = []
        for dept, count in departments:
            if dept:  # Skip null departments
                result.append({
                    'department': dept,
                    'alumni_count': count
                })
        
        return jsonify({
            'success': True,
            'data': result
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/alumni/batches', methods=['GET'])
def api_get_alumni_batches():
    """Get all batches with alumni counts"""
    try:
        sync_passed_out_students_to_alumni()
        # Join AlumniStudent with Batch to get batch information
        batch_data = db.session.query(
            Batch,
            Course,
            db.func.count(AlumniStudent.id).label('alumni_count')
        ).join(
            AlumniStudent, 
            Batch.id == AlumniStudent.batch_id
        ).join(
            Course,
            Batch.course_id == Course.id
        ).group_by(
            Batch.id, 
            Course.id
        ).all()
        
        result = []
        for batch, course, count in batch_data:
            result.append({
                'batch_id': batch.id,
                'course_name': get_display_course_name(batch, course.name),
                'start_year': batch.start_year,
                'end_year': batch.end_year,
                'alumni_count': count
            })
        
        return jsonify({
            'success': True,
            'data': result
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/alumni/search', methods=['GET'])
def api_search_alumni():
    """Search alumni by various criteria"""
    try:
        sync_passed_out_students_to_alumni()
        # Get query parameters
        search_term = request.args.get('search', '').strip()
        department = request.args.get('department', '').strip()
        batch_year = request.args.get('batch_year', '').strip()
        batch_id = request.args.get('batch_id', '').strip()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # Start building the query - explicitly specify the joins
        query = db.session.query(AlumniStudent, Batch, Course).select_from(AlumniStudent).join(Batch, AlumniStudent.batch_id == Batch.id).join(Course, Batch.course_id == Course.id)
        
        # Apply filters
        if search_term:
            search_filter = (
                AlumniStudent.name.ilike(f'%{search_term}%') |
                AlumniStudent.admission_number.ilike(f'%{search_term}%') |
                AlumniStudent.email.ilike(f'%{search_term}%')
            )
            query = query.filter(search_filter)
        
        if department:
            query = query.filter(AlumniStudent.department.ilike(f'%{department}%'))
        
        if batch_id:
            try:
                query = query.filter(Batch.id == int(batch_id))
            except ValueError:
                pass
        elif batch_year:
            try:
                batch_year_int = int(batch_year)
                query = query.filter(Batch.end_year == batch_year_int)
            except ValueError:
                pass  # Invalid year, ignore filter
        
        # Paginate results
        total = query.count()
        alumni_data = query.offset((page - 1) * per_page).limit(per_page).all()
        
        # Format results
        result = []
        for alum, batch, course in alumni_data:
            result.append({
                'id': alum.id,
                'admission_number': alum.admission_number,
                'name': alum.name,
                'email': alum.email,
                'department': alum.department,
                'course_name': get_display_course_name(batch, course.name) if batch and course else 'N/A',
                'batch_start_year': batch.start_year if batch else None,
                'batch_end_year': batch.end_year if batch else None,
                'passout_year': alum.passout_year,
                'created_at': alum.created_at.strftime('%Y-%m-%d') if alum.created_at else None
            })
        
        return jsonify({
            'success': True,
            'data': result,
            'pagination': {
                'current_page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/alumni/<string:admission_number>/mentor-notes', methods=['GET'])
def api_get_alumni_mentor_notes(admission_number):
    """Admin-only alumni note archive view."""
    try:
        alumni = AlumniStudent.query.filter_by(admission_number=admission_number.upper()).first()
        if not alumni:
            return jsonify({'success': False, 'message': 'Alumni record not found'}), 404

        notes = db.session.query(MentorPrivateNote, Faculty).outerjoin(
            Faculty, Faculty.id == MentorPrivateNote.mentor_id
        ).filter(
            MentorPrivateNote.student_admission_number == admission_number.upper(),
            MentorPrivateNote.transferred_to_admin == True
        ).order_by(MentorPrivateNote.created_at.desc()).all()

        result = []
        for note, mentor in notes:
            result.append({
                'id': note.id,
                'student_id': note.student_admission_number,
                'mentor_id': note.mentor_id,
                'mentor_name': mentor.name if mentor else 'Unknown mentor',
                'session_id': note.session_id,
                'note_type': note.note_type,
                'content': note.content,
                'visibility': note.visibility,
                'transferred_to_admin': bool(note.transferred_to_admin),
                'transferred_at': note.transferred_at.isoformat() if note.transferred_at else None,
                'created_at': note.created_at.isoformat() if note.created_at else None,
                'updated_at': note.updated_at.isoformat() if note.updated_at else None,
            })

        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/alumni/department/<department>/batches', methods=['GET'])
def api_get_alumni_by_department_batches(department):
    """Get batches for a specific department with alumni counts"""
    try:
        sync_passed_out_students_to_alumni()
        # Get batches for a specific department with alumni counts
        batch_data = db.session.query(
            Batch,
            db.func.count(AlumniStudent.id).label('alumni_count')
        ).join(
            AlumniStudent, 
            Batch.id == AlumniStudent.batch_id
        ).filter(
            AlumniStudent.department.ilike(f'%{department}%')
        ).group_by(
            Batch.id
        ).all()
        
        result = []
        for batch, count in batch_data:
            result.append({
                'batch_id': batch.id,
                'start_year': batch.start_year,
                'end_year': batch.end_year,
                'alumni_count': count,
                'status': batch.status
            })
        
        return jsonify({
            'success': True,
            'data': result
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/alumni/department/<department>/batch/<int:batch_id>', methods=['GET'])
def api_get_alumni_by_department_batch(department, batch_id):
    """Get alumni for a specific department and batch"""
    try:
        sync_passed_out_students_to_alumni()
        # Get alumni for specific department and batch
        alumni = AlumniStudent.query.join(Batch).filter(
            AlumniStudent.department.ilike(f'%{department}%'),
            AlumniStudent.batch_id == batch_id
        ).all()
        
        result = []
        for alum in alumni:
            result.append({
                'id': alum.id,
                'admission_number': alum.admission_number,
                'name': alum.name,
                'email': alum.email,
                'passout_year': alum.passout_year,
                'created_at': alum.created_at.strftime('%Y-%m-%d') if alum.created_at else None
            })
        
        return jsonify({
            'success': True,
            'data': result
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


from routes.admin_routes import admin_bp
from routes.auth_routes import auth_bp
from routes.student_routes import student_bp
from routes.mentor_routes import mentor_bp
from routes.subject_handler_routes import subject_handler_bp
from routes.hod_routes import hod_bp
from routes.schedule_routes import schedule_bp
from routes.ai_performance_routes import ai_performance_bp
from sqlalchemy import text


app.register_blueprint(admin_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(student_bp)
app.register_blueprint(mentor_bp)
app.register_blueprint(subject_handler_bp)
app.register_blueprint(hod_bp)
app.register_blueprint(schedule_bp)
app.register_blueprint(ai_performance_bp)

# Smart Adaptive Study Planner
from routes.smart_planner_routes import init_smart_planner
init_smart_planner(app)

# ═══════════════════════════════════════════════════════
# MENTORING SESSION ROUTES  (Student + Mentor)
# ═══════════════════════════════════════════════════════

# Slot definitions

SYSTEM_SLOTS = ["09:00", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00"]   # 9 AM – 5 PM
MENTOR_SLOTS = ["17:00", "18:00"]                                                            # 5 PM – 7 PM
ALL_SLOTS    = SYSTEM_SLOTS + MENTOR_SLOTS


def _is_mentor_busy_from_timetable(mentor_id, weekday_name, slot_hour):
    """Check if mentor has a class in the timetable at the given weekday and hour."""
    faculty = Faculty.query.get(mentor_id)
    if not faculty:
        return False
    # Map slot hour string → period number (rough mapping: period 1=9h, 2=10h, …)
    try:
        hour = int(slot_hour.split(":")[0])
        period = hour - 8   # period 1 starts at 9 h
    except Exception:
        return False

    entry = Timetable.query.filter_by(
        handler_id=mentor_id,
        day=weekday_name,
        period=period
    ).first()
    return entry is not None


def _slot_available(mentor_id, date_obj, slot):
    """Returns True when the slot is available for booking."""
    weekday = date_obj.strftime("%A")   # e.g. "Monday"

    # 1. Mentor leave check
    leaves = MentorLeave.query.filter_by(mentor_id=mentor_id, leave_date=date_obj).all()
    for leave in leaves:
        if leave.from_time is None:          # whole-day leave
            return False
        try:
            lf = int(leave.from_time.split(":")[0])
            lt = int(leave.to_time.split(":")[0])
            sh = int(slot.split(":")[0])
            if lf <= sh < lt:
                return False
        except Exception:
            pass

    # 2. For system slots (9-17): check timetable clash
    if slot in SYSTEM_SLOTS:
        if _is_mentor_busy_from_timetable(mentor_id, weekday, slot):
            return False

    # 3. Already booked (Pending or Approved)
    existing = MentoringSession.query.filter_by(
        mentor_id=mentor_id, date=date_obj, time_slot=slot
    ).filter(MentoringSession.status.in_(["Pending", "Approved"])).first()
    if existing:
        return False

    return True


def _session_datetime(session_date, time_slot):
    """Combine session date and slot time for past/upcoming checks."""
    time_str = (time_slot or "00:00").split('-')[0].strip()
    try:
        hour_str, minute_str = time_str.split(':', 1)
        hour = int(hour_str)
        minute = int(minute_str)
    except Exception:
        hour = 0
        minute = 0
    return datetime.combine(session_date, datetime.min.time()).replace(hour=hour, minute=minute)


# ─── Student: get my mentor info ───────────────────────
@app.route('/api/student/my-mentor/<admission_number>', methods=['GET'])
def api_student_my_mentor(admission_number):
    try:
        student = Student.query.get(admission_number)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        if not student.mentor_id:
            return jsonify({'success': True, 'data': None, 'message': 'No mentor assigned'}), 200
        mentor = Faculty.query.get(student.mentor_id)
        if not mentor:
            return jsonify({'success': True, 'data': None}), 200
        return jsonify({'success': True, 'data': {
            'id': mentor.id,
            'name': mentor.name,
            'email': mentor.email,
            'designation': mentor.designation,
            'department': mentor.department,
        }}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Student: get available slots for a date ───────────
@app.route('/api/session/available-slots', methods=['GET'])
def api_available_slots():
    """
    Query params: mentor_id, date (YYYY-MM-DD)
    Returns list of { slot, slot_type, available }
    Slot is available if mentor is not on leave, no timetable clash (for system slots),
    and no existing Pending/Approved booking.
    Date must be tomorrow → today + 28 days.
    """
    try:
        mentor_id = request.args.get('mentor_id', type=int)
        date_str  = request.args.get('date', '')
        if not mentor_id or not date_str:
            return jsonify({'success': False, 'message': 'mentor_id and date are required'}), 400

        from datetime import date as date_type, timedelta
        try:
            date_obj = date_type.fromisoformat(date_str)
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format'}), 400

        today = date_type.today()
        min_date = today + timedelta(days=1)
        max_date = today + timedelta(days=28)
        if not (min_date <= date_obj <= max_date):
            return jsonify({'success': False, 'message': 'Date must be between tomorrow and 28 days from today'}), 400

        result = []
        for slot in ALL_SLOTS:
            slot_type = 'mentor' if slot in MENTOR_SLOTS else 'system'
            avail = _slot_available(mentor_id, date_obj, slot)
            result.append({'slot': slot, 'slot_type': slot_type, 'available': avail})

        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Student: book a session ───────────────────────────
@app.route('/api/session/book', methods=['POST'])
def api_book_session():
    """
    Body: { admission_number, mentor_id, date, time_slot, session_type, notes }
    System slots (9-17) are auto-approved; Mentor slots (17-19) stay Pending until mentor approves.
    """
    try:
        data = request.get_json()
        admission_number = data.get('admission_number')
        mentor_id   = data.get('mentor_id')
        date_str    = data.get('date')
        time_slot   = data.get('time_slot')
        session_type = data.get('session_type', 'Online')
        notes       = data.get('notes', '')

        if not all([admission_number, mentor_id, date_str, time_slot]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400

        from datetime import date as date_type, timedelta
        date_obj = date_type.fromisoformat(date_str)
        today    = date_type.today()
        if not (today + timedelta(days=1) <= date_obj <= today + timedelta(days=28)):
            return jsonify({'success': False, 'message': 'Invalid booking date'}), 400

        if time_slot not in ALL_SLOTS:
            return jsonify({'success': False, 'message': 'Invalid time slot'}), 400

        if not _slot_available(int(mentor_id), date_obj, time_slot):
            return jsonify({'success': False, 'message': 'Slot is not available'}), 409

        slot_type = 'mentor' if time_slot in MENTOR_SLOTS else 'system'
        # System slots: auto-approve; Mentor slots stay Pending
        status = 'Approved' if slot_type == 'system' else 'Pending'

        session = MentoringSession(
            student_admission_number=admission_number,
            mentor_id=int(mentor_id),
            date=date_obj,
            time_slot=time_slot,
            slot_type=slot_type,
            session_type=session_type,
            status=status,
            notes=notes,
        )
        db.session.add(session)
        db.session.commit()

        return jsonify({'success': True, 'message': f'Session booked. Status: {status}', 'data': {
            'id': session.id, 'status': session.status, 'date': date_str, 'time_slot': time_slot
        }}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Student: list my sessions ─────────────────────────
@app.route('/api/session/student/<admission_number>', methods=['GET'])
def api_student_sessions(admission_number):
    """List all mentoring sessions for a student — uses raw SQL to match actual DB columns."""
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            rows = conn.execute(text(
                "SELECT ms.id, ms.date, ms.time, ms.type, ms.status, "
                "ms.meeting_link, ms.remarks, ms.mentor_id, ms.absence_reason, ms.attendance_marked_at, "
                "f.name as mentor_name "
                "FROM mentoring_sessions ms "
                "LEFT JOIN faculty f ON f.id = ms.mentor_id "
                "WHERE ms.student_admission_number = :adm "
                "ORDER BY ms.date DESC, ms.time DESC"
            ), {'adm': admission_number}).fetchall()

        result = []
        for r in rows:
            result.append({
                'id': r[0],
                'date': r[1].isoformat() if hasattr(r[1], 'isoformat') else str(r[1]),
                'time_slot': r[2] or '',
                'slot_type': 'mentor',
                'session_type': r[3] or 'Offline',
                'status': r[4] or 'Pending',
                'meeting_link': r[5] or '',
                'notes': r[6] or '',
                'mentor_name': r[10] or 'Your Mentor',
                'absence_reason': r[8] or '',
                'attendance_marked_at': r[9].isoformat() if r[9] and hasattr(r[9], 'isoformat') else (str(r[9]) if r[9] else None),
                'calendar_link': '',
            })
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Student: cancel or request reschedule ─────────────────────────
@app.route('/api/session/<int:session_id>/cancel', methods=['POST'])
def api_cancel_session(session_id):
    """Student cancels pending session OR requests reschedule for approved session with new date/time preference."""
    try:
        data = request.get_json()
        admission_number = data.get('admission_number')
        reason = data.get('reason', '')  # Required for approved sessions
        preferred_date = data.get('preferred_date')  # Optional: suggested new date
        preferred_time = data.get('preferred_time')  # Optional: suggested new time
        
        session = MentoringSession.query.get(session_id)
        if not session:
            return jsonify({'success': False, 'message': 'Session not found'}), 404
        if session.student_admission_number != admission_number:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        if session.status == 'Cancelled':
            return jsonify({'success': False, 'message': 'Already cancelled'}), 400
        
        # For approved sessions: create reschedule request
        if session.status == 'Approved':
            if not reason:
                return jsonify({'success': False, 'message': 'Reason is required for reschedule requests'}), 400
            
            # Build reschedule request note
            reschedule_note = f'[Reschedule Requested by Student: {reason}]'
            if preferred_date:
                reschedule_note += f' Preferred Date: {preferred_date}'
            if preferred_time:
                reschedule_note += f' Preferred Time: {preferred_time}'
            
            # Append to existing notes
            session.notes = (session.notes or '') + '\n' + reschedule_note
            session.status = 'Pending'  # Back to pending for mentor approval
            message = 'Reschedule request sent to mentor'
        else:
            # For pending sessions: just cancel
            session.status = 'Cancelled'
            message = 'Session cancelled'
        
        db.session.commit()
        return jsonify({'success': True, 'message': message}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Mentor: list pending/upcoming sessions ─────────────
@app.route('/api/session/mentor/<int:mentor_id>', methods=['GET'])
def api_mentor_sessions(mentor_id):
    try:
        from datetime import date as date_type
        status_filter = request.args.get('status')  # optional filter
        q = MentoringSession.query.filter_by(mentor_id=mentor_id)
        if status_filter:
            q = q.filter_by(status=status_filter)
        sessions = q.order_by(MentoringSession.date.asc(), MentoringSession.time_slot.asc()).all()

        result = []
        for s in sessions:
            student_name = s.student.full_name if s.student else s.student_admission_number
            result.append({
                'id': s.id,
                'date': s.date.isoformat(),
                'time_slot': s.time_slot,
                'slot_type': s.slot_type,
                'session_type': s.session_type,
                'status': s.status,
                'student_name': student_name,
                'student_id': s.student_admission_number,
                'notes': s.notes,
                'meeting_link': s.meeting_link,
                'absence_reason': s.absence_reason,
                'attendance_marked_at': s.attendance_marked_at.isoformat() if s.attendance_marked_at else None,
            })
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Mentor: approve / reject / reschedule a session ─────────────────
@app.route('/api/session/<int:session_id>/respond', methods=['POST'])
def api_respond_session(session_id):
    """Mentor approves/rejects/cancels a session. Also handles reschedule requests from students."""
    try:
        data = request.get_json()
        action       = data.get('action')           # 'approve' | 'reject' | 'cancel' | 'reschedule'
        mentor_id    = data.get('mentor_id')
        meeting_link = data.get('meeting_link', '')
        new_date     = data.get('date')             # For reschedule action
        new_time     = data.get('time_slot')        # For reschedule action
        message      = data.get('message', '')      # Optional message

        session = MentoringSession.query.get(session_id)
        if not session:
            return jsonify({'success': False, 'message': 'Session not found'}), 404
        if session.mentor_id != int(mentor_id):
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        if action == 'approve':
            session.status = 'Approved'
            if meeting_link:
                session.meeting_link = meeting_link
        elif action == 'reject':
            session.status = 'Rejected'
        elif action == 'cancel':
            session.status = 'Cancelled'
        elif action == 'reschedule':
            # Mentor accepts student's reschedule request with new time
            if new_date:
                session.date = new_date
            if new_time:
                session.time_slot = new_time
            if message:
                session.notes = (session.notes or '') + f'\n[Mentor Response: {message}]'
            session.status = 'Approved'
        else:
            return jsonify({'success': False, 'message': 'Invalid action'}), 400

        db.session.commit()
        return jsonify({'success': True, 'message': f'Session {action}d'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Mentor: mark leave / unavailability ─────────────────
@app.route('/api/session/<int:session_id>/status', methods=['POST'])
def api_session_mark_status(session_id):
    """Mentor records post-session attendance. Absence requires a reason."""
    try:
        data = request.get_json() or {}
        status = (data.get('status') or '').strip()
        mentor_id = data.get('mentor_id')
        absence_reason = (data.get('absence_reason') or '').strip()

        if status not in ['Attended', 'Absent']:
            return jsonify({'success': False, 'message': 'Invalid status'}), 400

        session = MentoringSession.query.get(session_id)
        if not session:
            return jsonify({'success': False, 'message': 'Session not found'}), 404
        if mentor_id is not None and str(session.mentor_id) != str(mentor_id):
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        if session.status != 'Approved':
            return jsonify({'success': False, 'message': 'Only approved sessions can be marked for attendance'}), 400
        if _session_datetime(session.date, session.time_slot) > datetime.now():
            return jsonify({'success': False, 'message': 'Attendance can be recorded only after the session time has passed'}), 400
        if status == 'Absent' and not absence_reason:
            return jsonify({'success': False, 'message': 'Absence reason is required'}), 400

        session.status = status
        session.attendance_marked_at = datetime.utcnow()
        session.absence_reason = absence_reason if status == 'Absent' else None
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Session marked successfully',
            'data': {
                'id': session.id,
                'status': session.status,
                'absence_reason': session.absence_reason,
                'attendance_marked_at': session.attendance_marked_at.isoformat() if session.attendance_marked_at else None,
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/leave', methods=['POST'])
def api_mentor_mark_leave():
    try:
        data = request.get_json()
        mentor_id  = data.get('mentor_id')
        date_str   = data.get('date')
        from_time  = data.get('from_time')    # None → whole day
        to_time    = data.get('to_time')
        reason     = data.get('reason', '')

        if not mentor_id or not date_str:
            return jsonify({'success': False, 'message': 'mentor_id and date are required'}), 400

        from datetime import date as date_type
        date_obj = date_type.fromisoformat(date_str)

        leave = MentorLeave(
            mentor_id=int(mentor_id),
            leave_date=date_obj,
            from_time=from_time,
            to_time=to_time,
            reason=reason,
        )
        db.session.add(leave)

        # Auto-cancel all Pending/Approved sessions on that date that fall within the leave window
        sessions = MentoringSession.query.filter_by(
            mentor_id=int(mentor_id), date=date_obj
        ).filter(MentoringSession.status.in_(["Pending", "Approved"])).all()

        cancelled = 0
        for s in sessions:
            if from_time is None:
                s.status = 'Cancelled'
                cancelled += 1
            else:
                try:
                    lf = int(from_time.split(":")[0])
                    lt = int(to_time.split(":")[0])
                    sh = int(s.time_slot.split(":")[0])
                    if lf <= sh < lt:
                        s.status = 'Cancelled'
                        cancelled += 1
                except Exception:
                    pass

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Leave marked. {cancelled} session(s) auto-cancelled.'
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Mentor: list / delete leaves ────────────────────────
@app.route('/api/mentor/<int:mentor_id>/leaves', methods=['GET'])
def api_mentor_leaves(mentor_id):
    try:
        from datetime import date as date_type
        leaves = MentorLeave.query.filter_by(mentor_id=mentor_id).order_by(MentorLeave.leave_date.asc()).all()
        result = [{
            'id': l.id,
            'date': l.leave_date.isoformat(),
            'from_time': l.from_time,
            'to_time': l.to_time,
            'reason': l.reason,
        } for l in leaves]
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/leave/<int:leave_id>', methods=['DELETE'])
def api_delete_mentor_leave(leave_id):
    try:
        leave = MentorLeave.query.get(leave_id)
        if not leave:
            return jsonify({'success': False, 'message': 'Leave not found'}), 404
        db.session.delete(leave)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Leave removed'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500




# ════════════════════════════════════════════════════════════════
# AI INSIGHTS ROUTES
# Priority: Gemini 2.0 Flash (free) → Gemini Flash Lite → Smart fallback
# ════════════════════════════════════════════════════════════════

_GEMINI_MODELS = ['gemini-1.5-flash', 'gemini-1.5-flash-8b', 'gemini-1.5-pro', 'gemini-2.0-flash-exp']

def _is_valid_key(key: str, bad_fragments=()) -> bool:
    """Returns True if key looks like a real API key."""
    if not key or len(key) < 20:
        return False
    low = key.lower()
    generics = ('placeholder', 'your-', 'your_', 'example', 'change-me', 'api-key-here')
    if any(b in low for b in generics + bad_fragments):
        return False
    return True


def _get_gemini_client():
    """Return a working (client, model_name) tuple or (None, None)."""
    key = os.getenv('GEMINI_API_KEY', '')
    if not _is_valid_key(key):
        return None, None
    try:
        from google import genai as _genai
        client = _genai.Client(api_key=key)
        return client, 'gemini-2.0-flash'
    except Exception:
        return None, None


# ════════════════════════════════════════════════════════════════


def _gemini_complete(client, model: str, system_prompt: str, user_prompt: str,
                     history=None, max_tokens: int = 800) -> str:
    """Call Gemini new SDK. Retries with lighter model on rate-limit."""
    from google import genai as _genai
    from google.genai import types as _gtypes

    # Build content list - Gemini requires alternating user/model roles
    contents = []
    last_role = None
    for h in (history or [])[-10:]:
        role = 'user' if h.get('role') == 'user' else 'model'
        if role == last_role: continue # Avoid consecutive same role
        contents.append(_gtypes.Content(role=role, parts=[_gtypes.Part(text=h.get('content', ''))]))
        last_role = role
    
    if last_role == 'user':
        # If last was user, we can't send another user immediately without a model response
        # In simple chat, we just skip history or append a small dummy model response if needed
        # But usually we want to ensure history ends with 'model' so current 'user' works
        pass 
    
    contents.append(_gtypes.Content(role='user', parts=[_gtypes.Part(text=user_prompt)]))

    config = _gtypes.GenerateContentConfig(
        system_instruction=system_prompt,
        max_output_tokens=max_tokens,
        temperature=0.7,
    )

    models_to_try = [model] + [m for m in _GEMINI_MODELS if m != model]
    last_err = None
    for m in models_to_try:
        try:
            resp = client.models.generate_content(model=m, contents=contents, config=config)
            return resp.text
        except Exception as e:
            err_str = str(e)
            # Rate-limit → try next model
            if '429' in err_str or 'quota' in err_str.lower() or 'rate' in err_str.lower():
                last_err = e
                continue
            raise  # Other errors bubble up
    raise last_err or RuntimeError("All Gemini models failed")
def _get_groq_client():
    """Return a working groq key or None."""
    key = os.getenv('GROQ_API_KEY', '')
    if not _is_valid_key(key):
        return None, None
    return 'api_key_only', 'llama-3.3-70b-versatile'


def _groq_complete(client, model: str, system_prompt: str, user_prompt: str,
                   history=None, max_tokens: int = 800) -> str:
    """Call Groq API using requests."""
    import requests
    key = os.getenv('GROQ_API_KEY')
    headers = {
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json"
    }
    messages = [{"role": "system", "content": system_prompt}]
    for h in (history or []):
        messages.append({"role": h.get("role"), "content": h.get("content")})
    messages.append({"role": "user", "content": user_prompt})

    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers=headers,
        json={
            "model": model,
            "messages": messages,
            "max_tokens": max_tokens,
            "temperature": 0.7
        },
        timeout=15
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]



def _get_ai_provider():
    """Returns ('groq', (client, model)), ('gemini', (client, model)) | (None, None)."""
    # Prefer Groq if available
    g_client, g_model = _get_groq_client()
    if g_client:
        return 'groq', (g_client, g_model)

    client, model = _get_gemini_client()
    if client:
        return 'gemini', (client, model)
    return None, None


def _ai_complete(provider, client_tuple, system_prompt: str, user_prompt: str,
                 history=None, max_tokens: int = 800) -> str:
    """Unified completion — wraps Groq and Gemini."""
    if provider == 'groq':
        client, model = client_tuple
        return _groq_complete(client, model, system_prompt, user_prompt, history, max_tokens)
    if provider == 'gemini':
        client, model = client_tuple
        return _gemini_complete(client, model, system_prompt, user_prompt, history, max_tokens)
    raise ValueError("No AI provider available")


# Backward-compat stub (used by older fallback paths)
def _get_openai_client():
    return None



def _build_student_context(admission_number: str) -> dict:
    """Gather all student metrics and return a structured context dict."""
    student = Student.query.get(admission_number)
    if not student:
        return {}

    ctx: dict = {
        "name": student.full_name,
        "admission_number": admission_number,
        "department": student.branch or "N/A",
        "batch": student.batch or "N/A",
    }

    dept_rows = Timetable.query.filter(Timetable.department == student.branch).all() if student.branch else []
    current_timetable = _filter_timetable_by_batch(dept_rows, student.batch) if student.batch else []
    current_subjects = sorted({row.subject for row in current_timetable if row.subject})
    ctx["current_subjects"] = current_subjects

    # Analytics
    sa = StudentAnalytics.query.filter_by(student_id=admission_number).first()
    if sa:
        ctx["attendance_pct"]     = round(sa.attendance_percentage, 1)
        ctx["attendance_trend"]   = "improving" if sa.attendance_slope > 0.02 else ("declining" if sa.attendance_slope < -0.05 else "stable")
        ctx["avg_marks"]          = round(sa.avg_internal_marks, 1)
        ctx["marks_trend"]        = "improving" if sa.marks_slope > 5 else ("declining" if sa.marks_slope < -5 else "stable")
        ctx["failure_count"]      = sa.failure_count
        ctx["risk_score"]         = round(sa.risk_score, 1)
        ctx["risk_level"]         = "critical" if sa.risk_score > 70 else ("at-risk" if sa.risk_score > 40 else "stable")
        ctx["academic_status"]    = sa.status

    # Subject-level marks
    from models import StudentMark
    marks = StudentMark.query.filter_by(student_id=admission_number).all()
    subject_marks: dict = {}
    for m in marks:
        # Use existing university mark if available, else average internals
        if m.university_mark is not None:
            subject_marks[m.subject_code] = m.university_mark
        else:
            vals = [v for v in [m.internal1, m.internal2, m.internal3] if v is not None]
            if vals:
                subject_marks[m.subject_code] = round(sum(vals)/len(vals), 1)
    if current_subjects:
        ctx["subject_marks"] = {s: subject_marks.get(s, 0) for s in current_subjects}
    else:
        ctx["subject_marks"] = subject_marks

    # Attendance per subject
    atts = Attendance.query.filter_by(student_admission_number=admission_number).all()
    ctx["subject_attendance"] = {a.subject_name: round(a.percentage, 1) for a in atts if a.percentage}

    progress_row = db.session.execute(text(
        "SELECT * FROM sp_progress WHERE student_id = :sid ORDER BY updated_at DESC, id DESC LIMIT 1"
    ), {"sid": admission_number}).fetchone()
    if progress_row:
        progress = dict(progress_row._mapping)
        ctx["study_plan_compliance"] = round(float(progress.get("compliance_score") or 0), 1)
        ctx["study_plan_status"] = progress.get("status_label")
        try:
            ctx["study_plan_progress_insights"] = json.loads(progress.get("generated_insights") or "[]")
        except Exception:
            ctx["study_plan_progress_insights"] = []

    plan_row = db.session.execute(text(
        "SELECT * FROM sp_plans WHERE student_id = :sid ORDER BY updated_at DESC, id DESC LIMIT 1"
    ), {"sid": admission_number}).fetchone()
    if plan_row:
        plan_data = dict(plan_row._mapping)
        try:
            raw_plan = json.loads(plan_data.get("raw_json") or "{}")
        except Exception:
            raw_plan = {}
        focus_subjects = raw_plan.get("focusSubjects") or []
        if focus_subjects:
            ctx["study_plan_subjects"] = [str(item.get("subject", "")).strip() for item in focus_subjects if str(item.get("subject", "")).strip()]
        ctx["study_plan_goal"] = raw_plan.get("weeklyGoal") or plan_data.get("weekly_goal")
        ctx["study_plan_alerts"] = raw_plan.get("alerts") or []

    if "study_plan_compliance" not in ctx:
        plan = WeeklyStudyPlan.query.filter_by(student_id=admission_number).order_by(WeeklyStudyPlan.id.desc()).first()
        if plan:
            total    = sum(s.allocated_hours for s in plan.subjects) if plan.subjects else 0
            done     = sum(s.completed_hours for s in plan.subjects) if plan.subjects else 0
            ctx["study_plan_compliance"] = round((done / total * 100) if total else 0, 1)
            ctx["study_plan_subjects"]   = [s.subject.name for s in plan.subjects if getattr(s, "subject", None)]

    return ctx


def _context_to_system_prompt(ctx: dict) -> str:
    marks_str = ", ".join(f"{s}: {v}/100" for s, v in ctx.get("subject_marks", {}).items()) or "no data"
    attend_str = ", ".join(f"{s}: {v}%" for s, v in ctx.get("subject_attendance", {}).items()) or "no data"
    compliance = ctx.get("study_plan_compliance", "N/A")
    subjects_str = ", ".join(ctx.get("study_plan_subjects", [])) or "none"
    current_subjects = ", ".join(ctx.get("current_subjects", [])) or "no timetable subjects found"

    return f"""You are MentAi — a highly empathetic, intelligent academic advisor embedded in a student academic management system.

STUDENT PROFILE (use this to give personalised, data-driven advice):
- Name: {ctx.get('name', 'Student')}
- Department: {ctx.get('department', 'N/A')}
- Batch: {ctx.get('batch', 'N/A')}
- Overall Attendance: {ctx.get('attendance_pct', 'N/A')}% (trend: {ctx.get('attendance_trend', 'N/A')})
- Average Internal Marks: {ctx.get('avg_marks', 'N/A')}/100 (trend: {ctx.get('marks_trend', 'N/A')})
- Subject Marks: {marks_str}
- Subject Attendance: {attend_str}
- Risk Score: {ctx.get('risk_score', 'N/A')}/100 ({ctx.get('risk_level', 'N/A')})
- Academic Status: {ctx.get('academic_status', 'N/A')}
- Study Plan Compliance: {compliance}%
- Current Timetable Subjects: {current_subjects}
- Current Study Plan Subjects: {subjects_str}

INSTRUCTIONS:
- Always refer to the student by name when appropriate.
- Ground every recommendation in the actual data above.
- Give detailed explanations with enough reasoning for the student to understand what to do next.
- Use the current timetable subjects as the source of truth for subject advice.
- If asked for a study plan, produce a structured weekly plan in Markdown with time blocks, breaks, and catch-up steps.
- If asked a general academic question, answer it clearly.
- Never hallucinate data — if something is N/A, say so honestly.
- Format responses in clean Markdown with short headers, bullets, and a clear next-action checklist.
- Aim for 250-450 words unless the student asks for a shorter answer.
"""


# ─── Auto-generate insights snapshot ─────────────────────────────
def _subject_performance_snapshot(row):
    mark_value = _subject_combined_score(row)
    if mark_value is None:
        internals = [value for value in (row.internal1, row.internal2, row.internal3) if value is not None]
        if not internals:
            return None, "not_started"
        mark_value = round(sum(internals) / len(internals), 1)

    if mark_value < 35:
        return mark_value, "failed"
    if mark_value < 50:
        return mark_value, "at_risk"
    if mark_value < 75:
        return mark_value, "ongoing"
    return mark_value, "strong"


def _resolve_subject_support_contacts(student, current_timetable):
    support_map = {}
    for row in current_timetable:
        subject_name = str(row.subject or '').strip()
        if not subject_name:
            continue

        support_map.setdefault(subject_name, {
            'handler_id': row.handler_id,
            'handler_name': row.handler.name if getattr(row, 'handler', None) else (row.handler_name or 'Subject Handler'),
            'slots': [],
        })
        if row.day and row.time_slot:
            support_map[subject_name]['slots'].append(f"{row.day} {row.time_slot}")

    fallback_handlers = Faculty.query.filter(
        Faculty.is_subject_handler == True,
        Faculty.status == 'Live'
    ).all()
    dept_fallback = next((faculty for faculty in fallback_handlers if _departments_match_simple(faculty.department, student.branch)), None)

    for item in support_map.values():
        if not item.get('handler_id') and dept_fallback:
            item['handler_id'] = dept_fallback.id
            item['handler_name'] = dept_fallback.name

    return support_map


def _build_student_support_report(admission_number: str) -> dict:
    student = Student.query.get(str(admission_number or '').strip().upper())
    if not student:
        return {}

    dept_rows = Timetable.query.filter(Timetable.department == student.branch).all() if student.branch else []
    current_timetable = _filter_timetable_by_batch(dept_rows, student.batch) if student.batch else []
    current_subjects = sorted({row.subject for row in current_timetable if row.subject})
    support_contacts = _resolve_subject_support_contacts(student, current_timetable)

    marks_rows = StudentMark.query.filter_by(student_id=student.admission_number).order_by(StudentMark.semester.asc(), StudentMark.id.asc()).all()
    attendance_rows = Attendance.query.filter_by(student_admission_number=student.admission_number).all()
    analytics = StudentAnalytics.query.filter_by(student_id=student.admission_number).first()
    cgpa_metrics = _calculate_student_cgpa_metrics(student.admission_number)

    attendance_map = {}
    for row in attendance_rows:
        key = str(row.subject_code or row.subject_name or '').strip()
        if key:
            attendance_map[key] = round(float(row.percentage or 0), 1)
        if row.subject_name:
            attendance_map.setdefault(str(row.subject_name).strip(), round(float(row.percentage or 0), 1))

    note_candidates = [
        note for note in PlaygroundNote.query.order_by(PlaygroundNote.created_at.desc()).all()
        if _playground_visible_for_student(note, student)
    ]

    subject_rows = {}
    subject_scores = []
    semester_scores = {}
    for row in marks_rows:
        subject_code = str(row.subject_code or '').strip()
        if not subject_code:
            continue
        subject_rows[subject_code] = row
        mark_value, _ = _subject_performance_snapshot(row)
        if mark_value is not None:
            subject_scores.append(mark_value)
            if row.semester is not None:
                semester_scores.setdefault(int(row.semester), []).append(mark_value)

    failed_subjects = []
    improving_areas = []
    attention_areas = []
    ongoing_areas = []
    available_notes = []
    seen_notes = set()
    focus_now = []
    today_plan_subjects = []
    weak_subjects_strategy = []
    failure_analysis = []
    subject_analysis = []

    for subject_code, row in subject_rows.items():
        mark_value, status = _subject_performance_snapshot(row)
        attendance_value = attendance_map.get(subject_code, attendance_map.get(subject_code.replace('_', ' ')))
        support = support_contacts.get(subject_code, {
            'handler_id': None,
            'handler_name': 'Subject Handler',
            'slots': [],
        })
        related_notes = [
            note for note in note_candidates
            if str(note.subject_code or '').strip().lower() == subject_code.lower()
        ]

        for note in related_notes[:2]:
            if note.id in seen_notes:
                continue
            seen_notes.add(note.id)
            available_notes.append({
                'id': note.id,
                'subject': subject_code,
                'title': note.title,
                'description': note.description,
                'download_url': f"/static/{note.file_path}" if note.file_path else None,
                'uploaded_by_name': note.uploader.name if note.uploader else 'Subject Handler',
                'created_at': note.created_at.isoformat() if note.created_at else None,
            })

        entry = {
            'subject': subject_code,
            'mark': mark_value,
            'combined_score': mark_value,
            'internal_score': _normalized_internal_score(row),
            'university_score': _valid_metric(row.university_mark),
            'attendance_pct': attendance_value,
            'handler_id': support.get('handler_id'),
            'handler_name': support.get('handler_name') or 'Subject Handler',
            'class_slots': support.get('slots', [])[:3],
            'notes_available': len(related_notes),
        }
        classification = _classification_for_score(mark_value)
        entry['classification'] = classification
        subject_analysis.append(entry.copy())

        if classification in ('critical', 'weak'):
            base_problem = "concept issue"
            if attendance_value is not None and attendance_value < 75 and (mark_value or 0) < 40:
                base_problem = "exposure problem"
            elif attendance_value is not None and attendance_value >= 75 and (mark_value or 0) < 40:
                base_problem = "understanding problem"
            elif (entry.get('internal_score') or 0) >= 70 and (entry.get('university_score') or 0) < 45:
                base_problem = "exam writing issue"
            elif (entry.get('internal_score') or 0) < 40 and (entry.get('university_score') or 0) < 40:
                base_problem = "concept issue"

            weak_subjects_strategy.append({
                'subject': subject_code,
                'problem': base_problem,
                'solution': f"Study 1 concept in {subject_code}, write a short summary, then solve 5 questions before ending the session.",
            })

        if status == 'failed':
            failed_subjects.append({
                **entry,
                'recommended_hours': 5 if attendance_value is None or attendance_value >= 75 else 6,
                'quick_fix': 'Rebuild one unit at a time, then solve previous questions without notes.',
                'pass_strategy': [
                    'Start with the last failed unit and write a one-page concept sheet.',
                    'Solve five previous or model questions after each study block.',
                    'Ask the subject handler for a short doubt-clearing session before the next internal.',
                ],
            })
            focus_now.append({
                'subject': subject_code,
                'reason': f"{subject_code} is urgent because the current score is {(mark_value if mark_value is not None else 'N/A')}/100 and it is in the failed range.",
                'action_today': f"Complete one concept-repair block and one practice block for {subject_code} today.",
            })
            failure_analysis.append({
                'subject': subject_code,
                'reason': weak_subjects_strategy[-1]['problem'] if weak_subjects_strategy else "concept issue",
                'fix': f"Rebuild the most recent weak unit in {subject_code}, then test yourself with 5 questions today.",
            })
            today_plan_subjects.append({'subject': subject_code, 'classification': classification})
            attention_areas.append({
                'title': subject_code,
                'detail': f"Needs backlog recovery support{f' ({mark_value}/100)' if mark_value is not None else ''}.",
            })
        elif status == 'strong':
            improving_areas.append({
                'title': subject_code,
                'detail': f"Strong academic control{f' at {mark_value}/100' if mark_value is not None else ''}. Keep it warm with light revision.",
            })
        elif status in ('ongoing', 'at_risk'):
            ongoing_areas.append({
                'title': subject_code,
                'detail': f"Needs steady follow-up{f' ({mark_value}/100)' if mark_value is not None else ''} so it does not slip into backlog.",
            })
            if classification == 'critical':
                focus_now.append({
                    'subject': subject_code,
                    'reason': f"{subject_code} is below the safe score band.",
                    'action_today': f"Do one focused concept-learning and recall session for {subject_code} today.",
                })
                today_plan_subjects.append({'subject': subject_code, 'classification': classification})
            elif classification == 'weak':
                today_plan_subjects.append({'subject': subject_code, 'classification': classification})

    attendance_trend = "stable"
    marks_trend = "stable"
    if analytics:
        attendance_trend = "improving" if analytics.attendance_slope > 0.02 else ("declining" if analytics.attendance_slope < -0.05 else "stable")
        marks_trend = "improving" if analytics.marks_slope > 5 else ("declining" if analytics.marks_slope < -5 else "stable")

    overall_attendance = round(float(analytics.attendance_percentage), 1) if analytics else round(
        sum(float(row.percentage or 0) for row in attendance_rows) / len(attendance_rows), 1
    ) if attendance_rows else 0.0
    avg_marks = round(float(analytics.avg_internal_marks), 1) if analytics else (round(sum(subject_scores) / len(subject_scores), 1) if subject_scores else 0.0)
    semester_average_scores = {
        semester: round(sum(values) / len(values), 2)
        for semester, values in semester_scores.items()
        if values
    }
    semester_trend = _trend_from_values([semester_average_scores[key] for key in sorted(semester_average_scores.keys())])

    progress_row = db.session.execute(text(
        "SELECT * FROM sp_progress WHERE student_id = :sid ORDER BY updated_at DESC, id DESC LIMIT 1"
    ), {'sid': student.admission_number}).fetchone()
    progress = dict(progress_row._mapping) if progress_row else {}

    plan_row = db.session.execute(text(
        "SELECT * FROM sp_plans WHERE student_id = :sid ORDER BY updated_at DESC, id DESC LIMIT 1"
    ), {'sid': student.admission_number}).fetchone()
    plan = dict(plan_row._mapping) if plan_row else {}

    sessions = []
    if plan.get('id'):
        session_rows = db.session.execute(text(
            "SELECT * FROM sp_sessions WHERE plan_id = :pid ORDER BY day, planned_start"
        ), {'pid': plan['id']}).fetchall()
        sessions = [dict(row._mapping) for row in session_rows]

    tracked_sessions = [session for session in sessions if session.get('status') != 'not_started']
    completed_sessions = [session for session in sessions if session.get('status') == 'completed']
    skipped_sessions = [session for session in sessions if session.get('status') == 'skipped']
    in_progress_sessions = [session for session in sessions if session.get('status') == 'in_progress']

    compliance_score = round(float(progress.get('compliance_score') or 0), 1)
    planner_component = compliance_score if tracked_sessions else (50.0 if sessions else 40.0)
    cgpa_value = float(cgpa_metrics.get('cgpa') or 0)
    cgpa_component = round((cgpa_value / 10) * 100, 1) if cgpa_value else (avg_marks if avg_marks else 0)

    overall_score = round(
        (overall_attendance * 0.25) +
        (avg_marks * 0.35) +
        (cgpa_component * 0.20) +
        (planner_component * 0.20),
        1,
    )
    overall_status = (
        'Strong Recovery Zone' if overall_score >= 80 else
        'Recoverable With Consistency' if overall_score >= 60 else
        'Needs Structured Support'
    )
    strict_status = (
        "On Track" if overall_score >= 80 and overall_attendance >= 75 and not failed_subjects else
        "Needs Support" if overall_score >= 60 and overall_attendance >= 75 else
        "Critical" if overall_attendance >= 65 else
        "Seriously Behind"
    )
    attendance_alert = (
        f"Critical alert: attendance is {overall_attendance}%. Attend every class this week, collect missed notes, and meet your mentor." if overall_attendance < 65 else
        f"Warning: attendance is {overall_attendance}%. Protect every class this week to stay exam-eligible." if overall_attendance < 75 else
        f"Attendance is {overall_attendance}%. Keep this safe buffer."
    )

    if overall_attendance >= 85:
        improving_areas.insert(0, {'title': 'Attendance', 'detail': f'Attendance buffer is healthy at {overall_attendance}%.'})
    elif overall_attendance < 75:
        attention_areas.insert(0, {'title': 'Attendance', 'detail': f'Attendance is {overall_attendance}%. Protect every class until you cross the safe line.'})
    else:
        ongoing_areas.insert(0, {'title': 'Attendance', 'detail': f'Attendance is {overall_attendance}%. Keep building a safer buffer.'})

    if compliance_score >= 70 and tracked_sessions:
        improving_areas.append({'title': 'Study Plan Follow-through', 'detail': f'Compliance score is {compliance_score}/100 with real logged sessions.'})
    elif skipped_sessions:
        attention_areas.append({'title': 'Skipped Sessions', 'detail': f'{len(skipped_sessions)} planned session(s) were skipped and need recovery blocks.'})
        for item in today_plan_subjects:
            if item['subject'] not in [entry['subject'] for entry in failure_analysis]:
                failure_analysis.append({
                    'subject': item['subject'],
                    'reason': 'discipline problem',
                    'fix': f"Complete at least one logged session for {item['subject']} today so the recovery plan becomes real progress.",
                })
    elif sessions and not tracked_sessions:
        ongoing_areas.append({'title': 'Study Plan Activation', 'detail': 'The plan is ready, but it still needs the first completed session to become data-backed.'})

    recommended_actions = []
    if failed_subjects:
        recommended_actions.append(f"Clear the first backlog subject, {failed_subjects[0]['subject']}, with two recovery blocks this week before adding extra new topics.")
    if overall_attendance < 75:
        recommended_actions.append('Attend every class this week and collect any missed notes on the same day.')
    if skipped_sessions:
        recommended_actions.append('Use one weekend recovery slot for each skipped session instead of rewriting the whole plan.')
    if not available_notes:
        recommended_actions.append('Request personalized notes from the subject handler for the weakest subject so revision starts from the right material.')
    recommended_actions.append('Log each study session honestly so the score reflects real recovery, not just a generated plan.')

    seen_focus = set()
    dedup_focus = []
    for item in focus_now:
        key = item['subject']
        if key in seen_focus:
            continue
        seen_focus.add(key)
        dedup_focus.append(item)
    focus_now = dedup_focus[:3]

    if not today_plan_subjects:
        today_plan_subjects = [{'subject': item.get('subject'), 'classification': item.get('classification')} for item in subject_analysis if item.get('classification') in ('critical', 'weak')][:3]
    if not today_plan_subjects:
        today_plan_subjects = [{'subject': item.get('subject'), 'classification': item.get('classification')} for item in subject_analysis[:3]]

    seen_plan_subjects = set()
    ordered_plan_subjects = []
    for item in today_plan_subjects:
        subject = item.get('subject')
        if not subject or subject in seen_plan_subjects:
            continue
        seen_plan_subjects.add(subject)
        ordered_plan_subjects.append(item)
    today_plan = _build_today_plan(ordered_plan_subjects)

    behavior_notes = []
    if skipped_sessions:
        behavior_notes.append("You are skipping planned sessions, so discipline is affecting recovery.")
    if failed_subjects:
        behavior_notes.append("Failed subjects must be scheduled before easier topics.")
    if compliance_score < 50:
        behavior_notes.append("Your study pattern is inconsistent; reduce load and complete only 2-3 high-priority sessions today.")
    behavior_correction = " ".join(behavior_notes) if behavior_notes else "Your current behavior is stable. Keep logging sessions and protect priority subjects first."

    motivation = (
        f"You are behind due to low marks and attendance, but recovery is possible if you complete {min(2, max(1, len(today_plan)))} focused sessions today."
        if strict_status in ("Critical", "Seriously Behind")
        else "You are still recoverable. Stay consistent with today's priority sessions and the numbers will move."
    )

    upcoming_remedials = RemedialClass.query.filter(
        RemedialClass.student_id == student.admission_number,
        RemedialClass.scheduled_date >= date.today() - timedelta(days=1)
    ).order_by(RemedialClass.scheduled_date.asc()).all()

    return {
        'student_id': student.admission_number,
        'student_name': student.full_name,
        'mentor_id': student.mentor_id,
        'current_subjects': current_subjects,
        'summary': {
            'overall_score': overall_score,
            'overall_status': overall_status,
            'status': strict_status,
            'attendance_pct': overall_attendance,
            'attendance_trend': attendance_trend,
            'avg_marks': avg_marks,
            'marks_trend': marks_trend,
            'cgpa': cgpa_metrics.get('cgpa'),
            'sgpa': cgpa_metrics.get('sgpa'),
            'semester_trend': semester_trend,
            'risk_score': round(float(getattr(analytics, 'adjusted_risk', 0) or getattr(analytics, 'risk_score', 0) or 0), 1) if analytics else 0,
            'planner_score': compliance_score,
            'planner_status': progress.get('status_label') or 'Not Started',
            'completed_sessions': len(completed_sessions),
            'skipped_sessions': len(skipped_sessions),
            'in_progress_sessions': len(in_progress_sessions),
            'planned_sessions': len(sessions),
            'failed_subject_count': len(failed_subjects),
            'available_note_count': len(available_notes),
        },
        'failed_subjects': failed_subjects,
        'improving_areas': improving_areas[:5],
        'attention_areas': attention_areas[:5],
        'ongoing_areas': ongoing_areas[:5],
        'recommended_actions': recommended_actions[:5],
        'available_notes': available_notes[:6],
        'attendance_alert': attendance_alert,
        'focus_now': focus_now,
        'today_plan': today_plan,
        'weak_subjects_strategy': weak_subjects_strategy[:6],
        'failure_analysis': failure_analysis[:6],
        'behavior_correction': behavior_correction,
        'motivation': motivation,
        'analysis': {
            'status': strict_status,
            'attendance_alert': attendance_alert,
            'focus_now': focus_now,
            'today_plan': today_plan,
            'weak_subjects_strategy': weak_subjects_strategy[:6],
            'failure_analysis': failure_analysis[:6],
            'behavior_correction': behavior_correction,
            'motivation': motivation,
            'subject_analysis': subject_analysis,
        },
        'support_contacts': [
            {
                'subject': subject,
                'handler_id': info.get('handler_id'),
                'handler_name': info.get('handler_name'),
                'class_slots': info.get('slots', [])[:3],
            }
            for subject, info in support_contacts.items()
        ],
        'upcoming_remedials': [{
            'id': item.id,
            'subject': item.subject_code,
            'title': item.title,
            'scheduled_date': item.scheduled_date.isoformat(),
            'time_slot': item.time_slot,
            'mode': item.mode,
            'meeting_link': item.meeting_link,
            'handler_name': item.handler.name if item.handler else 'Subject Handler',
        } for item in upcoming_remedials],
    }


@app.route('/api/ai/insights/<string:admission_number>', methods=['GET'])
def api_ai_insights(admission_number):
    """Return accurate rule-based cards grounded in live student data."""
    ctx = _build_student_context(admission_number)
    report = _build_student_support_report(admission_number)

    if not ctx or not report:
        return jsonify({'success': False, 'message': 'Student not found'}), 404

    insights = _fallback_insights(ctx, report)
    return jsonify({'success': True, 'data': insights, 'source': 'rule-based'}), 200


def _fallback_insights(ctx: dict) -> list:
    """Rule-based insights when OpenAI is not available."""
    cards = []
    att   = ctx.get('attendance_pct', 75)
    marks = ctx.get('avg_marks', 60)
    risk  = ctx.get('risk_score', 30)
    marks_dict   = ctx.get('subject_marks', {})
    attend_dict  = ctx.get('subject_attendance', {})
    compliance   = ctx.get('study_plan_compliance', 100)

    # Attendance card
    if att < 75:
        cards.append({"title": "Critical Attendance Alert", "body": f"Your overall attendance is {att}% — below the 75% minimum requirement. You risk being debarred from exams. Attend every possible class this week.", "type": "warning", "icon": "⚠️"})
    elif att < 85:
        cards.append({"title": "Attendance Needs Attention", "body": f"Attendance at {att}% is acceptable but leaves little room for absence. Aim for 90%+ to stay safe for semester-end.", "type": "info", "icon": "📊"})
    else:
        cards.append({"title": "Excellent Attendance!", "body": f"Great job! Your {att}% attendance shows strong commitment. Keep this momentum going.", "type": "success", "icon": "✅"})

    # Marks card
    worst = sorted(marks_dict.items(), key=lambda x: x[1])[:1]
    if worst:
        s, v = worst[0]
        cards.append({"title": f"Focus on {s}", "body": f"Your score in {s} ({v}/100) needs improvement. Dedicate extra study time to this subject before the next internal exam.", "type": "tip", "icon": "📚"})
    elif marks < 60:
        cards.append({"title": "Marks Below Average", "body": f"Average of {marks}/100 is below the passing threshold. Seek help from your subject handler and visit office hours.", "type": "warning", "icon": "⚠️"})
    else:
        cards.append({"title": "Solid Academic Performance", "body": f"Your average marks of {marks}/100 reflect good understanding. Push harder in weaker subjects to maximize your CGPA.", "type": "success", "icon": "✅"})

    # Risk card
    if risk > 60:
        cards.append({"title": "High Risk — Act Now", "body": f"Your risk score of {risk}/100 is concerning. Book a mentoring session immediately and follow the weekly study plan consistently.", "type": "warning", "icon": "⚠️"})
    elif risk > 35:
        cards.append({"title": "Monitor Your Progress", "body": f"Risk score of {risk}/100 puts you in the 'at-risk' zone. Regular mentor check-ins and consistent study sessions will help you improve.", "type": "info", "icon": "📊"})
    else:
        cards.append({"title": "Low Risk — Stay Sharp", "body": f"Your risk score of {risk}/100 is excellent. Maintain your discipline and help peers who may be struggling.", "type": "success", "icon": "🎯"})

    # Study plan card
    if compliance < 50:
        cards.append({"title": "Study Plan Compliance Low", "body": f"You've completed only {compliance}% of your weekly study goals. Set aside 1–2 hours each evening to catch up and log your sessions.", "type": "tip", "icon": "💡"})
    else:
        best = sorted(marks_dict.items(), key=lambda x: x[1], reverse=True)[:1]
        if best:
            s, v = best[0]
            cards.append({"title": f"Strength in {s}", "body": f"Your {v}/100 in {s} is your strongest subject. Leverage this confidence to tackle harder subjects with the same energy.", "type": "success", "icon": "💡"})
        else:
            cards.append({"title": "Keep Following Your Plan", "body": "Your study plan compliance is good. Consistent effort each week compounds into excellent semester results.", "type": "success", "icon": "🎯"})

    return cards[:4]


# ─── Chat / Q&A endpoint  ────────────────────────────────────────
@app.route('/api/student/support-report/<string:admission_number>', methods=['GET'])
def api_student_support_report(admission_number):
    report = _build_student_support_report(admission_number)
    if not report:
        return jsonify({'success': False, 'message': 'Student not found'}), 404
    return jsonify({'success': True, 'data': report}), 200


@app.route('/api/student/support-actions/<string:admission_number>', methods=['POST'])
def api_student_support_action(admission_number):
    try:
        sid = str(admission_number or '').strip().upper()
        student = Student.query.get(sid)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        payload = request.get_json(force=True) or {}
        action_type = str(payload.get('action_type') or '').strip().lower()
        subject = str(payload.get('subject') or '').strip()
        handler_id = payload.get('handler_id')
        custom_message = str(payload.get('message') or '').strip()
        preferred_date = str(payload.get('preferred_date') or '').strip()
        preferred_time = str(payload.get('preferred_time') or '').strip()

        if action_type not in {'session_request', 'notes_request'}:
            return jsonify({'success': False, 'message': 'Invalid action_type'}), 400
        if not subject:
            return jsonify({'success': False, 'message': 'subject is required'}), 400
        if not handler_id:
            return jsonify({'success': False, 'message': 'handler_id is required'}), 400

        handler = Faculty.query.get(int(handler_id))
        if not handler or not handler.is_subject_handler:
            return jsonify({'success': False, 'message': 'Subject handler not found'}), 404

        preferred_slot = " ".join([item for item in [preferred_date, preferred_time] if item]).strip()
        if action_type == 'session_request':
            category = 'Study Session Request'
            message = custom_message or (
                f"{student.full_name} is requesting a personalized study support session for {subject}."
                f"{f' Preferred slot: {preferred_slot}.' if preferred_slot else ''}"
                " Please suggest a remedial session or a focused recovery plan."
            )
            success_message = 'Study session request sent to the subject handler'
        else:
            category = 'Notes Request'
            message = custom_message or (
                f"{student.full_name} is requesting focused notes or revision material for {subject}."
                " Please share class notes, model questions, or a short reading sequence."
            )
            success_message = 'Notes request sent to the subject handler'

        row = SubjectHandlerMessage(
            student_id=sid,
            handler_id=int(handler_id),
            subject=subject,
            category=category,
            message=message,
            sender_role='student',
            status='open',
            is_read=False,
        )
        db.session.add(row)

        if student.mentor_id:
            db.session.add(MentorMessage(
                mentor_id=student.mentor_id,
                student_id=sid,
                message=f"[Student support request] {student.full_name} requested {category.lower()} for {subject}.",
                sender_role='student',
                is_read=False,
            ))

        db.session.add(Notification(
            student_id=sid,
            title='Support request sent',
            message=success_message,
            type='student_support_request',
            is_read=False,
        ))
        db.session.commit()
        return jsonify({'success': True, 'message': success_message, 'data': {'id': row.id}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


def _fallback_insights(ctx: dict, report: dict | None = None) -> list:
    report = report or {}
    cards = []
    att = ctx.get('attendance_pct', 75)
    marks = ctx.get('avg_marks', 60)
    risk = ctx.get('risk_score', 30)
    marks_dict = ctx.get('subject_marks', {})
    compliance = ctx.get('study_plan_compliance', 100)
    summary = report.get('summary', {})
    failed_subjects = report.get('failed_subjects', [])
    recommended_actions = report.get('recommended_actions', [])

    if failed_subjects:
        top_failed = failed_subjects[0]
        cards.append({
            "title": f"Backlog: {top_failed['subject']}",
            "body": f"You still need to clear {top_failed['subject']}. Give it {top_failed['recommended_hours']} focused hours this week, solve previous questions, and request support from {top_failed['handler_name']}.",
            "type": "warning",
            "icon": "📘",
        })

    if att < 75:
        cards.append({"title": "Attendance Needs Recovery", "body": f"Your overall attendance is {att}%. Stay present in every possible class this week so weak subjects do not become harder to recover.", "type": "warning", "icon": "⚠️"})
    elif att < 85:
        cards.append({"title": "Attendance Buffer Thin", "body": f"Attendance is {att}%. You are safe for now, but you need a bigger buffer so recovery sessions do not turn into class misses.", "type": "info", "icon": "📊"})
    else:
        cards.append({"title": "Attendance Is Helping", "body": f"Attendance at {att}% is supporting your recovery. Keep the same class discipline while fixing the weaker subjects.", "type": "success", "icon": "✅"})

    worst = sorted(marks_dict.items(), key=lambda x: x[1])[:1]
    if worst:
        subject_name, score = worst[0]
        cards.append({"title": f"Focus on {subject_name}", "body": f"Your current performance in {subject_name} is {score}/100. Start with concept repair, then a short practice set, then log the session so the score reflects real work.", "type": "tip", "icon": "🎯"})
    elif marks < 60:
        cards.append({"title": "Marks Need Lift", "body": f"Your average marks are {marks}/100. Build short daily revision blocks before trying long catch-up sessions.", "type": "warning", "icon": "⚠️"})
    else:
        cards.append({"title": "Marks Are Recoverable", "body": f"Your average marks are {marks}/100. Keep strong subjects warm and use extra time on the weakest subject first.", "type": "success", "icon": "📈"})

    if risk > 60 or summary.get('overall_score', 0) < 60:
        cards.append({"title": "Structured Support Needed", "body": f"Your overall recovery score is {summary.get('overall_score', 0)}/100. Use the plan, request subject support early, and avoid letting skipped sessions accumulate.", "type": "warning", "icon": "🧭"})
    elif risk > 35:
        cards.append({"title": "Progress Needs Tracking", "body": f"Your overall recovery score is {summary.get('overall_score', 0)}/100. Consistent session logging and class attendance will move you into a safer zone.", "type": "info", "icon": "📌"})
    else:
        cards.append({"title": "Recovery Is Moving", "body": f"Your current recovery score is {summary.get('overall_score', 0)}/100. Keep the momentum and use one extra revision block for the weakest area.", "type": "success", "icon": "🚀"})

    if compliance < 50:
        action_text = recommended_actions[0] if recommended_actions else "Complete the next session today and mark it honestly."
        cards.append({"title": "Activate The Plan", "body": f"Planner score is {compliance}/100. {action_text}", "type": "tip", "icon": "💡"})
    elif not failed_subjects:
        best = sorted(marks_dict.items(), key=lambda x: x[1], reverse=True)[:1]
        if best:
            subject_name, score = best[0]
            cards.append({"title": f"Strength: {subject_name}", "body": f"Your strongest area right now is {subject_name} at {score}/100. Use that confidence to stay steady while you repair weaker subjects.", "type": "success", "icon": "🌟"})

    return cards[:4]


@app.route('/api/ai/chat', methods=['POST'])
def api_ai_chat():
    """
    Body: { admission_number, message, history: [{role, content}] }
    Returns: { reply }
    """
    data             = request.get_json()
    admission_number = data.get('admission_number', '')
    user_message     = data.get('message', '').strip()
    history          = data.get('history', [])

    if not user_message:
        return jsonify({'success': False, 'message': 'Empty message'}), 400

    ctx              = _build_student_context(admission_number) if admission_number else {}
    provider, ai_client = _get_ai_provider()

    if not provider:
        reply = _fallback_chat(user_message, ctx)
        return jsonify({'success': True, 'reply': reply, 'source': 'rule-based'}), 200

    try:
        system_prompt = _context_to_system_prompt(ctx) if ctx else "You are MentAi, a helpful academic advisor."
        reply = _ai_complete(provider, ai_client, system_prompt, user_message, history=history, max_tokens=1200)
        return jsonify({'success': True, 'reply': reply, 'source': provider}), 200
    except Exception as e:
        reply = _fallback_chat(user_message, ctx)
        return jsonify({'success': True, 'reply': reply, 'source': 'rule-based', 'note': str(e)}), 200


def _fallback_chat(message: str, ctx: dict) -> str:
    """Simple keyword-based chat fallback when OpenAI is unavailable."""
    msg = message.lower()
    name = ctx.get('name', 'there').split()[0] if ctx.get('name') else 'there'
    current_subjects = ctx.get('current_subjects') or ctx.get('study_plan_subjects') or []
    subject_line = ", ".join(current_subjects[:5]) if current_subjects else "your current timetable subjects"

    if any(w in msg for w in ['study plan', 'plan', 'schedule', 'weekly']):
        subjects = ctx.get('study_plan_subjects', [])
        marks_dict = ctx.get('subject_marks', {})
        if marks_dict:
            sorted_subs = sorted(marks_dict.items(), key=lambda x: x[1])
            plan = "\n".join(f"- **{s}**: 2–3 hours (score: {v}/100)" for s, v in sorted_subs)
            return f"""Hi {name}, here is a practical weekly plan based on your current data.

**Main focus:** Start with the lowest-scoring subjects, but keep every current timetable subject active so you do not create a new backlog.

{plan}

**How to execute it**
- On college days, use the first evening block for the subject taught that day.
- Use a second shorter block for assignments, lab records, or pending notes.
- Keep one recovery slot on the weekend for anything missed during the week.
- After each session, log the completed time so your compliance score becomes accurate.

**Today:** Pick one subject from {subject_line}, study for 45-60 minutes, then do 20 minutes of practice questions."""
        return f"""Hi {name}, I do not have enough marks data yet, so use your timetable as the source of truth.

**Plan for this week**
- Rotate through these subjects: {subject_line}.
- Study one class subject the same evening it appears in the timetable.
- Keep 15-minute breaks between blocks so the plan is realistic.
- Reserve the weekend for revision, pending assignments, and self-tests.

Start today with a 60-minute block: 35 minutes concept review, 15 minutes examples, and 10 minutes quick recall."""

    if any(w in msg for w in ['attendance', 'absent', 'class']):
        att = ctx.get('attendance_pct', 'N/A')
        return f"""Your current attendance is **{att}%**. Most colleges treat **75%** as the minimum safe line.

**What this means**
- If you are below 75%, every upcoming class matters because one absence can keep the percentage low.
- If you are above 75%, keep a buffer so illness, travel, or unavoidable events do not create pressure later.

**Action plan**
- Attend all classes for the next two weeks.
- Avoid skipping repeated subjects because those usually affect both attendance and understanding.
- If absences were genuine, document the reason and discuss it with your mentor early."""

    if any(w in msg for w in ['marks', 'score', 'exam', 'cgpa', 'gpa']):
        avg = ctx.get('avg_marks', 'N/A')
        marks_dict = ctx.get('subject_marks', {})
        if marks_dict:
            worst = min(marks_dict.items(), key=lambda x: x[1])
            best  = max(marks_dict.items(), key=lambda x: x[1])
            return f"""Your average internal mark is **{avg}/100**.

**Performance read**
- Best area: **{best[0]}** ({best[1]}/100)
- Needs most attention: **{worst[0]}** ({worst[1]}/100)

**How to improve**
- Spend the next two sessions on **{worst[0]}** before revising stronger subjects.
- First session: rebuild notes and list formulas, definitions, or key steps.
- Second session: solve previous questions without looking at notes.
- After that, teach the topic aloud in 5 minutes. If you get stuck, that is the exact point to revise again."""
        return f"""Your average internal marks are **{avg}/100**.

I do not have clean subject-level marks yet, so use the current timetable subjects: {subject_line}. Pick the subject you feel least confident in and do one focused revision block today. Use this pattern: concept recap, worked example, practice question, quick self-test."""

    if any(w in msg for w in ['risk', 'danger', 'failing', 'fail']):
        risk = ctx.get('risk_score', 'N/A')
        level = ctx.get('risk_level', 'N/A')
        return f"""Your current risk score is **{risk}/100** ({level}).

**What drives this score**
- Attendance consistency
- Internal marks and recent drops
- Missed or completed study-plan sessions

**What to do next**
- If risk is high, book a mentor session and take your marks/attendance details with you.
- Complete the next two planned study sessions before trying to catch up everything.
- Focus on timetable subjects first: {subject_line}.
- Log study time honestly so the dashboard can reduce risk only when real progress happens."""

    if any(w in msg for w in ['mentor', 'meeting', 'session', 'book']):
        return f"You can book a mentoring session from the **Mentoring** tab in your dashboard. Sessions from 9 AM–5 PM are auto-approved if your mentor is free. Evening slots (5–7 PM) need mentor confirmation."

    if any(w in msg for w in ['tip', 'advice', 'suggest', 'help', 'improve']):
        att = ctx.get('attendance_pct', 75)
        avg = ctx.get('avg_marks', 60)
        tips = []
        if isinstance(att, float) and att < 80:
            tips.append("📌 Improve attendance — attend every class for the next 2 weeks to raise your percentage.")
        if isinstance(avg, float) and avg < 65:
            tips.append("📚 Spend 1 extra hour daily on your weakest subject.")
        tips.append("🎯 Log your study hours in the dashboard to track weekly plan compliance.")
        tips.append("💬 Book a mentor session at least once a month to discuss progress.")
        return f"Here are some personalised tips for you, {name}:\n\n" + "\n".join(tips)

    # Generic
    return f"""Hi {name}, I can help with attendance, marks, timetable-based study planning, risk level, assignments, and mentor preparation.

To give you the most useful answer, ask something like:
- "What should I study today?"
- "Explain my risk level."
- "Make a recovery plan for missed classes."
- "How do I improve in Cloud Computing?"

Based on what I know right now, your safest next move is to choose one current timetable subject from {subject_line}, study it for 45-60 minutes, and log the session afterward."""


# ─── Generate personalised weekly study plan ─────────────────────
@app.route('/api/ai/study-plan/<string:admission_number>', methods=['GET'])
def api_ai_study_plan(admission_number):
    """Generate a detailed weekly study plan using OpenAI."""
    ctx    = _build_student_context(admission_number)
    client = _get_openai_client()

    if not ctx:
        return jsonify({'success': False, 'message': 'Student not found'}), 404

    if not client:
        plan = _fallback_study_plan(ctx)
        return jsonify({'success': True, 'plan': plan, 'source': 'rule-based'}), 200

    try:
        marks_dict  = ctx.get('subject_marks', {})
        attend_dict = ctx.get('subject_attendance', {})
        marks_str   = "; ".join(f"{s}: {v}/100" for s, v in marks_dict.items()) or "no marks data"
        attend_str  = "; ".join(f"{s}: {v}%" for s, v in attend_dict.items()) or "no attendance data"

        prompt = f"""Create a detailed 7-day weekly study plan in Markdown for a student with the following academic profile. 
Be specific with daily time blocks and subject allocations. Prioritize subjects with low marks.

Name: {ctx.get('name')}
Department: {ctx.get('department')}
Subject Marks: {marks_str}
Subject Attendance: {attend_str}
Risk Level: {ctx.get('risk_level', 'stable')}
Academic Status: {ctx.get('academic_status', 'Stable')}

Format:
## 📅 Weekly Study Plan for {ctx.get('name')}
### Day-by-Day Schedule (Monday–Sunday)
For each day list 2-3 study blocks with subject, duration, and specific topic goal.
### 🎯 Key Focus Areas This Week
### 💡 Study Tips
Keep it motivating and actionable."""

        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
        {"role": "system", "content": f"You are MentAi — an expert academic advisor. {_context_to_system_prompt(ctx)}"},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1200,
            temperature=0.7,
        )
        plan = resp.choices[0].message.content
        return jsonify({'success': True, 'plan': plan, 'source': 'openai'}), 200
    except Exception as e:
        plan = _fallback_study_plan(ctx)
        return jsonify({'success': True, 'plan': plan, 'source': 'rule-based', 'note': str(e)}), 200


def _fallback_study_plan(ctx: dict) -> str:
    marks_dict = ctx.get('subject_marks', {})
    name = ctx.get('name', 'Student')
    sorted_subs = sorted(marks_dict.items(), key=lambda x: x[1]) if marks_dict else []

    day_targets = {}
    days  = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    subs  = [s for s, _ in sorted_subs] or ctx.get('study_plan_subjects', ['General Study'])

    for i, day in enumerate(days):
        primary = subs[i % len(subs)]
        secondary = subs[(i + 1) % len(subs)] if len(subs) > 1 else primary
        day_targets[day] = (primary, secondary)

    plan_lines = [f"## 📅 Weekly Study Plan for {name}\n"]
    for day, (p, s) in day_targets.items():
        if day == "Sunday":
            plan_lines.append(f"### {day}\n- 🔄 **Revision**: Review all subjects covered this week\n- 📝 **Self-test**: 30-minute quiz on weakest topic\n")
        else:
            plan_lines.append(f"### {day}\n- 🌅 **9:00–11:00 AM** — **{p}** (2 hrs) — Focus on theory & past questions\n- 🌇 **5:00–6:00 PM** — **{s}** (1 hr) — Practice problems\n")

    plan_lines.append("### 🎯 Key Focus Areas\n" + "\n".join(f"- **{s}** ({v}/100) — needs most attention" for s, v in sorted_subs[:3]))
    plan_lines.append("\n### 💡 Study Tips\n- Study high-priority subjects in the morning when focus is sharpest.\n- Use the 25-minute Pomodoro technique with 5-minute breaks.\n- Log every session in the Weekly Planner tab.")
    return "\n".join(plan_lines)



# ─────────────────────────────────────────────────────────────────────────────
# MENTOR STUDENT DETAIL + SESSION BOOKING APIs
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/mentor/private-notes/<string:student_id>', methods=['GET'])
def api_get_mentor_private_notes(student_id):
    try:
        mentor_id = request.args.get('mentor_id', type=int)
        if not mentor_id:
            return jsonify({'success': False, 'message': 'mentor_id is required'}), 400

        student_key = student_id.upper()
        student = Student.query.get(student_key)
        alumni = AlumniStudent.query.filter_by(admission_number=student_key).first()

        if student:
            if student.mentor_id != mentor_id:
                return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        elif not alumni or alumni.mentor_id != mentor_id:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        notes = MentorPrivateNote.query.filter_by(
            student_admission_number=student_key,
            mentor_id=mentor_id
        ).order_by(MentorPrivateNote.created_at.desc()).all()

        return jsonify({
            'success': True,
            'data': [{
                'id': note.id,
                'student_id': note.student_admission_number,
                'mentor_id': note.mentor_id,
                'session_id': note.session_id,
                'note_type': note.note_type,
                'content': note.content,
                'visibility': note.visibility,
                'transferred_to_admin': bool(note.transferred_to_admin),
                'created_at': note.created_at.isoformat() if note.created_at else None,
                'updated_at': note.updated_at.isoformat() if note.updated_at else None,
            } for note in notes]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/private-notes', methods=['POST'])
def api_create_mentor_private_note():
    try:
        data = request.get_json() or {}
        mentor_id = data.get('mentor_id')
        student_id = (data.get('student_id') or '').strip().upper()
        content = (data.get('content') or '').strip()
        session_id = data.get('session_id')
        note_type = (data.get('note_type') or 'private').strip().lower()

        if not mentor_id or not student_id or not content:
            return jsonify({'success': False, 'message': 'mentor_id, student_id and content are required'}), 400

        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        if str(student.mentor_id) != str(mentor_id):
            return jsonify({'success': False, 'message': 'This student is not assigned to you'}), 403

        if session_id:
            session_row = MentoringSession.query.get(session_id)
            if not session_row or session_row.student_admission_number != student_id or str(session_row.mentor_id) != str(mentor_id):
                return jsonify({'success': False, 'message': 'Invalid session for this mentor/student pair'}), 400

        if note_type not in ['private', 'session', 'abnormality']:
            note_type = 'private'

        note = MentorPrivateNote(
            student_admission_number=student_id,
            mentor_id=int(mentor_id),
            session_id=session_id,
            note_type=note_type,
            content=content,
            visibility='mentor_only'
        )
        db.session.add(note)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Private mentor note saved',
            'data': {
                'id': note.id,
                'student_id': note.student_admission_number,
                'mentor_id': note.mentor_id,
                'session_id': note.session_id,
                'note_type': note.note_type,
                'content': note.content,
                'visibility': note.visibility,
                'transferred_to_admin': bool(note.transferred_to_admin),
                'created_at': note.created_at.isoformat() if note.created_at else None,
                'updated_at': note.updated_at.isoformat() if note.updated_at else None,
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/student/detail/<string:adm>', methods=['GET'])
def api_student_detail(adm):
    """Full student profile for mentor view. Returns all data if profile completed, else basics."""
    try:
        student = Student.query.get(adm.upper())
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Build photo URL if photo_path exists
        photo_url = None
        if student.photo_path:
            photo_url = f"http://localhost:5000/uploads/{student.photo_path}"

        from academic_utils import calculate_academic_status
        # Basic data always returned
        result = {
            'admission_number': student.admission_number,
            'name': student.full_name,
            'email': student.email,
            'mobile': student.mobile_number,
            'branch': student.branch,
            'batch': student.batch,
            'blood_group': student.blood_group,
            'dob': student.dob.isoformat() if student.dob else None,
            'status': student.status,
            'photo_url': photo_url,
            'profile_completed': student.profile_completed,
            'academic_info': calculate_academic_status(student.batch, student.branch),
        }


        academic_metrics = _calculate_student_cgpa_metrics(adm.upper())

        if not student.academics and (
            academic_metrics['cgpa'] is not None or academic_metrics['sgpa'] is not None
        ):
            student.academics = Academic(
                student_admission_number=student.admission_number,
                cgpa=academic_metrics['cgpa'],
                sgpa=academic_metrics['sgpa'],
            )
            db.session.add(student.academics)
            db.session.commit()
        elif student.academics:
            touched = False
            if (student.academics.cgpa is None or float(student.academics.cgpa or 0) <= 0) and academic_metrics['cgpa'] is not None:
                student.academics.cgpa = academic_metrics['cgpa']
                touched = True
            if (student.academics.sgpa is None or float(student.academics.sgpa or 0) <= 0) and academic_metrics['sgpa'] is not None:
                student.academics.sgpa = academic_metrics['sgpa']
                touched = True
            if touched:
                db.session.commit()

        # Extended data if profile is complete
        if student.profile_completed:
            result['personal'] = {
                'religion': student.religion,
                'diocese': student.diocese,
                'parish': student.parish,
                'caste_category': student.caste_category,
                'permanent_address': student.permanent_address,
                'contact_address': student.contact_address,
            }

            if student.parents:
                p = student.parents
                result['parents'] = {
                    'father_name': p.father_name,
                    'father_profession': p.father_profession,
                    'father_mobile': p.father_mobile,
                    'mother_name': p.mother_name,
                    'mother_profession': p.mother_profession,
                    'mother_mobile': p.mother_mobile,
                }

            if student.guardian:
                g = student.guardian
                result['guardian'] = {
                    'name': g.name,
                    'address': g.address,
                    'mobile': g.mobile_number,
                }

            a = student.academics
            if a or academic_metrics['cgpa'] is not None or academic_metrics['sgpa'] is not None:
                effective_sgpa = a.sgpa if a and a.sgpa is not None and float(a.sgpa or 0) > 0 else academic_metrics['sgpa']
                effective_cgpa = a.cgpa if a and a.cgpa is not None and float(a.cgpa or 0) > 0 else academic_metrics['cgpa']
                result['academics'] = {
                    'school_10th': a.school_10th if a else None,
                    'percentage_10th': a.percentage_10th if a else None,
                    'school_12th': a.school_12th if a else None,
                    'percentage_12th': a.percentage_12th if a else None,
                    'college_ug': a.college_ug if a else None,
                    'percentage_ug': a.percentage_ug if a else None,
                    'sgpa': effective_sgpa,
                    'cgpa': effective_cgpa,
                    'entrance_rank': a.entrance_rank if a else None,
                    'nature_of_admission': a.nature_of_admission if a else None,
                    'semester_sgpa': academic_metrics['semester_sgpa'],
                    'latest_semester': academic_metrics['latest_semester'],
                }

            if student.other_info:
                o = student.other_info
                result['other'] = {
                    'accommodation_type': o.accommodation_type,
                    'staying_with': o.staying_with,
                    'hostel_name': o.hostel_name,
                    'transport_mode': o.transport_mode,
                }

        # Add analytics if available
        analytics = StudentAnalytics.query.filter_by(student_id=adm.upper()).first()
        if analytics:
            result['analytics'] = {
                'attendance_percentage': analytics.attendance_percentage or 0,
                'avg_internal_marks': analytics.avg_internal_marks or 0,
                'risk_score': analytics.risk_score or 0,
                'adjusted_risk': analytics.adjusted_risk or 0,
                'status': analytics.status,
                'failure_count': analytics.failure_count or 0,
            }

        # Recent sessions — using raw SQL to match actual DB columns
        from sqlalchemy import text
        with db.engine.connect() as conn:
            rows = conn.execute(text(
                "SELECT id, date, time, type, status, meeting_link, remarks "
                "FROM mentoring_sessions WHERE student_admission_number=:adm "
                "ORDER BY date DESC LIMIT 5"
            ), {'adm': adm.upper()}).fetchall()

        result['recent_sessions'] = [{
            'id': r[0],
            'date': r[1].isoformat() if hasattr(r[1], 'isoformat') else str(r[1]),
            'time_slot': r[2],
            'session_type': r[3],
            'status': r[4],
            'meeting_link': r[5],
            'notes': r[6],
        } for r in rows]

        return jsonify({'success': True, 'data': result}), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/session/book', methods=['POST'])
def api_mentor_book_session():
    """Mentor books a session for a specific mentee. Supports online (GMeet) and offline modes."""
    try:
        data = request.get_json() or {}

        mentor_id = data.get('mentor_id')
        student_id = data.get('student_id', '').strip().upper()
        session_date = data.get('date')          # "YYYY-MM-DD"
        time_slot = data.get('time_slot', '10:00')  # "HH:MM"
        session_type = data.get('session_type', 'Offline')  # Online / Offline
        meeting_link = data.get('meeting_link', '').strip()
        notes = data.get('notes', '').strip()

        if not all([mentor_id, student_id, session_date]):
            return jsonify({'success': False, 'message': 'mentor_id, student_id and date are required'}), 400

        # Verify mentee belongs to this mentor
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        if str(student.mentor_id) != str(mentor_id):
            return jsonify({'success': False, 'message': 'This student is not assigned to you'}), 403

        # Validate GMeet link for online sessions
        if session_type == 'Online' and meeting_link:
            if not (meeting_link.startswith('http://') or meeting_link.startswith('https://')):
                meeting_link = 'https://' + meeting_link

        from datetime import date as date_obj
        parsed_date = date_obj.fromisoformat(session_date)

        # Check for duplicate on same date + mentor
        from sqlalchemy import text
        with db.engine.connect() as conn:
            existing = conn.execute(text(
                "SELECT id FROM mentoring_sessions WHERE student_admission_number=:sid "
                "AND mentor_id=:mid AND date=:dt LIMIT 1"
            ), {'sid': student_id, 'mid': int(mentor_id), 'dt': session_date}).fetchone()

        if existing:
            return jsonify({
                'success': False,
                'message': f'A session already exists for {student.full_name} on {session_date}'
            }), 409

        # Insert using raw SQL to match actual columns: time, type, remarks
        with db.engine.begin() as conn:
            conn.execute(text(
                "INSERT INTO mentoring_sessions "
                "(student_admission_number, mentor_id, date, time, type, status, meeting_link, remarks) "
                "VALUES (:sid, :mid, :dt, :tm, :tp, 'Approved', :ml, :rem)"
            ), {
                'sid': student_id,
                'mid': int(mentor_id),
                'dt': parsed_date,
                'tm': time_slot,
                'tp': session_type,
                'ml': meeting_link if session_type == 'Online' else None,
                'rem': notes or None,
            })

        # Notify student
        try:
            notification = Notification(
                student_id=student_id,
                title='Mentoring Session Scheduled',
                message=f'Your mentor has scheduled a {session_type} session on {session_date} at {time_slot}.'
                        + (f' Join here: {meeting_link}' if meeting_link else ''),
                type='session',
            )
            db.session.add(notification)
            db.session.commit()
        except Exception:
            pass  # Notification failure should not block session creation

        return jsonify({
            'success': True,
            'message': f'{session_type} session booked successfully for {student.full_name}',
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/session/update/<int:session_id>', methods=['PUT'])
def api_mentor_update_session(session_id):
    """Mentor updates an existing session (reschedule, change mode, update meet link, add notes)."""
    try:
        data = request.get_json() or {}
        mentor_id = data.get('mentor_id')

        from sqlalchemy import text
        with db.engine.connect() as conn:
            row = conn.execute(text(
                "SELECT id, mentor_id FROM mentoring_sessions WHERE id=:sid LIMIT 1"
            ), {'sid': session_id}).fetchone()

        if not row:
            return jsonify({'success': False, 'message': 'Session not found'}), 404
        if str(row[1]) != str(mentor_id):
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        # Build dynamic SET clause
        updates = {}
        if 'date' in data:
            updates['date'] = data['date']
        if 'time_slot' in data:
            updates['time'] = data['time_slot']        # actual col = time
        if 'session_type' in data:
            updates['type'] = data['session_type']     # actual col = type
        if 'meeting_link' in data:
            link = data['meeting_link'].strip()
            if link and not (link.startswith('http://') or link.startswith('https://')):
                link = 'https://' + link
            updates['meeting_link'] = link or None
        if 'notes' in data:
            updates['remarks'] = data['notes']         # actual col = remarks
        if 'status' in data:
            updates['status'] = data['status']

        if updates:
            set_clause = ', '.join(f"{k}=:{k}" for k in updates)
            updates['sid'] = session_id
            with db.engine.begin() as conn:
                conn.execute(text(
                    f"UPDATE mentoring_sessions SET {set_clause} WHERE id=:sid"
                ), updates)

        return jsonify({'success': True, 'message': 'Session updated successfully'}), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/sessions/<int:mentor_id>', methods=['GET'])
def api_mentor_sessions_list(mentor_id):
    """List all sessions booked by a mentor, with student info and photo."""
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            rows = conn.execute(text(
                "SELECT ms.id, ms.student_admission_number, ms.date, ms.time, ms.type, "
                "ms.status, ms.meeting_link, ms.remarks, "
                "s.full_name, s.photo_path "
                "FROM mentoring_sessions ms "
                "LEFT JOIN students s ON s.admission_number = ms.student_admission_number "
                "WHERE ms.mentor_id=:mid "
                "ORDER BY ms.date DESC LIMIT 50"
            ), {'mid': mentor_id}).fetchall()

        result = []
        for r in rows:
            photo_url = None
            if r[9]:  # photo_path
                photo_url = f"http://localhost:5000/uploads/{r[9]}"
            result.append({
                'id': r[0],
                'student_id': r[1],
                'student_name': r[8] or 'Unknown',
                'student_photo': photo_url,
                'date': r[2].isoformat() if hasattr(r[2], 'isoformat') else str(r[2]),
                'time_slot': r[3],
                'session_type': r[4],
                'status': r[5],
                'meeting_link': r[6],
                'notes': r[7],
            })

        return jsonify({'success': True, 'data': result}), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500



# ─────────────────────────────────────────────────────────────────────────────
# STUDENT NOTIFICATION APIs
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/student/notifications/<string:admission_number>', methods=['GET'])
def api_student_notifications(admission_number):
    """Get all notifications for a student, newest first."""
    try:
        notifs = Notification.query.filter_by(
            student_id=admission_number
        ).order_by(Notification.created_at.desc()).limit(50).all()

        data = [{
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'type': n.type,
            'is_read': n.is_read,
            'created_at': n.created_at.isoformat() if n.created_at else None,
        } for n in notifs]

        unread_count = sum(1 for n in notifs if not n.is_read)
        return jsonify({'success': True, 'data': data, 'unread_count': unread_count}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/student/notifications/<int:notif_id>/read', methods=['PATCH'])
def api_mark_notification_read(notif_id):
    """Mark a single notification as read."""
    try:
        notif = Notification.query.get(notif_id)
        if not notif:
            return jsonify({'success': False, 'message': 'Not found'}), 404
        notif.is_read = True
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/student/notifications/<string:admission_number>/read-all', methods=['PATCH'])
def api_mark_all_notifications_read(admission_number):
    """Mark all notifications as read for a student."""
    try:
        Notification.query.filter_by(
            student_id=admission_number, is_read=False
        ).update({'is_read': True})
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── Public marks endpoint for React frontend ─────────────────────────────────
def _extract_subject_title(subject_name):
    raw_name = str(subject_name or '').strip()
    if not raw_name:
        return ''

    title = re.sub(r'^[A-Z]{2,5}\d{3,6}[A-Z]?\s*-\s*', '', raw_name).strip()
    title = re.sub(r'\s*\([^)]+\)\s*$', '', title).strip()
    return title or raw_name


def _resolve_subject_catalog_entry(subject_code, semester=None, student=None):
    normalized_code = str(subject_code or '').strip().upper()
    if not normalized_code:
        return {'subject_code': '', 'course_name': '', 'display_name': ''}

    query = Subject.query.join(Course, Subject.course_id == Course.id).join(Semester, Subject.semester_id == Semester.id)
    query = query.filter(Subject.name.ilike(f'{normalized_code}%'))

    if student and getattr(student, 'branch', None):
        preferred_rows = query.filter(Course.name == str(student.branch).strip()).all()
        if semester is not None:
            semester_label = f'Semester {int(semester)}'
            semester_rows = [row for row in preferred_rows if getattr(row.semester, 'name', None) == semester_label]
            if semester_rows:
                preferred_rows = semester_rows
        if preferred_rows:
            subject_row = preferred_rows[0]
            course_name = _extract_subject_title(subject_row.name)
            return {
                'subject_code': normalized_code,
                'course_name': course_name,
                'display_name': course_name or normalized_code,
            }

    subject_row = query.first()
    if subject_row:
        course_name = _extract_subject_title(subject_row.name)
        return {
            'subject_code': normalized_code,
            'course_name': course_name,
            'display_name': course_name or normalized_code,
        }

    return {
        'subject_code': normalized_code,
        'course_name': normalized_code,
        'display_name': normalized_code,
    }


def _calculate_current_semester_for_student(student):
    try:
        from academic_utils import calculate_academic_status
        academic_info = calculate_academic_status(student.batch, student.branch)
        semester = academic_info.get('current_semester')
        return int(semester) if semester else None
    except Exception:
        return None


def _average_numeric(values):
    nums = []
    for value in values:
        metric = _valid_metric(value)
        if metric is not None:
            nums.append(metric)
    return round(sum(nums) / len(nums), 2) if nums else None


def _subject_code_from_subject_name(subject_name):
    raw_name = str(subject_name or '').strip().upper()
    match = re.match(r'^([A-Z]{2,5}\d{3,6}[A-Z]?)', raw_name)
    return match.group(1) if match else raw_name


def _catalog_subjects_for_student(student, semester):
    semester_label = f'Semester {int(semester)}'
    query = Subject.query.join(Course, Subject.course_id == Course.id).join(Semester, Subject.semester_id == Semester.id)
    query = query.filter(
        Course.name == str(student.branch).strip(),
        Semester.name == semester_label,
    ).order_by(Subject.name.asc())
    return query.all()


def _build_semester_progression_payload(student):
    student_id = student.admission_number.upper()
    marks_rows = StudentMark.query.filter_by(student_id=student_id).all()
    attendance_rows = Attendance.query.filter_by(student_admission_number=student_id).all()
    current_semester = _calculate_current_semester_for_student(student)

    semester_subjects = {}

    def ensure_subject(semester, subject_code, fallback_name=''):
        if semester is None:
            return None
        sem_key = int(semester)
        normalized_code = str(subject_code or fallback_name or f'SEM{sem_key}').strip().upper()
        semester_subjects.setdefault(sem_key, {})
        if normalized_code not in semester_subjects[sem_key]:
            subject_meta = _resolve_subject_catalog_entry(normalized_code, semester=sem_key, student=student)
            semester_subjects[sem_key][normalized_code] = {
                'subject_code': normalized_code,
                'course_name': fallback_name or subject_meta.get('course_name') or subject_meta.get('display_name') or normalized_code,
                'internal1': None,
                'internal2': None,
                'internal3': None,
                'internal_avg': None,
                'attendance_percentage': None,
                'attended_classes': None,
                'total_classes': None,
                'university_mark': None,
                'combined_score': None,
            }
        return semester_subjects[sem_key][normalized_code]

    max_semester = current_semester or max(
        [int(row.semester) for row in marks_rows if row.semester is not None] +
        [int(row.semester) for row in attendance_rows if row.semester is not None] +
        [1]
    )

    for semester in range(1, max_semester + 1):
        semester_subjects.setdefault(semester, {})
        for catalog_subject in _catalog_subjects_for_student(student, semester):
            subject_code = _subject_code_from_subject_name(catalog_subject.name)
            course_name = _extract_subject_title(catalog_subject.name)
            ensure_subject(semester, subject_code, course_name)

    for row in marks_rows:
        if current_semester is not None and row.semester is not None and int(row.semester) > int(current_semester):
            continue
        subject = ensure_subject(row.semester, row.subject_code)
        if not subject:
            continue
        subject['internal1'] = _valid_metric(row.internal1)
        subject['internal2'] = _valid_metric(row.internal2)
        subject['internal3'] = _valid_metric(row.internal3)
        subject['internal_avg'] = _normalized_internal_score(row)
        subject['university_mark'] = _valid_metric(row.university_mark)
        subject['combined_score'] = _subject_combined_score(row)

    for row in attendance_rows:
        if current_semester is not None and row.semester is not None and int(row.semester) > int(current_semester):
            continue
        course_name = _extract_subject_title(row.subject_name)
        subject = ensure_subject(row.semester, row.subject_code or row.subject_name, course_name)
        if not subject:
            continue
        subject['attendance_percentage'] = _valid_metric(row.percentage)
        subject['attended_classes'] = row.attended_classes
        subject['total_classes'] = row.total_classes
        if course_name and (not subject.get('course_name') or subject.get('course_name') == subject.get('subject_code')):
            subject['course_name'] = course_name

    # Cap to current_semester — exclude any future semesters that may exist in attendance/catalog data
    max_allowed_sem = int(current_semester) if current_semester else max(semester_subjects.keys(), default=1)
    ordered_semesters = sorted(s for s in semester_subjects.keys() if s <= max_allowed_sem)
    semester_entries = []

    for semester in ordered_semesters:
        subjects = list(semester_subjects[semester].values())
        subjects.sort(key=lambda item: (str(item.get('subject_code') or ''), str(item.get('course_name') or '')))

        internal1_avg = _average_numeric([item.get('internal1') for item in subjects])
        internal2_avg = _average_numeric([item.get('internal2') for item in subjects])
        internal3_avg = _average_numeric([item.get('internal3') for item in subjects])
        internal_avg = _average_numeric([item.get('internal_avg') for item in subjects])
        attendance_avg = _average_numeric([item.get('attendance_percentage') for item in subjects])
        university_avg = _average_numeric([item.get('university_mark') for item in subjects])
        combined_avg = _average_numeric([item.get('combined_score') for item in subjects])
        progression_score = combined_avg
        if progression_score is None:
            fallback_parts = [internal_avg, university_avg, attendance_avg]
            fallback_parts = [value for value in fallback_parts if value is not None]
            progression_score = round(sum(fallback_parts) / len(fallback_parts), 2) if fallback_parts else None

        semester_entries.append({
            'semester': semester,
            'subject_count': len(subjects),
            'internal1_avg': internal1_avg,
            'internal2_avg': internal2_avg,
            'internal3_avg': internal3_avg,
            'internal_avg': internal_avg,
            'attendance_avg': attendance_avg,
            'university_avg': university_avg,
            'combined_avg': combined_avg,
            'progression_score': progression_score,
            'subjects': subjects,
        })

    previous_progression_score = None
    for entry in semester_entries:
        progression_score = entry.get('progression_score')
        if progression_score is not None and previous_progression_score is not None:
            entry['combined_delta'] = round(progression_score - previous_progression_score, 2)
        else:
            entry['combined_delta'] = None
        if progression_score is not None:
            previous_progression_score = progression_score

    return {
        'current_semester': current_semester,
        'semesters': semester_entries,
    }


@app.route('/api/student/semester-progression/<string:admission_number>', methods=['GET'])
def api_get_student_semester_progression(admission_number):
    try:
        student = Student.query.get(admission_number.upper())
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        data = _build_semester_progression_payload(student)
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/academics/overview/<int:mentor_id>', methods=['GET'])
def api_get_mentor_academics_overview(mentor_id):
    try:
        mentees = Student.query.filter_by(mentor_id=mentor_id).order_by(Student.full_name.asc()).all()
        semester_aggregate = {}
        students_summary = []
        verified_records = 0
        pending_records = 0

        for student in mentees:
            payload = _build_semester_progression_payload(student)
            semester_rows = payload.get('semesters', [])
            current_semester = payload.get('current_semester')
            current_entry = None
            if current_semester is not None:
                current_entry = next((row for row in semester_rows if int(row.get('semester', 0)) == int(current_semester)), None)
            if current_entry is None and semester_rows:
                current_entry = semester_rows[-1]

            marks_rows = StudentMark.query.filter_by(student_id=student.admission_number.upper()).all()
            verified_records += sum(1 for row in marks_rows if getattr(row, 'is_verified', False))
            pending_records += sum(1 for row in marks_rows if not getattr(row, 'is_verified', False))

            for row in semester_rows:
                sem = int(row.get('semester'))
                semester_aggregate.setdefault(sem, {
                    'semester': sem,
                    'progression_scores': [],
                    'attendance_scores': [],
                    'university_scores': [],
                    'combined_scores': [],
                })
                if row.get('progression_score') is not None:
                    semester_aggregate[sem]['progression_scores'].append(float(row['progression_score']))
                if row.get('attendance_avg') is not None:
                    semester_aggregate[sem]['attendance_scores'].append(float(row['attendance_avg']))
                if row.get('university_avg') is not None:
                    semester_aggregate[sem]['university_scores'].append(float(row['university_avg']))
                if row.get('combined_avg') is not None:
                    semester_aggregate[sem]['combined_scores'].append(float(row['combined_avg']))

            students_summary.append({
                'student_id': student.admission_number,
                'student_name': student.full_name,
                'batch': student.batch,
                'current_semester': current_semester,
                'progression_score': None if not current_entry or current_entry.get('progression_score') is None else round(float(current_entry['progression_score']), 2),
                'attendance_avg': None if not current_entry or current_entry.get('attendance_avg') is None else round(float(current_entry['attendance_avg']), 2),
                'university_avg': None if not current_entry or current_entry.get('university_avg') is None else round(float(current_entry['university_avg']), 2),
                'combined_avg': None if not current_entry or current_entry.get('combined_avg') is None else round(float(current_entry['combined_avg']), 2),
                'tracked_semesters': len(semester_rows),
            })

        semester_trends = []
        for sem in sorted(semester_aggregate.keys()):
            entry = semester_aggregate[sem]
            semester_trends.append({
                'semester': sem,
                'progression_score': _average_numeric(entry['progression_scores']),
                'attendance_avg': _average_numeric(entry['attendance_scores']),
                'university_avg': _average_numeric(entry['university_scores']),
                'combined_avg': _average_numeric(entry['combined_scores']),
            })

        return jsonify({
            'success': True,
            'data': {
                'student_count': len(mentees),
                'verified_records': verified_records,
                'pending_records': pending_records,
                'students': students_summary,
                'semester_trends': semester_trends,
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/student/marks/<string:admission_number>', methods=['GET'])
def api_get_student_marks_public(admission_number):
    """Return all StudentMark rows for a student as flat list. Used by React UI."""
    try:
        from models import StudentMark
        student = Student.query.get(admission_number.upper())
        marks = StudentMark.query.filter_by(student_id=admission_number.upper()).order_by(StudentMark.semester.asc(), StudentMark.subject_code.asc()).all()
        data = []
        for m in marks:
            subject_meta = _resolve_subject_catalog_entry(m.subject_code, semester=m.semester, student=student)
            data.append({
                'subject_code':     m.subject_code,
                'course_name':      subject_meta['course_name'],
                'display_name':     subject_meta['display_name'],
                'semester':         m.semester,
                'internal1':        m.internal1,
                'internal2':        m.internal2,
                'internal3':        m.internal3,
                'university_mark':  m.university_mark,
                'university_grade': m.exam_type if m.exam_type and not m.university_mark else None,
                'is_verified':      m.is_verified if hasattr(m, 'is_verified') else False,
            })
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/student/marksheet/upload', methods=['POST'])
def api_marksheet_upload():
    """Upload a KTU marksheet PDF and extract marks via AI."""
    try:
        student_id = request.form.get('student_id', '').upper()
        semester   = request.form.get('semester', '1')
        force      = request.form.get('force_replace', 'false').lower() == 'true'
        student    = Student.query.get(student_id)

        if not student_id:
            return jsonify({'success': False, 'message': 'student_id required'}), 400

        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400

        file = request.files['file']
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'message': 'Only PDF files allowed'}), 400

        from models import StudentMark
        import tempfile, os

        # Check existing marks for that semester
        existing = StudentMark.query.filter_by(student_id=student_id, semester=int(semester)).all()
        if existing and not force:
            return jsonify({'success': False, 'needs_confirmation': True,
                            'message': f'Marks for Semester {semester} already exist. Confirm replace.'}), 200

        # Save PDF to temp file and extract
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        extracted = []
        try:
            import pdfplumber, re as _re
            with pdfplumber.open(tmp_path) as pdf:
                raw_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

            # Simple regex for KTU-style marksheets: subject code + marks
            patterns = [
                r'([A-Z]{2,4}\d{3,4}[A-Z]?)\s+[^\d]*?(\d{1,3}(?:\.\d+)?)\s*(?:/\s*\d+)?',
                r'([A-Z]{2,4}\d{3,6})\s*\|\s*[^\|]+\|\s*(\d{1,3})',
            ]
            seen = set()
            for pat in patterns:
                for m in _re.finditer(pat, raw_text):
                    code = m.group(1).strip()
                    mark_val = float(m.group(2))
                    if code not in seen and 0 <= mark_val <= 100:
                        seen.add(code)
                        extracted.append({'subj': code, 'mark': mark_val})

            if not extracted:
                # Fall back: grab all float-looking numbers near word boundaries
                lines = raw_text.split('\n')
                for line in lines:
                    parts = line.split()
                    if len(parts) >= 2:
                        code_candidate = parts[0]
                        if _re.match(r'^[A-Z]{2,5}\d{3,6}', code_candidate) and code_candidate not in seen:
                            for part in parts[1:]:
                                try:
                                    v = float(part.replace(',', '.'))
                                    if 0 <= v <= 100:
                                        seen.add(code_candidate)
                                        extracted.append({'subj': code_candidate, 'mark': v})
                                        break
                                except ValueError:
                                    continue

        except ImportError:
            os.unlink(tmp_path)
            return jsonify({'success': False, 'message': 'pdfplumber not installed. Run: pip install pdfplumber'}), 500
        finally:
            try: os.unlink(tmp_path)
            except: pass

        if not extracted:
            return jsonify({'success': False, 'message': 'Could not extract any subject marks from PDF. Please check the format.'}), 200

        # Delete old marks for this semester if force
        if force and existing:
            for old in existing:
                db.session.delete(old)

        # Save extracted marks
        saved = []
        for item in extracted:
            sm = StudentMark(
                student_id=student_id,
                subject_code=item['subj'],
                semester=int(semester),
                university_mark=item['mark'],
            )
            db.session.add(sm)
            saved.append({
                **item,
                **_resolve_subject_catalog_entry(item['subj'], semester=semester, student=student),
            })

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Extracted and saved {len(saved)} subjects for Semester {semester}.',
            'extracted': saved,
            'raw_text_preview': raw_text[:800] if 'raw_text' in locals() else ''
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/student/marksheet/download/<string:admission_number>/<int:semester>', methods=['GET'])
def api_marksheet_download(admission_number, semester):
    """Generate a university-style PDF report card for a verified semester result."""
    try:
        from io import BytesIO
        from flask import send_file
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from models import StudentMark

        student = Student.query.get(admission_number.upper())
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        marks = StudentMark.query.filter_by(
            student_id=admission_number.upper(),
            semester=semester
        ).order_by(StudentMark.subject_code.asc()).all()
        if not marks:
            return jsonify({'success': False, 'message': 'No marks found for this semester'}), 404

        if any(not getattr(m, 'is_verified', False) for m in marks):
            return jsonify({'success': False, 'message': 'Marks must be verified by a mentor before download.'}), 403

        def fmt_num(value):
            if value is None:
                return '-'
            numeric = float(value)
            return f"{int(numeric)}" if numeric.is_integer() else f"{numeric:.2f}"

        def grade_point(mark):
            if mark is None:
                return None
            mark = float(mark)
            if mark >= 90:
                return 10
            if mark >= 80:
                return 9
            if mark >= 70:
                return 8
            if mark >= 60:
                return 7
            if mark >= 50:
                return 6
            if mark >= 45:
                return 5
            if mark >= 40:
                return 4
            return 0

        def grade_letter(mark):
            gp = grade_point(mark)
            mapping = {10: 'S', 9: 'A+', 8: 'A', 7: 'B+', 6: 'B', 5: 'C', 4: 'P', 0: 'F'}
            return '-' if gp is None else mapping.get(gp, '-')

        def calc_sgpa(rows):
            gps = [grade_point(row.university_mark) for row in rows if row.university_mark is not None]
            gps = [gp for gp in gps if gp is not None]
            return round(sum(gps) / len(gps), 2) if gps else None

        semester_sgpa = calc_sgpa(marks)
        cgpa_value = None
        if student.academics and student.academics.cgpa is not None:
            cgpa_value = round(float(student.academics.cgpa), 2)
        else:
            all_marks = StudentMark.query.filter_by(student_id=admission_number.upper()).all()
            by_semester = {}
            for row in all_marks:
                by_semester.setdefault(row.semester, []).append(row)
            sgpa_values = [calc_sgpa(rows) for rows in by_semester.values()]
            sgpa_values = [value for value in sgpa_values if value is not None]
            if sgpa_values:
                cgpa_value = round(sum(sgpa_values) / len(sgpa_values), 2)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title_style', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=15, leading=18, textColor=colors.HexColor('#8a1218'), alignment=TA_CENTER)
        sub_style = ParagraphStyle('sub_style', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=12, textColor=colors.HexColor('#0f172a'), alignment=TA_CENTER)
        meta_style = ParagraphStyle('meta_style', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#475569'), alignment=TA_CENTER)
        label_style = ParagraphStyle('label_style', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor('#8a1218'), alignment=TA_LEFT)
        value_style = ParagraphStyle('value_style', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#111827'), alignment=TA_LEFT)
        value_right = ParagraphStyle('value_right', parent=value_style, alignment=TA_RIGHT)

        story = []
        logo_path = os.path.join(app.root_path, 'static', 'branding', 'sjcet-logo.png')
        header = Table([[
            Image(logo_path, width=27 * mm, height=27 * mm) if os.path.exists(logo_path) else Paragraph('SJCET', title_style),
            [
                Paragraph("St. Joseph's College of Engineering and Technology, Palai", title_style),
                Paragraph("(Autonomous)", sub_style),
                Paragraph("Detailed Report Card", sub_style),
                Paragraph("University Mark Statement", meta_style),
            ],
            Paragraph(
                f"Semester {semester}<br/>Verified Result",
                ParagraphStyle('side_heading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=14, alignment=TA_RIGHT, textColor=colors.HexColor('#0f172a'))
            )
        ]], colWidths=[32 * mm, 110 * mm, 32 * mm])
        header.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header)
        story.append(Spacer(1, 5 * mm))

        info_table = Table([
            [Paragraph("Student Name", label_style), Paragraph(student.full_name, value_style), Paragraph("Admission No.", label_style), Paragraph(student.admission_number, value_style)],
            [Paragraph("Programme / Branch", label_style), Paragraph(student.branch or '-', value_style), Paragraph("Batch", label_style), Paragraph(student.batch or '-', value_style)],
        ], colWidths=[30 * mm, 65 * mm, 25 * mm, 52 * mm])
        info_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.7, colors.HexColor('#cbd5e1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 5 * mm))

        rows = [[
            'Sl. No.', 'Course', 'Code', 'Internal 1', 'Internal 2', 'Internal 3',
            'University', 'Total', 'Grade', 'Grade Point', 'Result'
        ]]
        for index, mark in enumerate(marks, start=1):
            subject_meta = _resolve_subject_catalog_entry(mark.subject_code, semester=semester, student=student)
            internal_total = sum(float(v) for v in [mark.internal1, mark.internal2, mark.internal3] if v is not None)
            university = float(mark.university_mark) if mark.university_mark is not None else None
            total = internal_total + university if university is not None else None
            gp = grade_point(university)
            rows.append([
                str(index),
                subject_meta['course_name'] or '-',
                mark.subject_code or '-',
                fmt_num(mark.internal1),
                fmt_num(mark.internal2),
                fmt_num(mark.internal3),
                fmt_num(university),
                fmt_num(total),
                grade_letter(university),
                '-' if gp is None else str(gp),
                'PASS' if gp is not None and gp > 0 else 'FAIL' if gp == 0 else '-',
            ])

        marks_table = Table(rows, colWidths=[12 * mm, 36 * mm, 20 * mm, 15 * mm, 15 * mm, 15 * mm, 18 * mm, 15 * mm, 11 * mm, 15 * mm, 16 * mm], repeatRows=1)
        marks_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8a1218')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.45, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(marks_table)
        story.append(Spacer(1, 5 * mm))

        summary_table = Table([[
            Paragraph("Semester SGPA", label_style),
            Paragraph('-' if semester_sgpa is None else f"{semester_sgpa:.2f}", value_right),
            Paragraph("Cumulative CGPA", label_style),
            Paragraph('-' if cgpa_value is None else f"{cgpa_value:.2f}", value_right),
        ]], colWidths=[38 * mm, 32 * mm, 38 * mm, 32 * mm])
        summary_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#8a1218')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fff7ed')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 6 * mm))

        footer = Table([[
            Paragraph("Controller of Examinations<br/><font size='8'>St. Joseph's College of Engineering and Technology, Palai</font>", meta_style),
            Paragraph(f"Generated on: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}<br/>System generated report card for student use.", meta_style),
        ]], colWidths=[85 * mm, 85 * mm])
        footer.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, -1), 0.6, colors.HexColor('#94a3b8')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(footer)

        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f"{student.admission_number}_semester_{semester}_report_card.pdf")
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/mentor/marks/verify/<string:student_id>/<int:semester>', methods=['POST'])
def api_mentor_verify_marks(student_id, semester):
    """Verify or Unlock all marks for a given student's semester."""
    try:
        data = request.get_json(force=True) or {}
        action = data.get('action', 'verify') # 'verify' or 'unlock'
        
        from models import StudentMark
        marks = StudentMark.query.filter_by(student_id=student_id.upper(), semester=semester).all()
        
        if not marks:
            return jsonify({'success': False, 'message': 'No marks found to verify'}), 404
            
        for m in marks:
            m.is_verified = (action == 'verify')
            
        db.session.commit()
        return jsonify({'success': True, 'message': f'Marks {"verified" if action == "verify" else "unlocked"} successfully.'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/student/routine-preferences/<string:admission_number>', methods=['GET'])
def api_get_routine_preferences(admission_number):
    try:
        from models import RoutinePrefs
        prefs = RoutinePrefs.query.filter_by(student_id=admission_number.upper()).first()
        if prefs:
            return jsonify({'success': True, 'data': {
                'wakeup_time': prefs.wakeup_time,
                'prayer_time': prefs.prayer_time,
                'breakfast_time': prefs.breakfast_time,
                'college_start': prefs.college_start,
                'college_end': prefs.college_end,
                'refresh_time': prefs.refresh_time,
                'play_time': prefs.play_time,
                'food_time': prefs.food_time,
                'bed_time': prefs.bed_time,
            }}), 200
        else:
            return jsonify({'success': True, 'data': {}}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/student/routine-preferences/<string:admission_number>', methods=['POST'])
def api_set_routine_preferences(admission_number):
    try:
        from models import RoutinePrefs
        data = request.get_json()
        prefs = RoutinePrefs.query.filter_by(student_id=admission_number.upper()).first()
        if not prefs:
            prefs = RoutinePrefs(student_id=admission_number.upper())
            db.session.add(prefs)
        
        prefs.wakeup_time = data.get('wakeup_time', prefs.wakeup_time)
        prefs.prayer_time = data.get('prayer_time', prefs.prayer_time)
        prefs.breakfast_time = data.get('breakfast_time', prefs.breakfast_time)
        prefs.college_start = data.get('college_start', prefs.college_start)
        prefs.college_end = data.get('college_end', prefs.college_end)
        prefs.refresh_time = data.get('refresh_time', prefs.refresh_time)
        prefs.play_time = data.get('play_time', prefs.play_time)
        prefs.food_time = data.get('food_time', prefs.food_time)
        prefs.bed_time = data.get('bed_time', prefs.bed_time)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Routine preferences saved successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/student/adaptive-timetable/<string:admission_number>', methods=['GET'])
def api_get_adaptive_timetable(admission_number):
    try:
        from models import WeeklyStudyPlan
        
        plan = WeeklyStudyPlan.query.filter_by(student_id=admission_number.upper()).order_by(WeeklyStudyPlan.id.desc()).first()
        
        subjects = []
        if plan and plan.subjects:
            for s in plan.subjects:
                if s.subject:
                    subjects.append(s.subject.name)
                elif hasattr(s, 'subject_name'):
                    subjects.append(getattr(s, 'subject_name'))
                    
        if not subjects:
            subjects = ["Core Subjects", "Assignments", "Self Study", "Revision"]
            
        while len(subjects) < 4:
            subjects.append(subjects[-1] if subjects else "Self Study")

        weekday_slots = [
            {"time": "18:00 - 19:00", "subject": subjects[0], "activity": "Deep learning block"},
            {"time": "19:00 - 20:00", "subject": subjects[1] if len(subjects) > 1 else subjects[0], "activity": "Practice & Revision"}
        ]
        
        weekend_slots = [
            {"time": "10:00 - 12:00", "subject": subjects[0], "activity": "Intensive study & Notes"},
            {"time": "14:00 - 16:00", "subject": subjects[1] if len(subjects) > 1 else subjects[0], "activity": "Assignments & Problem solving"},
            {"time": "18:00 - 20:00", "subject": subjects[2] if len(subjects) > 2 else subjects[0], "activity": "Weekly review & Mock tests"}
        ]
        
        return jsonify({'success': True, 'data': {'weekday': weekday_slots, 'weekend': weekend_slots}}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ─── Mentor Chat ──────────────────────────────────────────────────
@app.route('/api/chat/mentor/<string:admission_number>', methods=['GET'])
def api_get_mentor_chat(admission_number):
    try:
        from models import Student, Faculty, MentorMessage
        student = Student.query.get(admission_number.upper())
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        if not student.mentor_id:
            return jsonify({'success': True, 'no_mentor': True, 'message': 'Mentor not assigned yet'}), 200
            
        mentor = Faculty.query.get(student.mentor_id)
        msgs = MentorMessage.query.filter_by(student_id=admission_number.upper(), mentor_id=student.mentor_id).order_by(MentorMessage.sent_at.asc()).all()
        
        return jsonify({
            'success': True,
            'mentor': {'id': mentor.id, 'name': mentor.name, 'designation': mentor.designation},
            'data': [{
                'id': m.id,
                'message': m.message,
                'sender_role': m.sender_role,
                'sent_at': m.sent_at.isoformat()
            } for m in msgs]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/mentor/<string:admission_number>', methods=['POST'])
def api_send_mentor_message(admission_number):
    try:
        from models import Student, MentorMessage
        data = request.get_json()
        student = Student.query.get(admission_number.upper())
        if not student or not student.mentor_id:
            return jsonify({'success': False, 'message': 'No mentor assigned'}), 400
            
        msg = MentorMessage(
            student_id=admission_number.upper(),
            mentor_id=student.mentor_id,
            message=data.get('message'),
            sender_role='student'
        )
        db.session.add(msg)
        db.session.commit()
        return jsonify({'success': True}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ─── Subject Handler Messages ─────────────────────────────────────
def _departments_match_simple(left, right):
    left_norm = normalize_dept_name(str(left or '').strip())
    right_norm = normalize_dept_name(str(right or '').strip())
    if not left_norm or not right_norm:
        return False
    if left_norm == right_norm:
        return True
    compact_left = re.sub(r'[^a-z0-9]+', '', left_norm.lower())
    compact_right = re.sub(r'[^a-z0-9]+', '', right_norm.lower())
    if compact_left == compact_right:
        return True
    if compact_left in ('mca', 'imca') and 'computerapplication' in compact_right:
        return True
    if compact_right in ('mca', 'imca') and 'computerapplication' in compact_left:
        return True
    return False


def _playground_visible_for_student(note, student):
    if note.scope == 'student':
        return note.target_student_id == student.admission_number
    if note.batch and str(note.batch).strip() and str(student.batch or '').strip() and str(note.batch).strip().lower() != str(student.batch).strip().lower():
        return False
    if note.department and str(note.department).strip():
        return _departments_match_simple(note.department, student.branch)
    return True


def _extract_multiple_student_targets(raw_value):
    raw = str(raw_value or '').strip().upper()
    if not raw:
        return []
    parts = [part.strip() for part in re.split(r'[,\n;]+', raw) if part.strip()]
    return list(dict.fromkeys(parts))


def _is_student_academically_weak(student_id, subject_code=None):
    try:
        from models import StudentAnalytics
        analytics = StudentAnalytics.query.filter_by(student_id=student_id).first()
        if analytics and float(getattr(analytics, 'adjusted_risk', 0) or getattr(analytics, 'risk_score', 0) or 0) >= 60:
            return True
    except Exception:
        analytics = None

    normalized_subject = str(subject_code or '').strip().upper()
    if normalized_subject:
        handler_marks = SubjectHandlerMark.query.filter_by(student_id=student_id, subject_code=normalized_subject).all()
        if handler_marks:
            scores = []
            for mark in handler_marks:
                try:
                    maximum = float(mark.max_marks or 0)
                    obtained = float(mark.marks_obtained or 0)
                    if maximum > 0:
                        scores.append((obtained / maximum) * 100)
                except Exception:
                    continue
            if scores and (sum(scores) / len(scores)) < 50:
                return True

        attendance_rows = SubjectHandlerAttendance.query.filter_by(student_id=student_id, subject_code=normalized_subject).all()
        if attendance_rows:
            present = sum(1 for row in attendance_rows if str(row.status or '').strip().lower() == 'present')
            if len(attendance_rows) > 0 and ((present / len(attendance_rows)) * 100) < 75:
                return True

    mark_rows = StudentMark.query.filter_by(student_id=student_id).all()
    low_mark_hits = 0
    for mark in mark_rows:
        for field_name in ('internal1', 'internal2', 'internal3', 'university_mark'):
            value = getattr(mark, field_name, None)
            if value is not None and float(value) < 35:
                low_mark_hits += 1
                break
    return low_mark_hits >= 2


def _queue_personalized_note_alert(student, subject_code, note_title, sender_name):
    if not student or not getattr(student, 'mentor_id', None):
        return False
    if not _is_student_academically_weak(student.admission_number, subject_code=subject_code):
        return False

    db.session.add(Alert(
        student_admission_number=student.admission_number,
        mentor_id=student.mentor_id,
        type='PERSONALIZED_NOTE_REQUEST',
        message=(
            f"{sender_name} shared dedicated support material '{note_title}' for {student.full_name} "
            f"in {subject_code}. Please add personalized mentor notes and monitor progress closely."
        ),
        is_read=False,
    ))
    return True


@app.route('/api/messages/handlers/<string:department>', methods=['GET'])
def api_get_subject_handlers(department):
    try:
        dept = str(department or '').strip()
        handlers = Faculty.query.filter(
            Faculty.is_subject_handler == True,
            Faculty.status == 'Live'
        ).all()
        filtered = [h for h in handlers if _departments_match_simple(h.department, dept)]
        return jsonify({
            'success': True,
            'data': [{'id': h.id, 'name': h.name, 'designation': h.designation} for h in filtered]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/messages/handler/<string:admission_number>', methods=['GET'])
def api_get_handler_messages(admission_number):
    try:
        sid = str(admission_number or '').strip().upper()
        handler_id = request.args.get('handler_id', type=int)
        q = SubjectHandlerMessage.query.filter_by(student_id=sid)
        if handler_id:
            q = q.filter_by(handler_id=handler_id)
        rows = q.order_by(SubjectHandlerMessage.created_at.asc()).all()
        return jsonify({
            'success': True,
            'data': [{
                'id': row.id,
                'subject': row.subject,
                'message': row.message,
                'description': row.message,
                'status': row.status,
                'raised_at': row.created_at.isoformat() if row.created_at else None,
                'created_at': row.created_at.isoformat() if row.created_at else None,
                'resolution_notes': row.message if row.sender_role == 'handler' else None,
                'category': row.category,
                'sender_role': row.sender_role,
                'handler_id': row.handler_id,
                'handler_name': row.handler.name if row.handler else 'Subject Handler',
                'attachment_path': row.attachment_path,
                'is_read': bool(row.is_read),
            } for row in rows]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/messages/handler/<string:admission_number>', methods=['POST'])
def api_send_handler_message(admission_number):
    try:
        sid = str(admission_number or '').strip().upper()
        student = Student.query.get(sid)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        handler_id = request.form.get('handler_id', type=int)
        if not handler_id:
            return jsonify({'success': False, 'message': 'Please select a subject handler'}), 400
        handler = Faculty.query.get(handler_id)
        if not handler or not handler.is_subject_handler:
            return jsonify({'success': False, 'message': 'Subject handler not found'}), 404

        message = (request.form.get('message') or '').strip()
        attachment_path = None
        file = request.files.get('file')
        if file and file.filename:
            save_dir = os.path.join(app.root_path, 'static', 'uploads', 'handler_messages', sid)
            os.makedirs(save_dir, exist_ok=True)
            filename = secure_filename(f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}")
            full_path = os.path.join(save_dir, filename)
            file.save(full_path)
            attachment_path = os.path.relpath(full_path, os.path.join(app.root_path, 'static')).replace('\\', '/')

        if not message and not attachment_path:
            return jsonify({'success': False, 'message': 'Message or attachment is required'}), 400

        row = SubjectHandlerMessage(
            student_id=sid,
            handler_id=handler_id,
            subject=(request.form.get('subject') or '').strip() or 'General',
            category=(request.form.get('category') or 'Academic').strip() or 'Academic',
            message=message or 'Attachment shared',
            attachment_path=attachment_path,
            sender_role='student',
            status='open',
            is_read=False,
        )
        db.session.add(row)
        db.session.commit()
        return jsonify({'success': True, 'data': {'id': row.id}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/messages/<int:handler_id>', methods=['GET'])
def api_handler_messages(handler_id):
    try:
        student_id = str(request.args.get('student_id') or '').strip().upper()
        q = SubjectHandlerMessage.query.filter_by(handler_id=handler_id)
        if student_id:
            q = q.filter_by(student_id=student_id)
        rows = q.order_by(SubjectHandlerMessage.created_at.asc()).all()
        return jsonify({
            'success': True,
            'data': [{
                'id': row.id,
                'student_id': row.student_id,
                'student_name': row.student.full_name if row.student else row.student_id,
                'subject': row.subject,
                'category': row.category,
                'message': row.message,
                'attachment_path': row.attachment_path,
                'sender_role': row.sender_role,
                'status': row.status,
                'created_at': row.created_at.isoformat() if row.created_at else None,
                'is_read': bool(row.is_read),
            } for row in rows]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/messages/send', methods=['POST'])
def api_handler_messages_send():
    try:
        payload = request.get_json(force=True) or {}
        handler_id = int(payload.get('handler_id') or 0)
        student_id = str(payload.get('student_id') or '').strip().upper()
        message = str(payload.get('message') or '').strip()
        subject = str(payload.get('subject') or 'General').strip() or 'General'
        category = str(payload.get('category') or 'Academic').strip() or 'Academic'
        if not handler_id or not student_id or not message:
            return jsonify({'success': False, 'message': 'handler_id, student_id and message are required'}), 400

        handler = Faculty.query.get(handler_id)
        if not handler or not handler.is_subject_handler:
            return jsonify({'success': False, 'message': 'Subject handler not found'}), 404

        row = SubjectHandlerMessage(
            student_id=student_id,
            handler_id=handler_id,
            subject=subject,
            category=category,
            message=message,
            sender_role='handler',
            status='replied',
            is_read=False,
        )
        db.session.add(row)
        db.session.add(Notification(
            student_id=student_id,
            title='Reply from Subject Handler',
            message=f'{handler.name} replied regarding {subject}.',
            type='subject_handler_message',
            is_read=False,
        ))
        db.session.commit()
        return jsonify({'success': True, 'data': {'id': row.id}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/playground/<string:admission_number>', methods=['GET'])
def api_student_playground(admission_number):
    try:
        sid = str(admission_number or '').strip().upper()
        student = Student.query.get(sid)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        notes = PlaygroundNote.query.order_by(PlaygroundNote.created_at.desc()).all()
        visible_notes = [note for note in notes if _playground_visible_for_student(note, student)]
        return jsonify({
            'success': True,
            'data': [{
                'id': note.id,
                'title': note.title,
                'description': note.description,
                'subject_code': note.subject_code,
                'scope': note.scope,
                'target_student_id': note.target_student_id,
                'department': note.department,
                'batch': note.batch,
                'file_path': note.file_path,
                'download_url': f"/static/{note.file_path}" if note.file_path else None,
                'uploaded_by_name': note.uploader.name if note.uploader else 'Subject Handler',
                'created_at': note.created_at.isoformat() if note.created_at else None,
            } for note in visible_notes]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/playground/upload', methods=['POST'])
def api_handler_playground_upload():
    try:
        handler_id = request.form.get('handler_id', type=int)
        subject_code = (request.form.get('subject_code') or 'GENERAL').strip().upper()
        title = (request.form.get('title') or '').strip()
        scope = (request.form.get('scope') or 'class').strip().lower()
        target_student_id = (request.form.get('target_student_id') or '').strip().upper() or None
        department = (request.form.get('department') or '').strip() or None
        batch = (request.form.get('batch') or '').strip() or None
        description = (request.form.get('description') or '').strip()
        file = request.files.get('file')

        if not handler_id or not title or not file or not file.filename:
            return jsonify({'success': False, 'message': 'handler_id, title and file are required'}), 400
        invalid_targets = _extract_multiple_student_targets(target_student_id)
        if len(invalid_targets) > 1:
            return jsonify({'success': False, 'message': 'Dedicated notes can be shared to only one student at a time'}), 400
        if invalid_targets:
            target_student_id = invalid_targets[0]
        if scope == 'student' and not target_student_id:
            return jsonify({'success': False, 'message': 'target_student_id is required for dedicated student notes'}), 400
        if scope != 'student':
            target_student_id = None

        save_dir = os.path.join(app.root_path, 'static', 'uploads', 'playground', subject_code)
        os.makedirs(save_dir, exist_ok=True)
        filename = secure_filename(f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}")
        full_path = os.path.join(save_dir, filename)
        file.save(full_path)
        rel_path = os.path.relpath(full_path, os.path.join(app.root_path, 'static')).replace('\\', '/')

        note = PlaygroundNote(
            subject_code=subject_code,
            title=title,
            description=description,
            file_path=rel_path,
            scope='student' if scope == 'student' else 'class',
            target_student_id=target_student_id,
            department=department,
            batch=batch,
            uploaded_by_id=handler_id,
        )
        db.session.add(note)

        targets = []
        mentor_alerted = False
        if note.scope == 'student' and target_student_id:
            target_student = Student.query.get(target_student_id)
            if not target_student:
                return jsonify({'success': False, 'message': 'Target student not found'}), 404
            if batch and str(target_student.batch or '').strip().lower() != str(batch).strip().lower():
                return jsonify({'success': False, 'message': 'Selected student does not belong to this batch'}), 400
            if department and not _departments_match_simple(department, target_student.branch):
                return jsonify({'success': False, 'message': 'Selected student does not belong to this department'}), 400
            targets = [target_student_id]
            mentor_alerted = _queue_personalized_note_alert(
                target_student,
                subject_code=subject_code,
                note_title=title,
                sender_name=(Faculty.query.get(handler_id).name if Faculty.query.get(handler_id) else "Subject Handler"),
            )
        else:
            for student in Student.query.filter_by(status='Live').all():
                if _playground_visible_for_student(note, student):
                    targets.append(student.admission_number)

        for target in targets:
            db.session.add(Notification(
                student_id=target,
                title=f'New Playground Note: {title}',
                message=f'New {subject_code} study material has been shared in Playground.',
                type='playground_note',
                is_read=False,
            ))

        db.session.commit()
        return jsonify({
            'success': True,
            'data': {
                'id': note.id,
                'title': note.title,
                'scope': note.scope,
                'download_url': f"/static/{note.file_path}",
                'notified_count': len(targets),
                'mentor_alerted': mentor_alerted,
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/faculty/classes/<int:faculty_id>', methods=['GET'])
def api_faculty_classes(faculty_id):
    try:
        faculty = Faculty.query.get(faculty_id)
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404

        entries = Timetable.query.filter_by(handler_id=faculty_id).all()
        seen = set()
        rows = []
        for t in entries:
            dept = str(t.department or '').strip()
            batch = str(t.batch or '').strip()
            subject = str(t.subject or '').strip()
            if not dept or not batch:
                continue
            key = (dept.lower(), batch.lower(), subject.lower())
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                'department': dept,
                'batch': batch,
                'subject': subject,
            })
        rows.sort(key=lambda r: (r['department'].lower(), r['batch'].lower(), r['subject'].lower()))
        return jsonify({'success': True, 'data': rows}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/faculty/class-students', methods=['GET'])
def api_faculty_class_students():
    try:
        department = str(request.args.get('department') or '').strip()
        batch = str(request.args.get('batch') or '').strip()
        if not department or not batch:
            return jsonify({'success': False, 'message': 'department and batch are required'}), 400

        students = Student.query.filter_by(status='Live').all()
        result = []
        for s in students:
            if str(s.batch or '').strip().lower() != batch.lower():
                continue
            if not _departments_match_simple(department, s.branch):
                continue
            result.append({
                'admission_number': s.admission_number,
                'full_name': s.full_name,
                'email': s.email,
                'batch': s.batch,
                'branch': s.branch,
            })
        result.sort(key=lambda r: r['admission_number'])
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/faculty/notes/upload', methods=['POST'])
def api_faculty_notes_upload():
    try:
        faculty_id = request.form.get('faculty_id', type=int)
        subject_code = (request.form.get('subject_code') or 'GENERAL').strip().upper()
        title = (request.form.get('title') or '').strip()
        scope = (request.form.get('scope') or 'class').strip().lower()
        target_student_id = (request.form.get('target_student_id') or '').strip().upper() or None
        department = (request.form.get('department') or '').strip() or None
        batch = (request.form.get('batch') or '').strip() or None
        description = (request.form.get('description') or '').strip()
        file = request.files.get('file')

        if not faculty_id or not title or not file or not file.filename:
            return jsonify({'success': False, 'message': 'faculty_id, title and file are required'}), 400
        if not department or not batch:
            return jsonify({'success': False, 'message': 'department and batch are required'}), 400
        invalid_targets = _extract_multiple_student_targets(target_student_id)
        if len(invalid_targets) > 1:
            return jsonify({'success': False, 'message': 'Dedicated notes can be shared to only one student at a time'}), 400
        if invalid_targets:
            target_student_id = invalid_targets[0]
        if scope == 'student' and not target_student_id:
            return jsonify({'success': False, 'message': 'target_student_id is required for dedicated student notes'}), 400
        if scope != 'student':
            target_student_id = None

        faculty = Faculty.query.get(faculty_id)
        if not faculty or faculty.status != 'Live':
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404

        tt = Timetable.query.filter_by(handler_id=faculty_id).all()
        allowed = any(
            _departments_match_simple(getattr(t, 'department', ''), department) and str(getattr(t, 'batch', '') or '').strip().lower() == batch.lower()
            for t in tt
        )
        if not allowed:
            return jsonify({'success': False, 'message': 'Access denied for this class'}), 403

        save_dir = os.path.join(app.root_path, 'static', 'uploads', 'faculty_notes', subject_code)
        os.makedirs(save_dir, exist_ok=True)
        filename = secure_filename(f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}")
        full_path = os.path.join(save_dir, filename)
        file.save(full_path)
        rel_path = os.path.relpath(full_path, os.path.join(app.root_path, 'static')).replace('\\', '/')

        note = PlaygroundNote(
            subject_code=subject_code,
            title=title,
            description=description,
            file_path=rel_path,
            scope='student' if scope == 'student' else 'class',
            target_student_id=target_student_id,
            department=department,
            batch=batch,
            uploaded_by_id=faculty_id,
        )
        db.session.add(note)

        targets = []
        mentor_alerted = False
        if note.scope == 'student' and target_student_id:
            target_student = Student.query.get(target_student_id)
            if not target_student:
                return jsonify({'success': False, 'message': 'Target student not found'}), 404
            if str(target_student.batch or '').strip().lower() != str(batch or '').strip().lower() or not _departments_match_simple(department, target_student.branch):
                return jsonify({'success': False, 'message': 'Selected student does not belong to this class'}), 400
            targets = [target_student_id]
            mentor_alerted = _queue_personalized_note_alert(
                target_student,
                subject_code=subject_code,
                note_title=title,
                sender_name=faculty.name,
            )
        else:
            for student in Student.query.filter_by(status='Live').all():
                if _playground_visible_for_student(note, student):
                    targets.append(student.admission_number)

        for target in targets:
            db.session.add(Notification(
                student_id=target,
                title=f'New Note: {title}',
                message=f'{faculty.name} shared a PDF note for {subject_code}.',
                type='faculty_note',
                is_read=False,
            ))

        db.session.commit()
        return jsonify({
            'success': True,
            'data': {
                'id': note.id,
                'title': note.title,
                'scope': note.scope,
                'download_url': f"/static/{note.file_path}",
                'notified_count': len(targets),
                'mentor_alerted': mentor_alerted,
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/playground/<int:handler_id>', methods=['GET'])
def api_handler_playground_list(handler_id):
    try:
        notes = PlaygroundNote.query.filter_by(uploaded_by_id=handler_id).order_by(PlaygroundNote.created_at.desc()).all()
        return jsonify({
            'success': True,
            'data': [{
                'id': note.id,
                'title': note.title,
                'description': note.description,
                'subject_code': note.subject_code,
                'scope': note.scope,
                'target_student_id': note.target_student_id,
                'department': note.department,
                'batch': note.batch,
                'download_url': f"/static/{note.file_path}" if note.file_path else None,
                'created_at': note.created_at.isoformat() if note.created_at else None,
            } for note in notes]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ─── Notifications ────────────────────────────────────────────────
@app.route('/api/student/notifications/<string:admission_number>', methods=['GET'])
def api_get_student_notifications(admission_number):
    try:
        from models import Notification
        notifs = Notification.query.filter_by(student_id=admission_number.upper()).order_by(Notification.created_at.desc()).all()
        return jsonify({
            'success': True,
            'data': [{
                'id': n.id,
                'title': n.title,
                'message': n.message,
                'type': n.type,
                'is_read': n.is_read,
                'created_at': n.created_at.isoformat()
            } for n in notifs]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/student/notifications/<string:admission_number>/unread-count', methods=['GET'])
def api_get_unread_count(admission_number):
    try:
        from models import Notification
        count = Notification.query.filter_by(student_id=admission_number.upper(), is_read=False).count()
        return jsonify({'success': True, 'count': count}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============= SUBJECT HANDLER DATA OPERATIONS =============

def _norm_subject(subject_code):
    return str(subject_code or '').strip().upper()

def _norm_student(student_id):
    return str(student_id or '').strip().upper()

def _norm_scope_value(value):
    return str(value or '').strip()

def _scope_text_matches(left, right):
    return _norm_scope_value(left).lower() == _norm_scope_value(right).lower()

def _scope_department_matches(student_branch, scope_department):
    """
    Timetable departments are sometimes stored as long names (e.g. "Computer Applications")
    while student.branch may store abbreviations (e.g. "MCA"). Keep matching tolerant so
    subject handlers can still see their allocated students.
    """
    left = _norm_scope_value(student_branch).lower()
    right = _norm_scope_value(scope_department).lower()
    if not left or not right:
        return False
    if left == right:
        return True
    if left in right or right in left:
        return True

    def _compact(v):
        return re.sub(r'[^a-z0-9]+', '', str(v or '').lower())

    def _initialism(v):
        parts = re.split(r'\s+', re.sub(r'[^a-z0-9\s]+', ' ', str(v or '').lower()).strip())
        return ''.join(p[0] for p in parts if p)

    lc = _compact(left)
    rc = _compact(right)
    if lc and lc == rc:
        return True
    li = _initialism(left)
    ri = _initialism(right)
    if li and li == ri:
        return True

    # Common alias: MCA / IMCA are part of Computer Applications departments in many institutes.
    if lc in ('mca', 'imca') and ('computer' in right and 'application' in right):
        return True
    if rc in ('mca', 'imca') and ('computer' in left and 'application' in left):
        return True
    return False

def _scope_batch_matches(left, right):
    left_norm = _normalize_batch_label(left)
    right_norm = _normalize_batch_label(right)
    if not left_norm or not right_norm:
        return left_norm == right_norm
    if left_norm == right_norm:
        return True
    if left_norm.endswith(f" {right_norm}") or right_norm.endswith(f" {left_norm}"):
        return True
    left_years = _extract_batch_years(left)
    right_years = _extract_batch_years(right)
    return bool(left_years and right_years and left_years == right_years)

ACADEMIC_ENTRY_LIMITS = {
    'internal_assessment_score': float(os.environ.get('ACADEMIC_INTERNAL_MAX_MARKS', '100')),
    'practical_lab_score': float(os.environ.get('ACADEMIC_PRACTICAL_MAX_MARKS', '100')),
}

def _handler_subject_rows(handler_id):
    rows = Timetable.query.filter_by(handler_id=handler_id).all()
    faculty = Faculty.query.get(handler_id)
    mapped = []
    seen = set()
    for r in rows:
        subject_code = _norm_subject(r.subject)
        if not subject_code:
            continue
        department = _norm_scope_value(r.department)
        batch = _norm_scope_value(r.batch)
        key = (subject_code, department, batch)
        if key in seen:
            continue
        seen.add(key)
        mapped.append({
            'subject_code': subject_code,
            'department': department,
            'batch': batch,
            'faculty_id': handler_id,
            'faculty_name': faculty.name if faculty else '',
            'academic_limits': ACADEMIC_ENTRY_LIMITS,
        })
    return mapped

def _find_handler_subject_scope(handler_id, subject_code, department=None, batch=None):
    subject_code = _norm_subject(subject_code)
    department = _norm_scope_value(department)
    batch = _norm_scope_value(batch)
    for row in _handler_subject_rows(handler_id):
        if _norm_subject(row['subject_code']) != subject_code:
            continue
        if department and not _scope_text_matches(row.get('department'), department):
            continue
        if batch and not _scope_batch_matches(row.get('batch'), batch):
            continue
        return row
    return None

def _handler_has_subject(handler_id, subject_code, department=None, batch=None):
    if _find_handler_subject_scope(handler_id, subject_code, department=department, batch=batch):
        return True
    return False

def _student_ids_for_subject(subject_code):
    subject_code = _norm_subject(subject_code)
    ids = set()
    ids.update(m.student_id for m in StudentMark.query.filter(func.upper(StudentMark.subject_code) == subject_code).all())
    ids.update(m.student_id for m in SubjectHandlerMark.query.filter(func.upper(SubjectHandlerMark.subject_code) == subject_code).all())
    ids.update(a.student_id for a in SubjectHandlerAttendance.query.filter(func.upper(SubjectHandlerAttendance.subject_code) == subject_code).all())
    ids.update(e.student_id for e in SubjectAcademicEntry.query.filter(func.upper(SubjectAcademicEntry.subject_code) == subject_code).all())
    return {_norm_student(sid) for sid in ids if sid}

def _student_ids_for_handler_scope(handler_id, subject_code, department=None, batch=None):
    scope = _find_handler_subject_scope(handler_id, subject_code, department=department, batch=batch)
    if not scope:
        return set()

    scope_department = _norm_scope_value(scope.get('department'))
    scope_batch = _norm_scope_value(scope.get('batch'))

    # Start from all students and apply scope filters in Python so we can do tolerant matching.
    students = Student.query.all()

    if scope_department:
        dept_filtered = [s for s in students if _scope_department_matches(s.branch, scope_department)]
        # Fallback 1: some deployments keep the authoritative department on the Faculty profile
        # while timetable has a different label.
        if not dept_filtered:
            faculty = Faculty.query.get(handler_id)
            if faculty and getattr(faculty, 'department', None):
                dept_filtered = [s for s in students if _scope_department_matches(s.branch, faculty.department)]
        # If we still can't map the department, avoid leaking cross-department data.
        # When a batch is provided, it is usually specific enough to show the right cohort,
        # so we allow the batch filter to decide.
        if dept_filtered:
            students = dept_filtered
        elif not scope_batch:
            return set()

    if scope_batch:
        requested_norm = _normalize_batch_label(scope_batch)
        requested_years = _extract_batch_years(scope_batch)

        def _batch_matches(student_batch):
            student_norm = _normalize_batch_label(student_batch)
            if not student_norm:
                return False
            if (
                student_norm == requested_norm
                or student_norm.endswith(f" {requested_norm}")
                or requested_norm.endswith(f" {student_norm}")
            ):
                return True
            if requested_years and _extract_batch_years(student_batch) == requested_years:
                return True
            return False

        students = [s for s in students if _batch_matches(s.batch)]

    # Keep the handler view limited to the exact allocated batch/department scope.
    # Falling back to subject-wide rows leaks students from other allocations when
    # the same subject code is reused across multiple batches.
    return {_norm_student(s.admission_number) for s in students}

def _students_for_handler_allocations(handler_id):
    scoped_students = {}
    for row in _handler_subject_rows(handler_id):
        student_ids = _student_ids_for_handler_scope(
            handler_id,
            row.get('subject_code'),
            department=row.get('department'),
            batch=row.get('batch')
        )
        if not student_ids:
            continue
        students = Student.query.filter(Student.admission_number.in_(list(student_ids))).all()
        for student in students:
            sid = _norm_student(student.admission_number)
            scoped_students[sid] = student
    return list(scoped_students.values())

def _log_subject_change(handler_id, subject_code, action, entity, student_id=None, details=None):
    try:
        entry = SubjectDataAuditLog(
            subject_handler_id=handler_id,
            subject_code=_norm_subject(subject_code),
            action=action,
            entity=entity,
            student_id=_norm_student(student_id) if student_id else None,
            details=details
        )
        db.session.add(entry)
    except Exception:
        # Logging failures should not block primary action
        pass

def _attendance_summary_map(subject_code):
    subject_code = _norm_subject(subject_code)
    summary = {}
    rows = SubjectHandlerAttendance.query.filter(func.upper(SubjectHandlerAttendance.subject_code) == subject_code).all()
    for row in rows:
        sid = _norm_student(row.student_id)
        if sid not in summary:
            summary[sid] = {'total_classes': 0, 'present_count': 0}
        summary[sid]['total_classes'] += 1
        if str(row.status).strip().lower() in ('present', 'late'):
            summary[sid]['present_count'] += 1
    for sid in summary:
        total = summary[sid]['total_classes']
        present = summary[sid]['present_count']
        summary[sid]['attendance_percent'] = round((present / total) * 100, 2) if total else 0.0
    return summary

def _build_subject_ai_payload(subject_code):
    subject_code = _norm_subject(subject_code)
    mark_rows = SubjectHandlerMark.query.filter(func.upper(SubjectHandlerMark.subject_code) == subject_code).all()
    acad_rows = SubjectAcademicEntry.query.filter(func.upper(SubjectAcademicEntry.subject_code) == subject_code).all()
    attendance = _attendance_summary_map(subject_code)

    marks_map = {}
    for m in mark_rows:
        sid = _norm_student(m.student_id)
        marks_map.setdefault(sid, []).append(float(m.marks_obtained))

    assign_map = {}
    for a in acad_rows:
        assign_map[_norm_student(a.student_id)] = bool(a.assignment_submitted)

    student_ids = set(marks_map.keys()) | set(attendance.keys()) | set(assign_map.keys()) | _student_ids_for_subject(subject_code)
    students_payload = []
    for sid in sorted(student_ids):
        students_payload.append({
            'student_id': sid,
            'marks': marks_map.get(sid, []),
            'attendance_percent': attendance.get(sid, {}).get('attendance_percent', 0.0),
            'assignment_submitted': assign_map.get(sid, False),
        })

    return {
        'subject_code': subject_code,
        'students': students_payload
    }

def _trigger_subject_analysis(subject_code):
    payload = _build_subject_ai_payload(subject_code)
    try:
        with app.test_client() as client:
            client.post('/api/ai/ingest-subject-data', json=payload)
            client.post('/api/risk/predict', json=payload)
    except Exception:
        pass

@app.route('/api/handler/my-subjects/<int:handler_id>', methods=['GET'])
def api_handler_my_subjects(handler_id):
    try:
        return jsonify({'success': True, 'data': _handler_subject_rows(handler_id)}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/students', methods=['GET'])
def api_handler_students():
    try:
        handler_id = int(request.args.get('handler_id', 0))
        subject_code = _norm_subject(request.args.get('subject_code'))
        department = _norm_scope_value(request.args.get('department'))
        batch = _norm_scope_value(request.args.get('batch'))
        if not handler_id:
            return jsonify({'success': False, 'message': 'handler_id is required'}), 400

        if subject_code:
            if not _handler_has_subject(handler_id, subject_code, department=department, batch=batch):
                return jsonify({'success': False, 'message': 'Access denied for this subject'}), 403

            known_ids = _student_ids_for_handler_scope(handler_id, subject_code, department=department, batch=batch)
            students = Student.query.filter(Student.admission_number.in_(list(known_ids))).all() if known_ids else []
        else:
            students = _students_for_handler_allocations(handler_id)

        attendance_map = _attendance_summary_map(subject_code)
        result = []
        for s in students:
            sid = _norm_student(s.admission_number)
            summary = attendance_map.get(sid, {'total_classes': 0, 'attendance_percent': 0.0})
            result.append({
                'admission_number': sid,
                'full_name': s.full_name,
                'email': s.email,
                'batch': s.batch,
                'branch': s.branch,
                'total_classes': summary['total_classes'],
                'attendance_pct': summary['attendance_percent'],
                'attendance_warning': summary['attendance_percent'] < 75 if summary['total_classes'] > 0 else False,
            })
        result.sort(key=lambda r: r['admission_number'])
        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def _parse_marks_rows_from_file(file_storage):
    import csv
    import io

    filename = (file_storage.filename or '').lower()
    if filename.endswith('.csv'):
        text = file_storage.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        from openpyxl import load_workbook
        workbook = load_workbook(file_storage, data_only=True)
        sheet = workbook.active
        headers = []
        rows = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if idx == 1:
                headers = [str(v or '').strip() for v in row]
                continue
            if not any(v is not None and str(v).strip() != '' for v in row):
                continue
            rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return rows
    raise ValueError('Unsupported file type. Upload CSV or Excel file.')

def _upsert_handler_marks(rows, handler_id, subject_code, allowed_student_ids=None):
    updated = 0
    errors = []
    for idx, row in enumerate(rows, start=1):
        student_id = _norm_student(row.get('student_id') or row.get('admission_number'))
        exam_type = str(row.get('exam_type') or '').strip() or 'Quiz'
        max_marks_raw = row.get('max_marks', 100)
        marks_raw = row.get('marks_obtained')

        try:
            max_marks = float(max_marks_raw)
            marks_obtained = float(marks_raw)
            if max_marks <= 0:
                raise ValueError('max_marks must be greater than zero')
            if marks_obtained < 0 or marks_obtained > max_marks:
                raise ValueError(f'marks_obtained must be between 0 and {max_marks}')
        except Exception as ve:
            errors.append(f'Row {idx}: invalid marks - {ve}')
            continue

        student = Student.query.get(student_id)
        if not student:
            errors.append(f'Row {idx}: student {student_id} not found')
            continue
        if allowed_student_ids is not None and student_id not in allowed_student_ids:
            errors.append(f'Row {idx}: student {student_id} is not part of the selected subject allocation')
            continue

        record = SubjectHandlerMark.query.filter_by(
            student_id=student_id,
            subject_code=subject_code,
            exam_type=exam_type
        ).first()
        if not record:
            record = SubjectHandlerMark(
                student_id=student_id,
                subject_code=subject_code,
                exam_type=exam_type,
                marks_obtained=marks_obtained,
                max_marks=max_marks,
                subject_handler_id=handler_id
            )
            db.session.add(record)
        else:
            record.marks_obtained = marks_obtained
            record.max_marks = max_marks
            record.subject_handler_id = handler_id
        updated += 1
        _log_subject_change(
            handler_id,
            subject_code,
            'upsert',
            'marks',
            student_id=student_id,
            details=f'exam_type={exam_type}, marks={marks_obtained}, max={max_marks}'
        )

    return updated, errors

@app.route('/api/handler/marks/upload', methods=['POST'])
def api_handler_upload_marks():
    try:
        payload = request.get_json(silent=True) or {}
        handler_id_raw = request.form.get('handler_id') or request.form.get('subject_handler_id')
        if not handler_id_raw and request.is_json:
            handler_id_raw = payload.get('handler_id')
        handler_id = int(handler_id_raw or 0)
        subject_code = _norm_subject(request.form.get('subject_code') or payload.get('subject_code'))
        department = _norm_scope_value(request.form.get('department') or payload.get('department'))
        batch = _norm_scope_value(request.form.get('batch') or payload.get('batch'))
        if not handler_id or not subject_code:
            return jsonify({'success': False, 'message': 'handler_id and subject_code are required'}), 400
        if not _handler_has_subject(handler_id, subject_code, department=department, batch=batch):
            return jsonify({'success': False, 'message': 'Access denied for this subject'}), 403

        rows = []
        if 'file' in request.files:
            rows = _parse_marks_rows_from_file(request.files['file'])
        else:
            rows = payload.get('rows') or []

        if not rows:
            return jsonify({'success': False, 'message': 'No marks rows found'}), 400

        allowed_student_ids = _student_ids_for_handler_scope(handler_id, subject_code, department=department, batch=batch)
        updated, errors = _upsert_handler_marks(rows, handler_id, subject_code, allowed_student_ids=allowed_student_ids)
        db.session.commit()
        _trigger_subject_analysis(subject_code)
        return jsonify({
            'success': True,
            'updated': updated,
            'errors': errors,
            'message': f'Processed {updated} marks record(s)'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============= MENTOR DASHBOARD (UNIFIED API CONTRACT) =============
def _risk_band(score):
    score = float(score or 0)
    if score >= 80:
        return "High"
    if score >= 50:
        return "Medium"
    return "Low"


def _student_attendance_percent(student_id):
    rows = Attendance.query.filter_by(student_admission_number=student_id).all()
    total = sum(int(r.total_classes or 0) for r in rows)
    attended = sum(int(r.attended_classes or 0) for r in rows)
    return round((attended / total) * 100, 2) if total else 0.0


def _subject_risk_payload(student_id):
    marks_rows = StudentMark.query.filter_by(student_id=student_id).all()
    attendance_rows = Attendance.query.filter_by(student_admission_number=student_id).all()
    attendance_map = {}
    for a in attendance_rows:
        code = str(a.subject_code or a.subject_name or '').strip().upper()
        if not code:
            continue
        attendance_map[code] = float(a.percentage or 0.0)

    subject_data = {}
    for m in marks_rows:
        code = str(m.subject_code or '').strip().upper()
        if not code:
            continue
        vals = [v for v in (m.internal1, m.internal2, m.internal3) if v is not None]
        if code not in subject_data:
            subject_data[code] = {'marks': [], 'semester': m.semester}
        subject_data[code]['marks'].extend([float(v) for v in vals])
        subject_data[code]['semester'] = m.semester

    subject_risks = []
    for code, payload in subject_data.items():
        marks = payload['marks']
        avg_marks = (sum(marks) / len(marks)) if marks else 0.0
        att = attendance_map.get(code, _student_attendance_percent(student_id))
        risk_score = 0.0
        if avg_marks < 40:
            risk_score += 45
        elif avg_marks < 55:
            risk_score += 25
        if att < 65:
            risk_score += 40
        elif att < 75:
            risk_score += 20
        if len(marks) >= 2 and marks[-1] < marks[0]:
            risk_score += 15
        risk_score = min(100.0, risk_score)
        reasons = []
        if att < 65:
            reasons.append("Attendance below 65%")
        elif att < 75:
            reasons.append("Attendance below 75%")
        if avg_marks < 40:
            reasons.append("Consecutive low internal performance")
        elif avg_marks < 55:
            reasons.append("Dropping or weak internal scores")
        subject_risks.append({
            'subject_code': code,
            'risk_score': round(risk_score, 2),
            'risk_level': _risk_band(risk_score),
            'attendance_percent': round(att, 2),
            'avg_marks': round(avg_marks, 2),
            'explanation': " + ".join(reasons) if reasons else "Stable trend",
        })

    subject_risks.sort(key=lambda x: x['risk_score'], reverse=True)
    return subject_risks


def _mentor_students_core(mentor_id):
    students = Student.query.filter_by(mentor_id=mentor_id, status='Live').all()
    data = []
    for s in students:
        sa = StudentAnalytics.query.filter_by(student_id=s.admission_number).first()
        attendance_percent = _student_attendance_percent(s.admission_number)
        risk_score = float(sa.adjusted_risk if sa and sa.adjusted_risk is not None else (sa.risk_score if sa else 0.0))
        pending_interventions = MentorIntervention.query.filter_by(
            student_id=s.admission_number,
            mentor_id=mentor_id,
            escalated=False
        ).count()
        data.append({
            'student_id': s.admission_number,
            'student_name': s.full_name,
            'department': s.branch,
            'batch': s.batch,
            'attendance_percent': round(attendance_percent, 2),
            'risk_score': round(risk_score, 2),
            'risk_level': _risk_band(risk_score),
            'pending_interventions': pending_interventions,
        })
    data.sort(key=lambda x: x['risk_score'], reverse=True)
    return data


@app.route('/api/mentor/students', methods=['GET'])
def api_mentor_students_v2():
    try:
        mentor_id = request.args.get('mentor_id') or session.get('user_id')
        if not mentor_id:
            return jsonify({'success': False, 'message': 'mentor_id is required'}), 400
        data = _mentor_students_core(int(mentor_id))
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/risk/student/<string:student_id>', methods=['GET'])
def api_risk_student_v2(student_id):
    try:
        sid = str(student_id or '').strip().upper()
        student = Student.query.get(sid)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        sa = StudentAnalytics.query.filter_by(student_id=sid).first()
        overall_score = float(sa.adjusted_risk if sa and sa.adjusted_risk is not None else (sa.risk_score if sa else 0.0))
        subject_risks = _subject_risk_payload(sid)
        top_reason = subject_risks[0]['explanation'] if subject_risks else "No risk indicators"
        return jsonify({
            'success': True,
            'data': {
                'student_id': sid,
                'overall_risk_score': round(overall_score, 2),
                'overall_risk_level': _risk_band(overall_score),
                'subject_risks': subject_risks,
                'ai_explanation': top_reason,
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/alerts/mentor/<int:mentor_id>', methods=['GET'])
def api_mentor_alerts_v2(mentor_id):
    try:
        unread_only = str(request.args.get('unread', 'true')).lower() in ('1', 'true', 'yes')
        q = Alert.query.filter_by(mentor_id=mentor_id)
        if unread_only:
            q = q.filter_by(is_read=False)
        alerts = q.order_by(Alert.created_at.desc()).limit(200).all()
        data = []
        for a in alerts:
            data.append({
                'id': a.id,
                'student_id': a.student_admission_number,
                'type': a.type,
                'message': a.message,
                'created_at': a.created_at.isoformat() if a.created_at else None,
                'is_read': bool(a.is_read),
            })
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/alerts/mark-read', methods=['POST'])
def api_alerts_mark_read_v2():
    try:
        payload = request.get_json(force=True) or {}
        alert_ids = payload.get('alert_ids') or []
        if payload.get('alert_id'):
            alert_ids.append(payload.get('alert_id'))
        if not alert_ids:
            return jsonify({'success': False, 'message': 'alert_ids or alert_id is required'}), 400
        updated = 0
        for aid in alert_ids:
            alert = Alert.query.get(int(aid))
            if not alert:
                continue
            alert.is_read = True
            updated += 1
        db.session.commit()
        return jsonify({'success': True, 'updated': updated}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sessions', methods=['POST'])
def api_sessions_create_v2():
    try:
        payload = request.get_json(force=True) or {}
        student_id = str(payload.get('student_id') or '').strip().upper()
        mentor_id = int(payload.get('mentor_id') or 0)
        date_str = payload.get('date')
        time_slot = payload.get('time_slot')
        duration = int(payload.get('duration') or 30)
        mode = str(payload.get('mode') or 'Offline').strip()
        topic = str(payload.get('topic') or '').strip()
        notes = str(payload.get('notes') or '').strip()
        if not all([student_id, mentor_id, date_str, time_slot]):
            return jsonify({'success': False, 'message': 'student_id, mentor_id, date, time_slot required'}), 400
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        if student.mentor_id != mentor_id:
            return jsonify({'success': False, 'message': 'Student is not assigned to this mentor'}), 403
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
        existing = MentoringSession.query.filter_by(mentor_id=mentor_id, date=d, time_slot=time_slot).filter(
            MentoringSession.status.in_(['Pending', 'Approved'])
        ).first()
        if existing:
            return jsonify({'success': False, 'message': 'Mentor already booked for this slot'}), 409
        student_existing = MentoringSession.query.filter_by(student_admission_number=student_id, date=d, time_slot=time_slot).filter(
            MentoringSession.status.in_(['Pending', 'Approved'])
        ).first()
        if student_existing:
            return jsonify({'success': False, 'message': 'Student already booked for this slot'}), 409

        is_online = mode.lower() == 'online'
        # We can't mint a unique Google Meet URL without Google Calendar/Meet API OAuth.
        # `https://meet.google.com/new` is a safe default link that generates a meeting when opened.
        meeting_link = str(payload.get('meeting_link') or '').strip()
        if is_online and not meeting_link:
            meeting_link = 'https://meet.google.com/new'

        session_row = MentoringSession(
            student_admission_number=student_id,
            mentor_id=mentor_id,
            date=d,
            time_slot=time_slot,
            session_type='Online' if mode.lower() == 'online' else 'Offline',
            status='Approved',
            meeting_link=meeting_link if is_online else '',
            notes=f"[{duration} min] {topic}\n{notes}".strip()
        )
        db.session.add(session_row)

        # Notify student with easy join link
        msg_lines = [
            f"Session scheduled on {date_str} at {time_slot}.",
            f"Mode: {'Online' if is_online else 'Offline'}.",
        ]
        if topic:
            msg_lines.append(f"Topic: {topic}.")
        if is_online and meeting_link:
            msg_lines.append(f"Join: {meeting_link}")
        if notes:
            msg_lines.append(f"Notes: {notes}")

        db.session.add(Notification(
            student_id=student_id,
            title="Mentoring Session Scheduled",
            message="\n".join(msg_lines),
            type="session",
            is_read=False,
        ))

        db.session.commit()
        return jsonify({'success': True, 'data': {'id': session_row.id}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sessions/mentor/<int:mentor_id>', methods=['GET'])
def api_sessions_mentor_v2(mentor_id):
    try:
        export_format = str(request.args.get('format') or '').lower()
        q = MentoringSession.query.filter_by(mentor_id=mentor_id)
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        if from_date:
            q = q.filter(MentoringSession.date >= datetime.strptime(from_date, '%Y-%m-%d').date())
        if to_date:
            q = q.filter(MentoringSession.date <= datetime.strptime(to_date, '%Y-%m-%d').date())
        rows = q.order_by(MentoringSession.date.desc(), MentoringSession.time_slot.desc()).all()
        data = []
        for s in rows:
            data.append({
                'id': s.id,
                'date': s.date.isoformat() if s.date else None,
                'time_slot': s.time_slot,
                'duration': 60 if '[60 min]' in str(s.notes or '') else 30,
                'mode': s.session_type,
                'topic': str(s.notes or '').split('\n')[0].replace('[30 min] ', '').replace('[60 min] ', ''),
                'status': s.status,
                'student_id': s.student_admission_number,
                'student_name': s.student.full_name if s.student else s.student_admission_number,
                'notes': s.notes,
                'meeting_link': s.meeting_link,
            })
        if export_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['id', 'date', 'time_slot', 'duration', 'mode', 'topic', 'status', 'student_id', 'student_name', 'notes'])
            for row in data:
                writer.writerow([
                    row['id'],
                    row['date'],
                    row['time_slot'],
                    row['duration'],
                    row['mode'],
                    row['topic'],
                    row['status'],
                    row['student_id'],
                    row['student_name'],
                    row['notes'],
                ])
            csv_data = output.getvalue()
            output.close()
            return Response(
                csv_data,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=mentor_{mentor_id}_sessions.csv'}
            )
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/sessions/<int:session_id>/approve', methods=['PUT'])
def api_sessions_approve_v2(session_id):
    try:
        payload = request.get_json(force=True) or {}
        action = str(payload.get('action') or 'approve').strip().lower()
        reason = str(payload.get('reason') or '').strip()
        row = MentoringSession.query.get(session_id)
        if not row:
            return jsonify({'success': False, 'message': 'Session not found'}), 404
        if action == 'approve':
            row.status = 'Approved'
        elif action == 'reject':
            row.status = 'Rejected'
        elif action == 'cancel':
            row.status = 'Cancelled'
        elif action == 'complete':
            row.status = 'Completed'
        else:
            return jsonify({'success': False, 'message': 'Invalid action'}), 400
        if reason:
            row.notes = f"{row.notes or ''}\n[Action note] {reason}".strip()
        db.session.commit()
        return jsonify({'success': True, 'message': f'Session {action}d'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/interventions', methods=['POST'])
def api_interventions_create_v2():
    try:
        payload = request.get_json(force=True) or {}
        student_id = str(payload.get('student_id') or '').strip().upper()
        mentor_id = int(payload.get('mentor_id') or 0)
        intervention_type = str(payload.get('intervention_type') or 'Academic').strip()
        notes = str(payload.get('notes') or '').strip()
        date_str = payload.get('date')
        if not student_id or not mentor_id:
            return jsonify({'success': False, 'message': 'student_id and mentor_id required'}), 400
        d = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        if student.mentor_id != mentor_id:
            return jsonify({'success': False, 'message': 'Unauthorized mentor for this student'}), 403
        sa = StudentAnalytics.query.filter_by(student_id=student_id).first()
        row = MentorIntervention(
            student_id=student_id,
            mentor_id=mentor_id,
            week_start=d,
            risk_snapshot=float(sa.adjusted_risk if sa and sa.adjusted_risk is not None else 0.0),
            intervention_type=intervention_type,
            notes=notes,
            locked=True
        )
        db.session.add(row)
        db.session.commit()
        return jsonify({'success': True, 'data': {'id': row.id}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/interventions/student/<string:student_id>', methods=['GET'])
def api_interventions_student_v2(student_id):
    try:
        sid = str(student_id or '').strip().upper()
        rows = MentorIntervention.query.filter_by(student_id=sid).order_by(MentorIntervention.created_at.desc()).all()
        return jsonify({'success': True, 'data': [{
            'id': r.id,
            'student_id': r.student_id,
            'mentor_id': r.mentor_id,
            'intervention_type': r.intervention_type,
            'notes': r.notes,
            'date': r.week_start.isoformat() if r.week_start else None,
            'risk_snapshot': r.risk_snapshot,
            'created_at': r.created_at.isoformat() if r.created_at else None
        } for r in rows]}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/interventions/impact/<int:intervention_id>', methods=['GET'])
def api_interventions_impact_v2(intervention_id):
    try:
        row = MentorIntervention.query.get(intervention_id)
        if not row:
            return jsonify({'success': False, 'message': 'Intervention not found'}), 404
        marks = StudentMark.query.filter_by(student_id=row.student_id).all()
        attendance = Attendance.query.filter_by(student_admission_number=row.student_id).all()
        pre_marks = [float(v) for m in marks for v in (m.internal1, m.internal2) if v is not None]
        post_marks = [float(v) for m in marks for v in (m.internal3,) if v is not None]
        pre_mark_avg = round(sum(pre_marks) / len(pre_marks), 2) if pre_marks else 0.0
        post_mark_avg = round(sum(post_marks) / len(post_marks), 2) if post_marks else pre_mark_avg
        pre_att = round(sum(float(a.percentage or 0) for a in attendance) / len(attendance), 2) if attendance else 0.0
        post_att = pre_att
        sa = StudentAnalytics.query.filter_by(student_id=row.student_id).first()
        current_risk = float(sa.adjusted_risk if sa and sa.adjusted_risk is not None else 0.0)
        return jsonify({'success': True, 'data': {
            'intervention_id': row.id,
            'student_id': row.student_id,
            'before': {
                'marks_avg': pre_mark_avg,
                'attendance_percent': pre_att,
                'risk_score': float(row.risk_snapshot or 0.0),
            },
            'after': {
                'marks_avg': post_mark_avg,
                'attendance_percent': post_att,
                'risk_score': current_risk,
            },
            'risk_delta': round(current_risk - float(row.risk_snapshot or 0.0), 2)
        }}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/chat/<string:student_id>', methods=['GET'])
def api_chat_student_v2(student_id):
    try:
        sid = str(student_id or '').strip().upper()
        mentor_id = request.args.get('mentor_id') or session.get('user_id')
        if not mentor_id:
            return jsonify({'success': False, 'message': 'mentor_id is required'}), 400
        msgs = MentorMessage.query.filter_by(student_id=sid, mentor_id=int(mentor_id)).order_by(MentorMessage.sent_at.asc()).all()
        return jsonify({'success': True, 'data': [{
            'id': m.id,
            'student_id': m.student_id,
            'mentor_id': m.mentor_id,
            'message': m.message,
            'sender_role': m.sender_role,
            'sent_at': m.sent_at.isoformat() if m.sent_at else None,
            'is_read': bool(m.is_read),
        } for m in msgs]}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/chat/send', methods=['POST'])
def api_chat_send_v2():
    try:
        payload = request.get_json(force=True) or {}
        sid = str(payload.get('student_id') or '').strip().upper()
        mentor_id = int(payload.get('mentor_id') or 0)
        message = str(payload.get('message') or '').strip()
        sender_role = str(payload.get('sender_role') or 'mentor').strip().lower()
        if not sid or not mentor_id or not message:
            return jsonify({'success': False, 'message': 'student_id, mentor_id, message required'}), 400
        msg = MentorMessage(
            student_id=sid,
            mentor_id=mentor_id,
            message=message,
            sender_role='mentor' if sender_role not in ('mentor', 'student') else sender_role,
            is_read=False
        )
        db.session.add(msg)
        db.session.commit()
        return jsonify({'success': True, 'data': {'id': msg.id}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/chat/unread-count', methods=['GET'])
def api_chat_unread_count_v2():
    try:
        mentor_id = request.args.get('mentor_id') or session.get('user_id')
        if not mentor_id:
            return jsonify({'success': False, 'message': 'mentor_id is required'}), 400
        count = MentorMessage.query.filter_by(mentor_id=int(mentor_id), sender_role='student', is_read=False).count()
        return jsonify({'success': True, 'data': {'unread_count': count}}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/reports/mentor/<int:mentor_id>', methods=['GET'])
def api_reports_mentor_v2(mentor_id):
    try:
        period = str(request.args.get('period') or 'month').lower()
        students = _mentor_students_core(mentor_id)
        total_students = len(students)
        high_risk = [s for s in students if s['risk_level'] == 'High']
        month_start = date.today().replace(day=1)
        sessions_q = MentoringSession.query.filter(
            MentoringSession.mentor_id == mentor_id,
            MentoringSession.date >= month_start
        )
        sessions_held = sessions_q.count()
        interventions = MentorIntervention.query.filter_by(mentor_id=mentor_id).all()
        improved = 0
        for iv in interventions:
            sa = StudentAnalytics.query.filter_by(student_id=iv.student_id).first()
            current_risk = float(sa.adjusted_risk if sa and sa.adjusted_risk is not None else 0.0)
            if iv.risk_snapshot is not None and current_risk < float(iv.risk_snapshot):
                improved += 1
        success_rate = round((improved / len(interventions)) * 100, 2) if interventions else 0.0

        progress_rows = []
        for s in students:
            sid = s['student_id']
            sa = StudentAnalytics.query.filter_by(student_id=sid).first()
            initial = 0.0
            first_iv = MentorIntervention.query.filter_by(student_id=sid, mentor_id=mentor_id).order_by(MentorIntervention.created_at.asc()).first()
            if first_iv and first_iv.risk_snapshot is not None:
                initial = float(first_iv.risk_snapshot)
            current = float(sa.adjusted_risk if sa and sa.adjusted_risk is not None else s['risk_score'])
            marks = StudentMark.query.filter_by(student_id=sid).all()
            vals = [float(v) for m in marks for v in (m.internal1, m.internal2, m.internal3) if v is not None]
            marks_change = 0.0
            if len(vals) >= 2:
                mid = max(1, len(vals) // 2)
                before = sum(vals[:mid]) / len(vals[:mid])
                after = sum(vals[mid:]) / len(vals[mid:])
                marks_change = round(after - before, 2)
            last_iv = MentorIntervention.query.filter_by(student_id=sid, mentor_id=mentor_id).order_by(MentorIntervention.created_at.desc()).first()
            progress_rows.append({
                'student_id': sid,
                'student_name': s['student_name'],
                'initial_risk': round(initial, 2),
                'current_risk': round(current, 2),
                'attendance_change': 0.0,
                'marks_change': marks_change,
                'last_intervention_date': last_iv.created_at.isoformat() if last_iv and last_iv.created_at else None,
                'improvement_score': round(initial - current, 2),
            })

        progress_rows.sort(key=lambda x: x['improvement_score'], reverse=True)

        return jsonify({'success': True, 'data': {
            'period': period,
            'summary': {
                'total_students_assigned': total_students,
                'high_risk_students': len(high_risk),
                'sessions_held_this_month': sessions_held,
                'intervention_success_rate': success_rate,
            },
            'students_progress': progress_rows
        }}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def _date_range(start_dt, end_dt):
    cur = start_dt
    while cur <= end_dt:
        yield cur
        cur = cur + timedelta(days=1)

@app.route('/api/handler/attendance/upload', methods=['POST'])
def api_handler_upload_attendance():
    try:
        data = request.get_json(force=True) or {}
        handler_id = int(data.get('handler_id') or 0)
        subject_code = _norm_subject(data.get('subject_code'))
        department = _norm_scope_value(data.get('department'))
        batch = _norm_scope_value(data.get('batch'))
        entries = data.get('entries') or []
        if not handler_id or not subject_code or not entries:
            return jsonify({'success': False, 'message': 'handler_id, subject_code, and entries are required'}), 400
        if not _handler_has_subject(handler_id, subject_code, department=department, batch=batch):
            return jsonify({'success': False, 'message': 'Access denied for this subject'}), 403

        from datetime import datetime as _dt
        start_date = data.get('start_date') or data.get('date')
        end_date = data.get('end_date') or start_date
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'date or start_date/end_date is required'}), 400

        start_dt = _dt.strptime(start_date, '%Y-%m-%d').date()
        end_dt = _dt.strptime(end_date, '%Y-%m-%d').date()
        if end_dt < start_dt:
            return jsonify({'success': False, 'message': 'end_date must be on or after start_date'}), 400

        valid_status = {'present', 'absent', 'late'}
        updated = 0
        errors = []
        allowed_student_ids = _student_ids_for_handler_scope(handler_id, subject_code, department=department, batch=batch)
        for rec in entries:
            student_id = _norm_student(rec.get('student_id'))
            status = str(rec.get('status') or '').strip().lower()
            if not student_id:
                errors.append('Missing student_id in one row')
                continue
            if status not in valid_status:
                errors.append(f'Invalid status for {student_id}: {status}')
                continue
            student = Student.query.get(student_id)
            if not student:
                errors.append(f'Student not found: {student_id}')
                continue
            if student_id not in allowed_student_ids:
                errors.append(f'Student {student_id} is not part of the selected subject allocation')
                continue
            for date_val in _date_range(start_dt, end_dt):
                item = SubjectHandlerAttendance.query.filter_by(
                    student_id=student_id,
                    subject_code=subject_code,
                    date=date_val
                ).first()
                if not item:
                    item = SubjectHandlerAttendance(
                        student_id=student_id,
                        subject_code=subject_code,
                        date=date_val,
                        status=status.capitalize(),
                        subject_handler_id=handler_id
                    )
                    db.session.add(item)
                else:
                    item.status = status.capitalize()
                    item.subject_handler_id = handler_id
                updated += 1
                _log_subject_change(
                    handler_id,
                    subject_code,
                    'upsert',
                    'attendance',
                    student_id=student_id,
                    details=f'date={date_val.isoformat()}, status={status}'
                )

        db.session.commit()
        _trigger_subject_analysis(subject_code)

        summary = _attendance_summary_map(subject_code)
        students = [{
            'student_id': sid,
            'total_classes': v['total_classes'],
            'attendance_percent': v['attendance_percent'],
            'warning': v['attendance_percent'] < 75
        } for sid, v in summary.items() if sid in allowed_student_ids]

        return jsonify({
            'success': True,
            'updated': updated,
            'errors': errors,
            'summary': students
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/attendance/summary', methods=['GET'])
def api_handler_attendance_summary():
    try:
        handler_id = int(request.args.get('handler_id', 0))
        subject_code = _norm_subject(request.args.get('subject_code'))
        department = _norm_scope_value(request.args.get('department'))
        batch = _norm_scope_value(request.args.get('batch'))
        if not handler_id or not subject_code:
            return jsonify({'success': False, 'message': 'handler_id and subject_code are required'}), 400
        if not _handler_has_subject(handler_id, subject_code, department=department, batch=batch):
            return jsonify({'success': False, 'message': 'Access denied for this subject'}), 403
        allowed_student_ids = _student_ids_for_handler_scope(handler_id, subject_code, department=department, batch=batch)
        summary = _attendance_summary_map(subject_code)
        data = [{
            'student_id': sid,
            'total_classes': row['total_classes'],
            'attendance_percent': row['attendance_percent'],
            'warning': row['attendance_percent'] < 75
        } for sid, row in summary.items() if sid in allowed_student_ids]
        data.sort(key=lambda d: d['student_id'])
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/academic-grid', methods=['GET'])
def api_handler_get_academic_grid():
    try:
        handler_id = int(request.args.get('handler_id', 0))
        subject_code = _norm_subject(request.args.get('subject_code'))
        department = _norm_scope_value(request.args.get('department'))
        batch = _norm_scope_value(request.args.get('batch'))
        if not handler_id or not subject_code:
            return jsonify({'success': False, 'message': 'handler_id and subject_code are required'}), 400
        scope = _find_handler_subject_scope(handler_id, subject_code, department=department, batch=batch)
        if not scope:
            return jsonify({'success': False, 'message': 'Access denied for this subject'}), 403

        student_ids = _student_ids_for_handler_scope(handler_id, subject_code, department=department, batch=batch)

        students = Student.query.filter(Student.admission_number.in_(list(student_ids))).all() if student_ids else []
        entries = SubjectAcademicEntry.query.filter(
            func.upper(SubjectAcademicEntry.subject_code) == subject_code,
            SubjectAcademicEntry.student_id.in_(list(student_ids))
        ).all() if student_ids else []
        by_student = {_norm_student(e.student_id): e for e in entries}

        data = []
        for s in students:
            sid = _norm_student(s.admission_number)
            e = by_student.get(sid)
            data.append({
                'student_id': sid,
                'student_name': s.full_name,
                'internal_assessment_score': None if not e else e.internal_assessment_score,
                'assignment_submitted': False if not e else bool(e.assignment_submitted),
                'practical_lab_score': None if not e else e.practical_lab_score
            })
        data.sort(key=lambda r: r['student_id'])
        return jsonify({
            'success': True,
            'data': data,
            'meta': {
                'subject': scope,
                'academic_limits': ACADEMIC_ENTRY_LIMITS,
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/academic-grid/bulk-update', methods=['POST'])
def api_handler_bulk_update_academic():
    try:
        payload = request.get_json(force=True) or {}
        handler_id = int(payload.get('handler_id') or 0)
        subject_code = _norm_subject(payload.get('subject_code'))
        department = _norm_scope_value(payload.get('department'))
        batch = _norm_scope_value(payload.get('batch'))
        entries = payload.get('entries') or []
        if not handler_id or not subject_code or not entries:
            return jsonify({'success': False, 'message': 'handler_id, subject_code, and entries are required'}), 400
        if not _handler_has_subject(handler_id, subject_code, department=department, batch=batch):
            return jsonify({'success': False, 'message': 'Access denied for this subject'}), 403

        updated = 0
        errors = []
        allowed_student_ids = _student_ids_for_handler_scope(handler_id, subject_code, department=department, batch=batch)
        for row in entries:
            sid = _norm_student(row.get('student_id'))
            if not sid:
                errors.append('One row is missing student_id')
                continue
            if not Student.query.get(sid):
                errors.append(f'Student not found: {sid}')
                continue
            if sid not in allowed_student_ids:
                errors.append(f'Student {sid} is not part of the selected subject allocation')
                continue

            ia = row.get('internal_assessment_score')
            practical = row.get('practical_lab_score')
            assignment = bool(row.get('assignment_submitted', False))
            try:
                ia_val = None if ia in (None, '') else float(ia)
                practical_val = None if practical in (None, '') else float(practical)
            except Exception:
                errors.append(f'Invalid numeric values for {sid}')
                continue
            if ia_val is not None and (ia_val < 0 or ia_val > ACADEMIC_ENTRY_LIMITS['internal_assessment_score']):
                errors.append(f'Internal assessment for {sid} must be between 0 and {ACADEMIC_ENTRY_LIMITS["internal_assessment_score"]}')
                continue
            if practical_val is not None and (practical_val < 0 or practical_val > ACADEMIC_ENTRY_LIMITS['practical_lab_score']):
                errors.append(f'Practical/Lab score for {sid} must be between 0 and {ACADEMIC_ENTRY_LIMITS["practical_lab_score"]}')
                continue

            record = SubjectAcademicEntry.query.filter_by(student_id=sid, subject_code=subject_code).first()
            if not record:
                record = SubjectAcademicEntry(
                    student_id=sid,
                    subject_code=subject_code,
                    subject_handler_id=handler_id
                )
                db.session.add(record)
            record.internal_assessment_score = ia_val
            record.assignment_submitted = assignment
            record.practical_lab_score = practical_val
            record.subject_handler_id = handler_id
            updated += 1
            _log_subject_change(
                handler_id,
                subject_code,
                'upsert',
                'academic_entry',
                student_id=sid,
                details=f'ia={ia_val}, assignment_submitted={assignment}, practical={practical_val}'
            )

        db.session.commit()
        _trigger_subject_analysis(subject_code)
        return jsonify({'success': True, 'updated': updated, 'errors': errors}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/handler/audit-logs', methods=['GET'])
def api_handler_audit_logs():
    try:
        handler_id = int(request.args.get('handler_id', 0))
        subject_code = _norm_subject(request.args.get('subject_code'))
        if not handler_id:
            return jsonify({'success': False, 'message': 'handler_id is required'}), 400
        query = SubjectDataAuditLog.query.filter_by(subject_handler_id=handler_id)
        if subject_code:
            query = query.filter(func.upper(SubjectDataAuditLog.subject_code) == subject_code)
        rows = query.order_by(SubjectDataAuditLog.created_at.desc()).limit(200).all()
        return jsonify({
            'success': True,
            'data': [{
                'id': r.id,
                'subject_code': r.subject_code,
                'action': r.action,
                'entity': r.entity,
                'student_id': r.student_id,
                'details': r.details,
                'created_at': r.created_at.isoformat() if r.created_at else None
            } for r in rows]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/ai/ingest-subject-data', methods=['POST'])
def api_ai_ingest_subject_data():
    try:
        payload = request.get_json(force=True) or {}
        subject_code = _norm_subject(payload.get('subject_code'))
        if not subject_code:
            return jsonify({'success': False, 'message': 'subject_code is required'}), 400

        log = SubjectAnalysisLog(
            subject_code=subject_code,
            endpoint='ai_ingest',
            payload=json.dumps(payload)
        )
        db.session.add(log)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Subject data ingested for AI analysis'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/risk/predict', methods=['POST'])
def api_risk_predict_subject():
    try:
        payload = request.get_json(force=True) or {}
        subject_code = _norm_subject(payload.get('subject_code'))
        students = payload.get('students') or []
        if not subject_code:
            return jsonify({'success': False, 'message': 'subject_code is required'}), 400

        predictions = []
        for s in students:
            sid = _norm_student(s.get('student_id'))
            marks = [float(m) for m in (s.get('marks') or []) if m is not None]
            attendance_percent = float(s.get('attendance_percent') or 0)
            assignment_submitted = bool(s.get('assignment_submitted', False))
            avg_marks = (sum(marks) / len(marks)) if marks else 0

            risk_score = 0
            if attendance_percent < 75:
                risk_score += 40
            if avg_marks < 50:
                risk_score += 40
            if not assignment_submitted:
                risk_score += 20

            if risk_score >= 70:
                level = 'high'
            elif risk_score >= 40:
                level = 'medium'
            else:
                level = 'low'

            predictions.append({
                'student_id': sid,
                'risk_level': level,
                'risk_score': min(100, risk_score),
                'avg_marks': round(avg_marks, 2),
                'attendance_percent': attendance_percent
            })

        log = SubjectAnalysisLog(
            subject_code=subject_code,
            endpoint='risk_predict',
            payload=json.dumps({'subject_code': subject_code, 'students': students, 'predictions': predictions})
        )
        db.session.add(log)
        db.session.commit()
        return jsonify({'success': True, 'subject_code': subject_code, 'predictions': predictions}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


def _current_role():
    role = session.get('user_role') or session.get('role') or ''
    return str(role).strip().lower().replace(' ', '-').replace('_', '-')


def _current_faculty_id():
    """Return faculty id from session or None."""
    fid = session.get('user_id')
    try:
        return int(fid) if fid is not None else None
    except Exception:
        return None


def _require_any_role(*roles: str):
    allowed = {str(r).strip().lower().replace(' ', '-').replace('_', '-') for r in roles}
    role = _current_role()
    if role not in allowed:
        abort(403)


def _normalized_roles(values):
    vals = values or []
    try:
        it = list(vals)
    except Exception:
        it = []
    return [str(v).strip().lower().replace(' ', '-').replace('_', '-') for v in it]


def _auto_switch_role_if_allowed(target_role: str) -> bool:
    """If the user is allowed to use target_role, switch session role keys and return True."""
    target = str(target_role).strip().lower().replace(' ', '-').replace('_', '-')
    allowed_roles = _normalized_roles(session.get('user_roles') or session.get('allowed_roles'))
    if target in allowed_roles:
        session['user_role'] = target
        session['role'] = target
        return True
    return False


def _faculty_has_active_subjects_api(faculty_id: int) -> bool:
    if SubjectAllocation.query.filter_by(faculty_id=faculty_id).first() is not None:
        return True
    if Timetable.query.filter_by(handler_id=faculty_id).first() is not None:
        return True
    return False


def _faculty_allowed_roles_from_db(faculty: Faculty) -> list[str]:
    if faculty.designation and str(faculty.designation).strip().lower() == 'admin':
        return ['admin']

    roles: list[str] = []
    if bool(getattr(faculty, 'is_hod', False)):
        roles.append('hod')

    # Mentor role (eligible or has mentees).
    if bool(getattr(faculty, 'is_mentor_eligible', False)) or bool(getattr(faculty, 'mentees', None)):
        roles.append('mentor')

    # Subject handler role (flagged OR actively handles subjects).
    if bool(getattr(faculty, 'is_subject_handler', False)) or _faculty_has_active_subjects_api(faculty.id):
        roles.append('subject-handler')

    if not roles:
        roles = ['mentor']

    seen: set[str] = set()
    ordered: list[str] = []
    for r in roles:
        r2 = str(r).strip().lower().replace(' ', '-').replace('_', '-')
        if r2 and r2 not in seen:
            ordered.append(r2)
            seen.add(r2)
    return ordered


def _ensure_faculty_roles_in_session():
    """If a faculty is logged in but roles aren't present, infer from DB and store in session."""
    fid = _current_faculty_id()
    if not fid:
        return
    # If roles already present, keep.
    existing = _normalized_roles(session.get('user_roles') or session.get('allowed_roles'))
    if existing:
        return
    faculty = Faculty.query.get(fid)
    if not faculty:
        return
    roles = _faculty_allowed_roles_from_db(faculty)
    session['user_roles'] = roles
    session['allowed_roles'] = roles


def _certificate_access_allowed(admission_number):
    sid = str(admission_number or '').strip().upper()
    role = _current_role()
    user_id = str(session.get('user_id') or '').strip().upper()
    if role == 'mentor':
        mentor_id = _current_faculty_id()
        if not mentor_id:
            return False, sid
        student = Student.query.get(sid)
        if not student or student.mentor_id != mentor_id:
            return False, sid
        return True, sid
    if role in ('subject-handler',):
        return False, sid
    if role == 'student' and user_id and user_id != sid:
        return False, sid
    return True, sid


@app.route('/mentor/dashboard')
def mentor_dashboard():
    _ensure_faculty_roles_in_session()
    # If the user navigates directly, auto-switch if they have access.
    if _current_role() not in ('mentor', 'hod'):
        if not _auto_switch_role_if_allowed('mentor'):
            return redirect('/subject-handler/dashboard')
    _require_any_role('mentor', 'hod')
    mentor_id = _current_faculty_id()
    if not mentor_id:
        return redirect(url_for('index'))

    mentor = Faculty.query.get(mentor_id)
    if not mentor:
        session.clear()
        return redirect(url_for('index'))

    allowed_roles = session.get('user_roles') or session.get('allowed_roles') or ['mentor']
    return render_template(
        'mentor_dashboard.html',
        faculty=mentor,
        active_role=_current_role(),
        allowed_roles=allowed_roles,
    )


@app.route('/subject-handler/dashboard')
def subject_handler_dashboard():
    _ensure_faculty_roles_in_session()
    if _current_role() not in ('subject-handler', 'hod'):
        if not _auto_switch_role_if_allowed('subject-handler'):
            return redirect('/mentor/dashboard')
    _require_any_role('subject-handler', 'hod')
    handler_id = _current_faculty_id()
    if not handler_id:
        return redirect(url_for('index'))

    handler = Faculty.query.get(handler_id)
    if not handler:
        session.clear()
        return redirect(url_for('index'))

    allowed_roles = session.get('user_roles') or session.get('allowed_roles') or ['subject-handler']
    return render_template(
        'subject_handler_dashboard.html',
        faculty=handler,
        active_role=_current_role(),
        allowed_roles=allowed_roles,
    )


@app.route('/subject_handler/dashboard')
def subject_handler_dashboard_legacy():
    return redirect('/subject-handler/dashboard')


@app.route('/mentor/certificates')
def mentor_certificates():
    _ensure_faculty_roles_in_session()
    if _current_role() not in ('mentor', 'hod'):
        _auto_switch_role_if_allowed('mentor')
    _require_any_role('mentor', 'hod')
    mentor_id = _current_faculty_id()
    if not mentor_id:
        return redirect(url_for('index'))

    mentor = Faculty.query.get(mentor_id)
    mentees = Student.query.filter_by(mentor_id=mentor_id).order_by(Student.full_name.asc()).all()
    allowed_roles = session.get('user_roles') or session.get('allowed_roles') or ['mentor']
    return render_template(
        'mentor_certificates.html',
        faculty=mentor,
        mentees=mentees,
        active_role=_current_role(),
        allowed_roles=allowed_roles,
    )


@app.route('/mentor/certificates/<string:admission_number>')
def mentor_certificates_student(admission_number):
    _ensure_faculty_roles_in_session()
    if _current_role() not in ('mentor', 'hod'):
        _auto_switch_role_if_allowed('mentor')
    _require_any_role('mentor', 'hod')
    mentor_id = _current_faculty_id()
    if not mentor_id:
        return redirect(url_for('index'))


@app.route('/hod/dashboard')
def hod_dashboard():
    return redirect('/mentor/dashboard')

    sid = str(admission_number or '').strip().upper()
    student = Student.query.get_or_404(sid)
    if student.mentor_id != mentor_id and _current_role() != 'hod':
        abort(403)

    certs = Certificate.query.filter_by(student_id=sid).order_by(Certificate.uploaded_at.desc()).all()
    mentor = Faculty.query.get(mentor_id)
    allowed_roles = session.get('user_roles') or session.get('allowed_roles') or ['mentor']
    return render_template(
        'mentor_certificates_student.html',
        faculty=mentor,
        student=student,
        certificates=[_format_certificate_row(c) for c in certs],
        active_role=_current_role(),
        allowed_roles=allowed_roles,
    )


def _format_certificate_row(c):
    return {
        'id': c.id,
        'title': c.title or c.activity_name or 'Certificate',
        'issuing_org': c.issuing_org or c.category or '',
        'issue_date': c.issue_date.isoformat() if getattr(c, 'issue_date', None) else (c.date_of_event.isoformat() if c.date_of_event else None),
        'expiry_date': c.expiry_date.isoformat() if getattr(c, 'expiry_date', None) else None,
        'file_url': c.file_path,
        'download_url': f"/static/{c.file_path}" if c.file_path else None,
        'created_at': c.uploaded_at.isoformat() if c.uploaded_at else None,
    }


@app.route('/api/certificates/<string:admission_number>', methods=['GET'])
def api_get_certificates_v2(admission_number):
    try:
        allowed, sid = _certificate_access_allowed(admission_number)
        if not allowed:
            return jsonify({'success': False, 'message': 'Access denied: Mentor cannot view student certificates'}), 403

        certs = Certificate.query.filter_by(student_id=sid).order_by(Certificate.uploaded_at.desc()).all()
        return jsonify({'success': True, 'data': [_format_certificate_row(c) for c in certs]}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/certificates/<string:admission_number>', methods=['POST'])
def api_upload_certificate_v2(admission_number):
    try:
        allowed, sid = _certificate_access_allowed(admission_number)
        if not allowed:
            return jsonify({'success': False, 'message': 'Access denied: Mentor cannot view student certificates'}), 403

        student = Student.query.get(sid)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'success': False, 'message': 'Certificate file is required'}), 400

        save_dir = os.path.join(app.root_path, 'static', 'uploads', 'student_certificates', sid)
        os.makedirs(save_dir, exist_ok=True)
        filename = secure_filename(f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}")
        file_path = os.path.join(save_dir, filename)
        file.save(file_path)
        rel_path = os.path.relpath(file_path, os.path.join(app.root_path, 'static')).replace('\\', '/')

        def _parse_date(value):
            value = str(value or '').strip()
            if not value:
                return None
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except Exception:
                return None

        title = (request.form.get('title') or request.form.get('activity_name') or '').strip()
        issuing_org = (request.form.get('issuing_org') or request.form.get('category') or '').strip()
        issue_date = _parse_date(request.form.get('issue_date') or request.form.get('date_of_event'))
        expiry_date = _parse_date(request.form.get('expiry_date'))

        cert = Certificate(
            student_id=sid,
            title=title,
            issuing_org=issuing_org,
            issue_date=issue_date,
            expiry_date=expiry_date,
            activity_name=title,
            category=issuing_org,
            date_of_event=issue_date,
            file_path=rel_path,
            share_with_mentor=False,
        )
        db.session.add(cert)
        db.session.commit()
        return jsonify({'success': True, 'data': _format_certificate_row(cert)}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/certificates/<string:admission_number>/<int:certificate_id>', methods=['DELETE'])
def api_delete_certificate_v2(admission_number, certificate_id):
    try:
        allowed, sid = _certificate_access_allowed(admission_number)
        if not allowed:
            return jsonify({'success': False, 'message': 'Access denied: Mentor cannot view student certificates'}), 403

        cert = Certificate.query.filter_by(id=certificate_id, student_id=sid).first()
        if not cert:
            return jsonify({'success': False, 'message': 'Certificate not found'}), 404

        if cert.file_path:
            full_path = os.path.join(app.root_path, 'static', cert.file_path.replace('/', os.sep))
            if os.path.exists(full_path):
                try:
                    os.remove(full_path)
                except Exception:
                    pass

        db.session.delete(cert)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Certificate deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


def _semester_gpa(entries):
    if not entries:
        return None
    total = 0.0
    count = 0
    for row in entries:
        total_marks = float(row.total_marks or 0)
        marks = float(row.marks_obtained or 0)
        if total_marks <= 0:
            continue
        total += max(0.0, min(10.0, (marks / total_marks) * 10.0))
        count += 1
    return round(total / count, 2) if count else None


def _group_results_by_semester(rows):
    grouped = {}
    for row in rows:
        grouped.setdefault(int(row.semester), []).append(row)
    payload = []
    for semester in sorted(grouped):
        semester_rows = grouped[semester]
        payload.append({
            'semester': semester,
            'sgpa': _semester_gpa(semester_rows),
            'results': [{
                'id': row.id,
                'subject': row.subject,
                'marks_obtained': row.marks_obtained,
                'total_marks': row.total_marks,
                'percentage': round((float(row.marks_obtained or 0) / float(row.total_marks or 1)) * 100, 2) if row.total_marks else None,
                'status': row.status,
                'result_date': row.result_date.isoformat() if row.result_date else None,
                'verified_by_mentor_id': row.verified_by_mentor_id,
                'verified_at': row.verified_at.isoformat() if row.verified_at else None,
                'mentor_comment': row.mentor_comment,
                'created_at': row.created_at.isoformat() if row.created_at else None,
            } for row in semester_rows]
        })
    return payload


@app.route('/api/results/upload', methods=['POST'])
def api_upload_university_result():
    try:
        role = _current_role()
        student_id = str(session.get('user_id') or request.form.get('student_id') or '').strip().upper()
        if role in ('mentor', 'subject-handler', 'hod', 'admin'):
            return jsonify({'success': False, 'message': 'Student login required'}), 403
        if role == 'student' and student_id and session.get('user_id') and str(session.get('user_id')).strip().upper() != student_id:
            return jsonify({'success': False, 'message': 'Student login required'}), 403

        semester = request.form.get('semester', type=int)
        subject = (request.form.get('subject') or '').strip()
        marks_obtained = request.form.get('marks_obtained', type=float)
        total_marks = request.form.get('total_marks', type=float) or 100
        result_date_raw = request.form.get('result_date') or ''
        if not semester or not subject or marks_obtained is None:
            return jsonify({'success': False, 'message': 'semester, subject, and marks_obtained are required'}), 400

        try:
            result_date = datetime.strptime(result_date_raw, '%Y-%m-%d').date() if result_date_raw else None
        except Exception:
            result_date = None

        result = UniversityResult(
            student_id=student_id,
            semester=semester,
            subject=subject,
            marks_obtained=marks_obtained,
            total_marks=total_marks,
            result_date=result_date,
            status='pending_verification',
        )
        db.session.add(result)
        db.session.commit()
        return jsonify({'success': True, 'data': {'id': result.id, 'status': result.status}}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/results/pending-verification', methods=['GET'])
def api_pending_results():
    try:
        role = _current_role()
        mentor_id = session.get('user_id')
        if role != 'mentor' or not mentor_id:
            return jsonify({'success': False, 'message': 'Mentor access required'}), 403

        student_ids = [s.admission_number for s in Student.query.filter_by(mentor_id=mentor_id).all()]
        rows = UniversityResult.query.filter(
            UniversityResult.student_id.in_(student_ids),
            UniversityResult.status == 'pending_verification'
        ).order_by(UniversityResult.created_at.desc()).all()
        return jsonify({'success': True, 'data': [{
            'id': row.id,
            'student_id': row.student_id,
            'student_name': row.student.full_name if row.student else row.student_id,
            'semester': row.semester,
            'subject': row.subject,
            'marks_obtained': row.marks_obtained,
            'total_marks': row.total_marks,
            'result_date': row.result_date.isoformat() if row.result_date else None,
            'status': row.status,
            'created_at': row.created_at.isoformat() if row.created_at else None,
        } for row in rows]}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/results/<int:result_id>/verify', methods=['PUT'])
def api_verify_university_result(result_id):
    try:
        role = _current_role()
        mentor_id = session.get('user_id')
        if role != 'mentor' or not mentor_id:
            return jsonify({'success': False, 'message': 'Mentor access required'}), 403

        result = UniversityResult.query.get(result_id)
        if not result:
            return jsonify({'success': False, 'message': 'Result not found'}), 404
        if not result.student or result.student.mentor_id != mentor_id:
            return jsonify({'success': False, 'message': 'Access denied'}), 403

        data = request.get_json(silent=True) or {}
        action = str(data.get('action') or 'verify').lower()
        comment = (data.get('comment') or '').strip()
        if action not in ('verify', 'reject'):
            return jsonify({'success': False, 'message': 'Invalid action'}), 400

        result.status = 'verified' if action == 'verify' else 'rejected'
        result.verified_by_mentor_id = mentor_id
        result.verified_at = datetime.utcnow()
        result.mentor_comment = comment or None
        db.session.commit()
        return jsonify({'success': True, 'message': f"Result {result.status}", 'data': {'id': result.id, 'status': result.status}}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/results/student/<string:admission_number>', methods=['GET'])
def api_student_results(admission_number):
    try:
        sid = str(admission_number or '').strip().upper()
        role = _current_role()
        user_id = str(session.get('user_id') or '').strip().upper()
        if role == 'mentor':
            mentor = Faculty.query.get(session.get('user_id'))
            if not mentor:
                return jsonify({'success': False, 'message': 'Access denied'}), 403
            student = Student.query.get(sid)
            if not student or student.mentor_id != mentor.id:
                return jsonify({'success': False, 'message': 'Access denied'}), 403
            rows = UniversityResult.query.filter_by(student_id=sid, status='verified').order_by(UniversityResult.semester.asc(), UniversityResult.created_at.asc()).all()
        else:
            if role in ('subject-handler', 'hod', 'admin'):
                return jsonify({'success': False, 'message': 'Access denied'}), 403
            if role == 'student' and user_id and user_id != sid:
                return jsonify({'success': False, 'message': 'Access denied'}), 403
            rows = UniversityResult.query.filter_by(student_id=sid).order_by(UniversityResult.semester.asc(), UniversityResult.created_at.asc()).all()

        return jsonify({'success': True, 'data': {'student_id': sid, 'semesters': _group_results_by_semester(rows)}}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':


    with app.app_context():
        # Create tables if they don't exist
        db.create_all()
    
    # Initialize APScheduler for daily schedule regeneration
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from services.schedule_service import regenerate_all_schedules
        import atexit
        
        scheduler = BackgroundScheduler()
        
        # Run daily at 6:00 AM to regenerate schedules
        scheduler.add_job(
            func=regenerate_all_schedules,
            trigger='cron',
            hour=6,
            minute=0,
            id='daily_schedule_regeneration',
            replace_existing=True
        )
        
        scheduler.start()
        
        # Shutdown scheduler on app exit
        atexit.register(lambda: scheduler.shutdown())
        
        print("APScheduler started - daily schedule regeneration at 6:00 AM")
    except Exception as e:
        print(f"Warning: APScheduler initialization failed: {e}")
    
    app.run(debug=True, port=5000)
